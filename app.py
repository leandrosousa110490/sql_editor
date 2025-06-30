#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
SQL Editor Application

A modern SQL editor with support for multiple database types including SQLite and DuckDB.
Features include syntax highlighting, query execution, result visualization, and schema browsing.
"""

import sys
import os
import sqlite3
import duckdb
import pandas as pd
from datetime import datetime
import json
import re
import warnings
import time
import glob

# Suppress pandas warnings about DuckDB connections
warnings.filterwarnings('ignore', message='pandas only supports SQLAlchemy connectable.*')

# Additional imports for export functionality
try:
    import openpyxl
except ImportError:
    openpyxl = None

# Import for CSV delimiter detection
try:
    import csv
except ImportError:
    csv = None

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, QHBoxLayout,
    QSplitter, QTreeWidget, QTreeWidgetItem, QTextEdit, QTableView, QHeaderView,
    QLabel, QPushButton, QComboBox, QFileDialog, QMessageBox, QDialog, QLineEdit,
    QFormLayout, QDialogButtonBox, QToolBar, QStatusBar, QMenu, QInputDialog,
    QSizePolicy, QFrame, QToolButton, QGroupBox, QRadioButton, QCheckBox, QListWidget,
    QCompleter, QListWidgetItem, QProgressDialog, QGridLayout, QScrollArea, QProgressBar
)
from PyQt6.QtGui import (
    QAction, QFont, QColor, QSyntaxHighlighter, QTextCharFormat, QIcon,
    QTextCursor, QPalette, QKeySequence, QShortcut, QStandardItemModel, QStandardItem
)
from PyQt6.QtCore import (
    Qt, QAbstractTableModel, QModelIndex, QSize, QThread, pyqtSignal,
    QRegularExpression, QSettings, QTimer, QStringListModel, pyqtSlot
)
import qtawesome as qta

# Bulk Excel import will be loaded dynamically when needed

# CSV merger functionality removed - no longer importing csv_merger

# Import CSV automation functionality
try:
    from csv_automation import show_csv_automation_dialog
    CSV_AUTOMATION_AVAILABLE = True
except ImportError:
    CSV_AUTOMATION_AVAILABLE = False

# Import Polars for ultra-fast Excel processing
try:
    import polars as pl
    POLARS_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False

# Set application style
QApplication.setStyle('Fusion')

# Define color scheme
class ColorScheme:
    # Dark theme colors
    BACKGROUND = QColor(45, 45, 45)
    SIDEBAR_BG = QColor(35, 35, 35)
    TEXT = QColor(240, 240, 240)
    HIGHLIGHT = QColor(58, 150, 221)
    ACCENT = QColor(75, 160, 240)
    SUCCESS = QColor(95, 200, 115)
    WARNING = QColor(255, 170, 0)
    ERROR = QColor(255, 85, 85)
    
    # Enhanced SQL syntax highlighting colors
    COMMENT = QColor(128, 128, 128)          # Gray for comments
    STRING = QColor(152, 195, 121)           # Green for strings
    KEYWORD = QColor(198, 120, 221)          # Purple for SQL keywords
    NUMBER = QColor(229, 192, 123)           # Orange for numbers
    FUNCTION = QColor(97, 175, 239)          # Blue for functions
    OPERATOR = QColor(86, 182, 194)          # Cyan for operators
    TABLE_NAME = QColor(224, 108, 117)       # Red for table names
    COLUMN_NAME = QColor(152, 195, 121)      # Light green for column names
    DATA_TYPE = QColor(209, 154, 102)        # Tan for data types

# SQL Auto-completion and suggestions
class SQLCompleter(QCompleter):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_completions()
        self.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        self.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.setWrapAround(False)
        
        # Customize the popup appearance
        popup = self.popup()
        popup.setStyleSheet("""
            QListView {
                background-color: #2d2d2d;
                color: #f0f0f0;
                border: 1px solid #4a4a4a;
                selection-background-color: #3a7bd5;
                selection-color: white;
                outline: none;
            }
            QListView::item {
                padding: 5px;
                border-bottom: 1px solid #3a3a3a;
            }
            QListView::item:hover {
                background-color: #3a3a3a;
            }
            QListView::item:selected {
                background-color: #3a7bd5;
            }
        """)
        
    def setup_completions(self):
        # SQL Keywords and commands
        self.sql_keywords = [
            # Core SQL commands
            'SELECT', 'FROM', 'WHERE', 'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'ALTER', 'DROP',
            'TRUNCATE', 'BEGIN', 'COMMIT', 'ROLLBACK', 'SAVEPOINT', 'RELEASE',
            
            # Clauses and modifiers
            'DISTINCT', 'ALL', 'AS', 'INTO', 'VALUES', 'SET', 'ON', 'USING',
            'GROUP BY', 'ORDER BY', 'HAVING', 'LIMIT', 'OFFSET', 'FETCH',
            'UNION', 'UNION ALL', 'INTERSECT', 'EXCEPT', 'MINUS',
            
            # Joins
            'JOIN', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN',
            'CROSS JOIN', 'NATURAL JOIN', 'LEFT OUTER JOIN', 'RIGHT OUTER JOIN',
            'FULL OUTER JOIN',
            
            # Conditional logic
            'CASE', 'WHEN', 'THEN', 'ELSE', 'END', 'IF', 'IFNULL', 'NULLIF',
            
            # Operators
            'AND', 'OR', 'NOT', 'IN', 'NOT IN', 'LIKE', 'NOT LIKE', 'ILIKE',
            'BETWEEN', 'NOT BETWEEN', 'EXISTS', 'NOT EXISTS', 'IS', 'IS NOT',
            'IS NULL', 'IS NOT NULL', 'REGEXP', 'RLIKE',
            
            # Window functions
            'OVER', 'PARTITION BY', 'WINDOW', 'ROWS', 'RANGE', 'PRECEDING',
            'FOLLOWING', 'UNBOUNDED', 'CURRENT ROW',
            
            # CTEs and subqueries
            'WITH', 'RECURSIVE', 'LATERAL',
            
            # Data types
            'INTEGER', 'INT', 'BIGINT', 'SMALLINT', 'TINYINT', 'SERIAL',
            'VARCHAR', 'CHAR', 'TEXT', 'STRING', 'CLOB', 'NVARCHAR', 'NCHAR',
            'REAL', 'FLOAT', 'DOUBLE', 'NUMERIC', 'DECIMAL', 'MONEY',
            'DATE', 'TIME', 'TIMESTAMP', 'DATETIME', 'YEAR', 'INTERVAL',
            'BOOLEAN', 'BOOL', 'BIT', 'BLOB', 'BINARY', 'VARBINARY',
            'JSON', 'JSONB', 'XML', 'UUID', 'ARRAY',
            
            # Constraints and table operations
            'PRIMARY KEY', 'FOREIGN KEY', 'UNIQUE', 'NOT NULL', 'DEFAULT',
            'CHECK', 'REFERENCES', 'CASCADE', 'RESTRICT', 'SET NULL', 'SET DEFAULT',
            'AUTO_INCREMENT', 'IDENTITY', 'GENERATED', 'ALWAYS', 'BY DEFAULT',
            
            # Database objects
            'TABLE', 'VIEW', 'INDEX', 'TRIGGER', 'PROCEDURE', 'FUNCTION',
            'SCHEMA', 'DATABASE', 'CATALOG', 'SEQUENCE', 'DOMAIN', 'TYPE',
            
            # Permissions and security
            'GRANT', 'REVOKE', 'DENY', 'ROLE', 'USER', 'LOGIN', 'PASSWORD',
            'PRIVILEGES', 'USAGE', 'EXECUTE', 'REFERENCES', 'TEMPORARY',
            
            # Optimization and hints
            'EXPLAIN', 'ANALYZE', 'VERBOSE', 'COSTS', 'BUFFERS', 'TIMING',
            'PLAN', 'EXECUTION', 'STATISTICS', 'HINT', 'USE INDEX', 'FORCE INDEX'
        ]
        
        # SQL Functions
        self.sql_functions = [
            # Aggregate functions
            'COUNT', 'SUM', 'AVG', 'MIN', 'MAX', 'STDDEV', 'VARIANCE',
            'GROUP_CONCAT', 'STRING_AGG', 'ARRAY_AGG', 'JSON_AGG',
            
            # String functions
            'CONCAT', 'SUBSTRING', 'SUBSTR', 'LENGTH', 'CHAR_LENGTH',
            'UPPER', 'LOWER', 'TRIM', 'LTRIM', 'RTRIM', 'REPLACE',
            'REGEXP_REPLACE', 'SPLIT_PART', 'POSITION', 'INSTR',
            'LEFT', 'RIGHT', 'REVERSE', 'REPEAT', 'LPAD', 'RPAD',
            'ASCII', 'CHR', 'INITCAP', 'TRANSLATE', 'SOUNDEX',
            
            # Mathematical functions
            'ABS', 'CEIL', 'CEILING', 'FLOOR', 'ROUND', 'TRUNC', 'TRUNCATE',
            'POWER', 'POW', 'SQRT', 'EXP', 'LN', 'LOG', 'LOG10',
            'SIN', 'COS', 'TAN', 'ASIN', 'ACOS', 'ATAN', 'ATAN2',
            'DEGREES', 'RADIANS', 'PI', 'RAND', 'RANDOM', 'SIGN',
            'MOD', 'GREATEST', 'LEAST',
            
            # Date/Time functions
            'NOW', 'CURRENT_DATE', 'CURRENT_TIME', 'CURRENT_TIMESTAMP',
            'TODAY', 'YESTERDAY', 'TOMORROW', 'DATE', 'TIME', 'DATETIME',
            'EXTRACT', 'DATE_PART', 'DATE_TRUNC', 'DATE_ADD', 'DATE_SUB',
            'DATEDIFF', 'DATEADD', 'YEAR', 'MONTH', 'DAY', 'HOUR',
            'MINUTE', 'SECOND', 'MICROSECOND', 'DAYOFWEEK', 'DAYOFYEAR',
            'WEEK', 'WEEKDAY', 'QUARTER', 'LAST_DAY', 'NEXT_DAY',
            'AGE', 'TIMEZONE', 'TO_TIMESTAMP', 'FROM_UNIXTIME',
            
            # Type conversion functions
            'CAST', 'CONVERT', 'TRY_CAST', 'TRY_CONVERT', 'SAFE_CAST',
            'TO_CHAR', 'TO_NUMBER', 'TO_DATE', 'PARSE_DATE', 'PARSE_DATETIME',
            
            # Conditional functions
            'COALESCE', 'ISNULL', 'IFNULL', 'NULLIF', 'NVL', 'NVL2',
            'DECODE', 'CHOOSE', 'IIF',
            
            # Window functions
            'ROW_NUMBER', 'RANK', 'DENSE_RANK', 'PERCENT_RANK', 'CUME_DIST',
            'NTILE', 'LAG', 'LEAD', 'FIRST_VALUE', 'LAST_VALUE', 'NTH_VALUE',
            
            # JSON functions
            'JSON_EXTRACT', 'JSON_UNQUOTE', 'JSON_ARRAY', 'JSON_OBJECT',
            'JSON_KEYS', 'JSON_LENGTH', 'JSON_VALID', 'JSON_TYPE',
            
            # Array functions
            'ARRAY_LENGTH', 'ARRAY_POSITION', 'ARRAY_REMOVE', 'ARRAY_REPLACE',
            'ARRAY_APPEND', 'ARRAY_PREPEND', 'ARRAY_CONTAINS', 'UNNEST',
            
            # System functions
            'VERSION', 'USER', 'CURRENT_USER', 'SESSION_USER', 'SYSTEM_USER',
            'DATABASE', 'SCHEMA', 'CONNECTION_ID', 'LAST_INSERT_ID',
            'ROW_COUNT', 'FOUND_ROWS'
        ]
        
        # Combine all completions
        all_completions = self.sql_keywords + self.sql_functions
        
        # Create model
        self.model = QStringListModel(all_completions)
        self.setModel(self.model)
        
        # Store for dynamic updates
        self.table_names = []
        self.column_names = []
        
    def update_table_names(self, table_names):
        """Update the list of available table names"""
        self.table_names = table_names
        self.refresh_completions()
        
    def update_column_names(self, column_names):
        """Update the list of available column names"""
        self.column_names = column_names
        self.refresh_completions()
        
    def refresh_completions(self):
        """Refresh the completion model with current keywords, functions, tables, and columns"""
        all_completions = (self.sql_keywords + self.sql_functions + 
                         self.table_names + self.column_names)
        
        # Remove duplicates while preserving order
        seen = set()
        unique_completions = []
        for item in all_completions:
            if item.upper() not in seen:
                unique_completions.append(item)
                seen.add(item.upper())
                
        self.model.setStringList(unique_completions)

# Enhanced SQL Text Editor with auto-completion
class SQLTextEdit(QTextEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.completer = SQLCompleter(self)
        self.completer.setWidget(self)
        self.completer.activated.connect(self.insert_completion)
        
        # Set up the text editor
        self.setFont(QFont("Consolas", 11))
        self.setLineWrapMode(QTextEdit.LineWrapMode.NoWrap)
        self.setTabStopDistance(40)  # 4 spaces for tab
        
        # Connect to text change for auto-completion
        self.textChanged.connect(self.on_text_changed)
        
        # Track cursor position for better completion
        self.cursorPositionChanged.connect(self.on_cursor_changed)
        
    def insert_completion(self, completion):
        """Insert the selected completion into the text"""
        cursor = self.textCursor()
        
        # Find the start of the current word
        current_word = self.get_current_word()
        if current_word:
            # Replace the current word with the completion
            cursor.movePosition(QTextCursor.MoveOperation.Left, 
                              QTextCursor.MoveMode.KeepAnchor, len(current_word))
            cursor.removeSelectedText()
            
        cursor.insertText(completion)
        self.setTextCursor(cursor)
        
    def get_current_word(self):
        """Get the word currently being typed"""
        cursor = self.textCursor()
        cursor.select(QTextCursor.SelectionType.WordUnderCursor)
        return cursor.selectedText()
        
    def on_text_changed(self):
        """Handle text changes for auto-completion"""
        current_word = self.get_current_word()
        
        # Only show completions if we have at least 2 characters
        if len(current_word) >= 2:
            self.completer.setCompletionPrefix(current_word)
            
            # Position the completion popup
            cursor_rect = self.cursorRect()
            cursor_rect.setWidth(self.completer.popup().sizeHintForColumn(0) + 
                               self.completer.popup().verticalScrollBar().sizeHint().width())
            self.completer.complete(cursor_rect)
        else:
            self.completer.popup().hide()
            
    def on_cursor_changed(self):
        """Handle cursor position changes"""
        # Hide completion popup if cursor moves away from the word being completed
        if not self.get_current_word():
            self.completer.popup().hide()
            
    def keyPressEvent(self, event):
        """Handle key press events"""
        # Handle special keys for completion
        if self.completer.popup().isVisible():
            if event.key() in (Qt.Key.Key_Enter, Qt.Key.Key_Return, Qt.Key.Key_Tab):
                event.ignore()
                return
            elif event.key() == Qt.Key.Key_Escape:
                self.completer.popup().hide()
                return
                
        # Auto-indentation for new lines
        if event.key() in (Qt.Key.Key_Enter, Qt.Key.Key_Return):
            cursor = self.textCursor()
            cursor.insertText('\n')
            
            # Get the indentation of the current line
            cursor.movePosition(QTextCursor.MoveOperation.Up)
            cursor.movePosition(QTextCursor.MoveOperation.StartOfLine)
            cursor.movePosition(QTextCursor.MoveOperation.EndOfLine, 
                              QTextCursor.MoveMode.KeepAnchor)
            line_text = cursor.selectedText()
            
            # Calculate indentation
            indent = ''
            for char in line_text:
                if char in ' \t':
                    indent += char
                else:
                    break
                    
            # Insert the same indentation on the new line
            cursor.movePosition(QTextCursor.MoveOperation.Down)
            cursor.insertText(indent)
            
            self.setTextCursor(cursor)
            return
            
        # Auto-completion for parentheses, quotes, etc.
        if event.text() == '(':
            cursor = self.textCursor()
            cursor.insertText('()')
            cursor.movePosition(QTextCursor.MoveOperation.Left)
            self.setTextCursor(cursor)
            return
        elif event.text() == '[':
            cursor = self.textCursor()
            cursor.insertText('[]')
            cursor.movePosition(QTextCursor.MoveOperation.Left)
            self.setTextCursor(cursor)
            return
        elif event.text() in ['"', "'"]:
            cursor = self.textCursor()
            quote = event.text()
            cursor.insertText(quote + quote)
            cursor.movePosition(QTextCursor.MoveOperation.Left)
            self.setTextCursor(cursor)
            return
            
        super().keyPressEvent(event)
        
    def update_completions(self, table_names=None, column_names=None):
        """Update the completer with new table and column names"""
        if table_names is not None:
            self.completer.update_table_names(table_names)
        if column_names is not None:
            self.completer.update_column_names(column_names)

# SQL Syntax Highlighter
class SQLHighlighter(QSyntaxHighlighter):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.highlighting_rules = []
        self.setup_highlighting_rules()

    def setup_highlighting_rules(self):
        # Clear existing rules
        self.highlighting_rules = []

        # SQL Keywords (Primary commands) - Most important, put first
        keyword_format = QTextCharFormat()
        keyword_format.setForeground(QColor(198, 120, 221))  # Purple - explicit color
        keyword_format.setFontWeight(QFont.Weight.Bold)
        
        sql_keywords = [
            'SELECT', 'FROM', 'WHERE', 'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'ALTER', 'DROP',
            'TRUNCATE', 'BEGIN', 'COMMIT', 'ROLLBACK', 'DISTINCT', 'AS', 'INTO', 'VALUES', 'SET',
            'GROUP BY', 'ORDER BY', 'HAVING', 'LIMIT', 'OFFSET', 'UNION', 'UNION ALL',
            'INTERSECT', 'EXCEPT', 'WITH', 'RECURSIVE', 'JOIN', 'INNER JOIN', 'LEFT JOIN',
            'RIGHT JOIN', 'FULL JOIN', 'FULL OUTER JOIN', 'CROSS JOIN', 'ON', 'USING',
            'CASE', 'WHEN', 'THEN', 'ELSE', 'END', 'IF', 'IFNULL', 'ISNULL',
            'PRIMARY KEY', 'FOREIGN KEY', 'REFERENCES', 'CONSTRAINT', 'UNIQUE', 'INDEX',
            'VIEW', 'TRIGGER', 'PROCEDURE', 'FUNCTION', 'DECLARE', 'RETURN'
        ]
        
        for keyword in sql_keywords:
            # Use word boundaries to match whole words only
            pattern = f"\\b{keyword}\\b"
            regex = QRegularExpression(pattern, QRegularExpression.CaseInsensitiveOption)
            self.highlighting_rules.append((regex, keyword_format))

        # SQL Operators and Logic
        operator_format = QTextCharFormat()
        operator_format.setForeground(QColor(86, 182, 194))  # Cyan - explicit color
        operator_format.setFontWeight(QFont.Weight.Bold)
        
        operators = [
            'AND', 'OR', 'NOT', 'IN', 'LIKE', 'BETWEEN', 'EXISTS', 'IS', 'NULL',
            'IS NULL', 'IS NOT NULL', 'ALL', 'ANY', 'SOME'
        ]
        
        for operator in operators:
            pattern = f"\\b{operator}\\b"
            regex = QRegularExpression(pattern, QRegularExpression.CaseInsensitiveOption)
            self.highlighting_rules.append((regex, operator_format))

        # Operator symbols
        operator_symbols = ['=', '!=', '<>', '<', '>', '<=', '>=', '\\+', '-', '\\*', '/', '%']
        for symbol in operator_symbols:
            regex = QRegularExpression(symbol)
            self.highlighting_rules.append((regex, operator_format))

        # SQL Functions
        function_format = QTextCharFormat()
        function_format.setForeground(QColor(97, 175, 239))  # Blue - explicit color
        function_format.setFontWeight(QFont.Weight.Bold)
        
        functions = [
            'COUNT', 'SUM', 'AVG', 'MAX', 'MIN', 'GROUP_CONCAT', 'COALESCE', 'NULLIF',
            'CAST', 'CONVERT', 'SUBSTRING', 'SUBSTR', 'LENGTH', 'UPPER', 'LOWER',
            'TRIM', 'LTRIM', 'RTRIM', 'REPLACE', 'NOW', 'CURRENT_DATE', 'CURRENT_TIME',
            'ROUND', 'FLOOR', 'CEIL', 'ABS', 'SQRT', 'POWER', 'MOD', 'RANDOM',
            'CONCAT', 'CONCAT_WS', 'LEFT', 'RIGHT', 'REVERSE', 'REPEAT', 'SPACE',
            'STRFTIME', 'DATE', 'TIME', 'DATETIME', 'JULIANDAY', 'UNIXEPOCH',
            'ROW_NUMBER', 'RANK', 'DENSE_RANK', 'LAG', 'LEAD', 'FIRST_VALUE', 'LAST_VALUE',
            'NTILE', 'PERCENT_RANK', 'CUME_DIST', 'OVER', 'PARTITION BY',
            'JSON_EXTRACT', 'JSON_ARRAY', 'JSON_OBJECT', 'JSON_VALID'
        ]
        
        for function in functions:
            pattern = f"\\b{function}\\b"
            regex = QRegularExpression(pattern, QRegularExpression.CaseInsensitiveOption)
            self.highlighting_rules.append((regex, function_format))

        # JOIN keywords
        join_format = QTextCharFormat()
        join_format.setForeground(QColor(75, 160, 240))  # Accent blue - explicit color
        join_format.setFontWeight(QFont.Weight.Bold)
        
        joins = ['JOIN', 'INNER JOIN', 'LEFT JOIN', 'RIGHT JOIN', 'FULL JOIN', 'CROSS JOIN', 'ON', 'USING']
        for join in joins:
            pattern = f"\\b{join}\\b"
            regex = QRegularExpression(pattern, QRegularExpression.CaseInsensitiveOption)
            self.highlighting_rules.append((regex, join_format))

        # Data Types
        datatype_format = QTextCharFormat()
        datatype_format.setForeground(QColor(209, 154, 102))  # Tan - explicit color
        datatype_format.setFontWeight(QFont.Weight.Bold)
        
        datatypes = [
            'INTEGER', 'INT', 'BIGINT', 'SMALLINT', 'TINYINT', 'VARCHAR', 'CHAR', 'TEXT',
            'REAL', 'FLOAT', 'DOUBLE', 'NUMERIC', 'DECIMAL', 'DATE', 'TIME', 'TIMESTAMP',
            'DATETIME', 'BOOLEAN', 'BOOL', 'BLOB', 'BINARY'
        ]
        
        for datatype in datatypes:
            pattern = f"\\b{datatype}\\b"
            regex = QRegularExpression(pattern, QRegularExpression.CaseInsensitiveOption)
            self.highlighting_rules.append((regex, datatype_format))

        # String literals (quoted strings)
        string_format = QTextCharFormat()
        string_format.setForeground(QColor(152, 195, 121))  # Green - explicit color
        string_format.setFontItalic(True)
        
        # Single quoted strings
        self.highlighting_rules.append((QRegularExpression("'[^']*'"), string_format))
        # Double quoted strings
        self.highlighting_rules.append((QRegularExpression('"[^"]*"'), string_format))

        # Numbers
        number_format = QTextCharFormat()
        number_format.setForeground(QColor(229, 192, 123))  # Orange - explicit color
        number_format.setFontWeight(QFont.Weight.Bold)
        
        # Decimal numbers
        self.highlighting_rules.append((QRegularExpression("\\b\\d+\\.\\d+\\b"), number_format))
        # Integer numbers
        self.highlighting_rules.append((QRegularExpression("\\b\\d+\\b"), number_format))

        # Comments
        comment_format = QTextCharFormat()
        comment_format.setForeground(QColor(128, 128, 128))  # Gray - explicit color
        comment_format.setFontItalic(True)
        
        # Single line comments
        self.highlighting_rules.append((QRegularExpression("--[^\n]*"), comment_format))
        
        # Multi-line comment setup
        self.multiline_comment_format = QTextCharFormat()
        self.multiline_comment_format.setForeground(QColor(128, 128, 128))  # Gray - explicit color
        self.multiline_comment_format.setFontItalic(True)
        self.comment_start_expression = QRegularExpression("/\\*")
        self.comment_end_expression = QRegularExpression("\\*/")

    def highlightBlock(self, text):
        # Debug: Print what we're highlighting (comment out in production)
        # print(f"Highlighting text: '{text}'")
        
        # Apply single-line highlighting rules
        for pattern, format_obj in self.highlighting_rules:
            match_iterator = pattern.globalMatch(text)
            while match_iterator.hasNext():
                match = match_iterator.next()
                start = match.capturedStart()
                length = match.capturedLength()
                matched_text = text[start:start+length]
                # Debug: Print matches (comment out in production)
                # print(f"  Matched: '{matched_text}' at {start}-{start+length}")
                self.setFormat(start, length, format_obj)

        # Handle multi-line comments
        self.setCurrentBlockState(0)
        start_index = 0
        if self.previousBlockState() != 1:
            match = self.comment_start_expression.match(text)
            start_index = match.capturedStart() if match.hasMatch() else -1

        while start_index >= 0:
            match_end = self.comment_end_expression.match(text, start_index)
            end_index = match_end.capturedStart() if match_end.hasMatch() else -1
            comment_length = 0
            
            if end_index == -1:
                self.setCurrentBlockState(1)
                comment_length = len(text) - start_index
            else:
                comment_length = end_index - start_index + match_end.capturedLength()
            
            self.setFormat(start_index, comment_length, self.multiline_comment_format)
            
            if end_index != -1:
                next_match = self.comment_start_expression.match(text, start_index + comment_length)
                start_index = next_match.capturedStart() if next_match.hasMatch() else -1
            else:
                start_index = -1

# Table model for displaying query results with lazy loading for massive datasets
class LazyLoadTableModel(QAbstractTableModel):
    """High-performance table model that lazy loads data for billion+ row datasets"""
    
    def __init__(self, connection, query, chunk_size=None):
        super().__init__()
        self.connection = connection
        self.query = query
        
        # Load settings
        settings = QSettings('SQLEditor', 'QuerySettings')
        self.chunk_size = chunk_size or settings.value('chunk_size', 1000, type=int)
        self.max_cache_size = settings.value('max_cache_chunks', 50, type=int)
        
        # Cache for loaded chunks: {start_row: DataFrame}
        self.data_cache = {}
        
        # Metadata
        self.total_rows = 0
        self.columns = []
        self.column_count = 0
        
        # Initialize metadata
        self._initialize_metadata()
    
    def _initialize_metadata(self):
        """Initialize table metadata without loading all data"""
        try:
            # Get total row count efficiently
            count_query = f"SELECT COUNT(*) FROM ({self.query}) AS count_subquery"
            
            if isinstance(self.connection, sqlite3.Connection):
                cursor = self.connection.cursor()
                cursor.execute(count_query)
                self.total_rows = cursor.fetchone()[0]
                
                # Get column info from first row
                cursor.execute(f"SELECT * FROM ({self.query}) AS sample_subquery LIMIT 1")
                sample_row = cursor.fetchone()
                if sample_row:
                    self.columns = [desc[0] for desc in cursor.description]
                    self.column_count = len(self.columns)
                    
            elif isinstance(self.connection, duckdb.DuckDBPyConnection):
                result = self.connection.execute(count_query).fetchone()
                self.total_rows = result[0]
                
                # Get column info
                sample_result = self.connection.execute(f"SELECT * FROM ({self.query}) AS sample_subquery LIMIT 1")
                if sample_result:
                    self.columns = [desc[0] for desc in sample_result.description]
                    self.column_count = len(self.columns)
                    
        except Exception as e:
            print(f"Error initializing lazy table metadata: {e}")
            self.total_rows = 0
            self.columns = []
            self.column_count = 0
    
    def _get_chunk_start(self, row):
        """Get the starting row of the chunk containing the given row"""
        return (row // self.chunk_size) * self.chunk_size
    
    def _load_chunk(self, start_row):
        """Load a chunk of data starting from start_row"""
        try:
            # Check if already cached
            if start_row in self.data_cache:
                return self.data_cache[start_row]
            
            # Clean cache if too large
            if len(self.data_cache) >= self.max_cache_size:
                # Remove oldest entries (simple FIFO)
                oldest_keys = list(self.data_cache.keys())[:10]
                for key in oldest_keys:
                    del self.data_cache[key]
            
            # Load chunk from database
            limit_query = f"SELECT * FROM ({self.query}) AS chunked_subquery LIMIT {self.chunk_size} OFFSET {start_row}"
            
            if isinstance(self.connection, sqlite3.Connection):
                df = pd.read_sql_query(limit_query, self.connection)
            elif isinstance(self.connection, duckdb.DuckDBPyConnection):
                df = self.connection.execute(limit_query).df()
            else:
                df = pd.DataFrame()  # Fallback
            
            # Cache the chunk
            self.data_cache[start_row] = df
            return df
            
        except Exception as e:
            print(f"Error loading chunk at row {start_row}: {e}")
            return pd.DataFrame()
    
    def rowCount(self, parent=QModelIndex()):
        return self.total_rows
    
    def columnCount(self, parent=QModelIndex()):
        return self.column_count
    
    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        
        row = index.row()
        col = index.column()
        
        if row >= self.total_rows or col >= self.column_count:
            return None
        
        if role == Qt.ItemDataRole.DisplayRole:
            # Determine which chunk this row belongs to
            chunk_start = self._get_chunk_start(row)
            chunk_df = self._load_chunk(chunk_start)
            
            if chunk_df.empty:
                return "Loading..."
            
            # Get the row within the chunk
            chunk_row = row - chunk_start
            if chunk_row >= len(chunk_df):
                return "Loading..."
            
            try:
                value = chunk_df.iloc[chunk_row, col]
                if pd.isna(value):
                    return "NULL"
                elif isinstance(value, (float, int)):
                    return str(value)
                else:
                    return str(value)
            except (IndexError, KeyError):
                return "Loading..."
                
        elif role == Qt.ItemDataRole.TextAlignmentRole:
            # Try to get value for alignment
            chunk_start = self._get_chunk_start(row)
            chunk_df = self._load_chunk(chunk_start)
            
            if not chunk_df.empty:
                chunk_row = row - chunk_start
                if chunk_row < len(chunk_df):
                    try:
                        value = chunk_df.iloc[chunk_row, col]
                        if isinstance(value, (int, float)):
                            return int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    except (IndexError, KeyError):
                        pass
            
            return int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        
        return None
    
    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                if section < len(self.columns):
                    return str(self.columns[section])
                return f"Column {section + 1}"
            else:
                return str(section + 1)
        return None
    
    def sort(self, column, order):
        """Sorting requires recreating the model with ORDER BY clause"""
        if column >= self.column_count:
            return
        
        column_name = self.columns[column]
        order_direction = "ASC" if order == Qt.SortOrder.AscendingOrder else "DESC"
        
        # Add ORDER BY to the original query
        # This is a simplified approach - in production, you'd want more sophisticated query parsing
        if "ORDER BY" in self.query.upper():
            # Replace existing ORDER BY (simple approach)
            import re
            self.query = re.sub(r'ORDER BY.*$', f'ORDER BY "{column_name}" {order_direction}', self.query, flags=re.IGNORECASE)
        else:
            self.query += f' ORDER BY "{column_name}" {order_direction}'
        
        # Clear cache and reinitialize
        self.data_cache.clear()
        self._initialize_metadata()
        
        # Notify view that everything changed
        self.layoutAboutToBeChanged.emit()
        self.layoutChanged.emit()
    
    def canFetchMore(self, parent):
        """Enable fetch more for better lazy loading"""
        return True
    
    def fetchMore(self, parent):
        """Preload next chunks when scrolling"""
        # This is called by the view when it needs more data
        # We can use this to preload upcoming chunks
        pass

# Table model for displaying regular query results (backward compatibility)
class PandasTableModel(QAbstractTableModel):
    def __init__(self, data=None):
        super().__init__()
        self._data = pd.DataFrame() if data is None else data

    def rowCount(self, parent=QModelIndex()):
        return len(self._data)

    def columnCount(self, parent=QModelIndex()):
        return len(self._data.columns)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
            
        if role == Qt.ItemDataRole.DisplayRole:
            value = self._data.iloc[index.row(), index.column()]
            if pd.isna(value):
                return "NULL"
            elif isinstance(value, (float, int)):
                return str(value)
            else:
                return str(value)
                
        elif role == Qt.ItemDataRole.TextAlignmentRole:
            value = self._data.iloc[index.row(), index.column()]
            if isinstance(value, (int, float)):
                return int(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            return int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                
        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return str(self._data.columns[section])
            else:
                return str(section + 1)
        return None

    def sort(self, column, order):
        self.layoutAboutToBeChanged.emit()
        self._data = self._data.sort_values(
            self._data.columns[column],
            ascending=(order == Qt.SortOrder.AscendingOrder)
        )
        self.layoutChanged.emit()

# Enhanced query worker that supports both lazy and regular loading
class EnhancedQueryWorker(QThread):
    # Signals for regular loading
    finished = pyqtSignal(object, float)  # DataFrame, execution_time
    error = pyqtSignal(str)
    
    # Signals for lazy loading
    lazy_finished = pyqtSignal(object, float)  # LazyLoadTableModel, execution_time
    metadata_ready = pyqtSignal(int, list, float)  # total_rows, columns, execution_time
    
    def __init__(self, connection, query, use_lazy_loading=True, row_limit=100000):
        super().__init__()
        self.connection = connection
        self.query = query
        self.use_lazy_loading = use_lazy_loading
        self.row_limit = row_limit  # Threshold for using lazy loading
        
    def run(self):
        try:
            start_time = datetime.now()
            
            # Check if this is a DDL statement that doesn't return results
            query_upper = self.query.strip().upper()
            ddl_keywords = ['ALTER TABLE', 'CREATE TABLE', 'DROP TABLE', 'CREATE VIEW', 'DROP VIEW', 
                          'CREATE INDEX', 'DROP INDEX', 'INSERT INTO', 'UPDATE ', 'DELETE FROM']
            is_ddl = any(query_upper.startswith(keyword) for keyword in ddl_keywords)
            
            if is_ddl:
                # Execute DDL statement directly without expecting results
                if isinstance(self.connection, sqlite3.Connection):
                    cursor = self.connection.cursor()
                    cursor.execute(self.query)
                    self.connection.commit()
                elif isinstance(self.connection, duckdb.DuckDBPyConnection):
                    self.connection.execute(self.query)
                else:
                    raise ValueError("Unsupported database connection type")
                
                # Return empty DataFrame for DDL operations
                execution_time = (datetime.now() - start_time).total_seconds()
                empty_df = pd.DataFrame({'Result': ['Query executed successfully']})
                self.finished.emit(empty_df, execution_time)
                return
            
            if self.use_lazy_loading:
                # Check if we should use lazy loading by estimating row count
                should_use_lazy = self._should_use_lazy_loading()
                
                if should_use_lazy:
                    # Create lazy loading model
                    lazy_model = LazyLoadTableModel(self.connection, self.query)
                    execution_time = (datetime.now() - start_time).total_seconds()
                    self.lazy_finished.emit(lazy_model, execution_time)
                    return
            
            # Regular loading for smaller results
            if isinstance(self.connection, sqlite3.Connection) or isinstance(self.connection, duckdb.DuckDBPyConnection):
                df = pd.read_sql_query(self.query, self.connection)
            else:
                raise ValueError("Unsupported database connection type")
                
            execution_time = (datetime.now() - start_time).total_seconds()
            self.finished.emit(df, execution_time)
            
        except Exception as e:
            # Enhanced error handling with more specific messages
            error_msg = str(e)
            
            # Provide more helpful error messages for common issues
            if "no such table" in error_msg.lower():
                error_msg += "\n\nTip: Check if the table name is correct and exists in the database."
            elif "no such column" in error_msg.lower():
                error_msg += "\n\nTip: Check if the column name is correct and exists in the table."
            elif "syntax error" in error_msg.lower():
                error_msg += "\n\nTip: Check your SQL syntax. Common issues include missing quotes, incorrect keywords, or typos."
            elif "alter table" in error_msg.lower():
                error_msg += "\n\nTip: ALTER TABLE operations have limitations. For SQLite, some operations require recreating the table."
            elif "permission" in error_msg.lower() or "access" in error_msg.lower():
                error_msg += "\n\nTip: Check if you have the necessary permissions to perform this operation."
            
            self.error.emit(error_msg)
    
    def _should_use_lazy_loading(self):
        """Determine if lazy loading should be used based on estimated result size"""
        try:
            # Quick row count estimation
            count_query = f"SELECT COUNT(*) FROM ({self.query}) AS count_subquery"
            
            if isinstance(self.connection, sqlite3.Connection):
                cursor = self.connection.cursor()
                cursor.execute(count_query)
                estimated_rows = cursor.fetchone()[0]
            elif isinstance(self.connection, duckdb.DuckDBPyConnection):
                result = self.connection.execute(count_query).fetchone()
                estimated_rows = result[0]
            else:
                return False
            
            return estimated_rows > self.row_limit
            
        except Exception as e:
            print(f"Error estimating query size: {e}")
            # If we can't estimate, default to lazy loading for safety
            return True

# Worker thread for executing queries (backward compatibility)
class QueryWorker(QThread):
    finished = pyqtSignal(object, float)
    error = pyqtSignal(str)
    
    def __init__(self, connection, query):
        super().__init__()
        self.connection = connection
        self.query = query
        
    def run(self):
        try:
            start_time = datetime.now()
            if isinstance(self.connection, sqlite3.Connection) or isinstance(self.connection, duckdb.DuckDBPyConnection):
                df = pd.read_sql_query(self.query, self.connection)
            else:
                # Handle other database types if needed
                raise ValueError("Unsupported database connection type")
                
            execution_time = (datetime.now() - start_time).total_seconds()
            self.finished.emit(df, execution_time)
        except Exception as e:
            self.error.emit(str(e))

class ImportWorker(QThread):
    """Worker thread for non-blocking data imports with optimized performance"""
    progress = pyqtSignal(int, str)  # progress percentage, status message
    finished = pyqtSignal(bool, str)  # success, message
    error = pyqtSignal(str)
    
    def __init__(self, main_app, import_info):
        super().__init__()
        self.main_app = main_app
        self.import_info = import_info
        self.cancelled = False
    
    def cancel(self):
        """Cancel the import operation"""
        self.cancelled = True
    
    def run(self):
        try:
            self.progress.emit(5, "Starting optimized file import...")
            success = self.main_app.import_data_optimized(self.import_info, self)
            
            if self.cancelled:
                self.finished.emit(False, "Import was cancelled by user.")
            elif success:
                self.progress.emit(100, "Import completed successfully!")
                self.finished.emit(True, "Data imported successfully!")
            else:
                self.finished.emit(False, "Import failed. Check console for details.")
                
        except Exception as e:
            if not self.cancelled:
                self.error.emit(f"Import error: {str(e)}")

class ProgressDialog(QDialog):
    """Progress dialog for long-running operations"""
    
    def __init__(self, parent=None, title="Processing..."):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setFixedSize(400, 150)
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.CustomizeWindowHint | Qt.WindowType.WindowTitleHint)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Status label
        self.status_label = QLabel("Initializing...")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.status_label)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        
        # Cancel button
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #c82333; }
        """)
        
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_button)
        layout.addLayout(button_layout)
    
    def update_progress(self, value, message):
        """Update progress bar and status message"""
        self.progress_bar.setValue(value)
        self.status_label.setText(message)
        
        # Auto-close when complete
        if value >= 100:
            QTimer.singleShot(1000, self.accept)  # Close after 1 second

# Connection dialog
class ConnectionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Connect to Database")
        self.resize(450, 250)
        
        self.layout = QVBoxLayout()
        
        # Database type selection
        self.db_type_layout = QHBoxLayout()
        self.db_type_label = QLabel("Database Type:")
        self.db_type_combo = QComboBox()
        self.db_type_combo.addItems(["SQLite", "DuckDB"])
        self.db_type_combo.currentIndexChanged.connect(self.update_form)
        self.db_type_layout.addWidget(self.db_type_label)
        self.db_type_layout.addWidget(self.db_type_combo)
        self.layout.addLayout(self.db_type_layout)
        
        # Form layout for connection details
        self.form_layout = QFormLayout()
        self.file_path_edit = QLineEdit()
        self.browse_button = QPushButton("Browse...")
        self.browse_button.clicked.connect(self.browse_file)
        self.create_button = QPushButton("Create New Database")
        self.create_button.clicked.connect(self.create_new_database)
        
        self.file_layout = QHBoxLayout()
        self.file_layout.addWidget(self.file_path_edit)
        self.file_layout.addWidget(self.browse_button)
        self.file_layout.addWidget(self.create_button)
        
        self.form_layout.addRow("Database File:", self.file_layout)
        self.layout.addLayout(self.form_layout)
        
        # Buttons
        self.button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.layout.addWidget(self.button_box)
        
        self.setLayout(self.layout)
    
    def update_form(self):
        # Update form based on selected database type
        db_type = self.db_type_combo.currentText()
        # For now, both SQLite and DuckDB just need a file path
        # This can be expanded for other database types
    
    def browse_file(self):
        db_type = self.db_type_combo.currentText()
        file_filter = "SQLite Database (*.db *.sqlite);;All Files (*)" if db_type == "SQLite" else "DuckDB Database (*.duckdb);;All Files (*)"
        file_path, _ = QFileDialog.getOpenFileName(self, "Select Database File", "", file_filter)
        if file_path:
            self.file_path_edit.setText(file_path)
    
    def create_new_database(self):
        db_type = self.db_type_combo.currentText()
        
        # Get database name from user
        name, ok = QInputDialog.getText(self, "Create New Database", "Enter database name:")
        if not ok or not name.strip():
            return
        
        # Clean the name (remove invalid characters)
        name = re.sub(r'[<>:"/\\|?*]', '', name.strip())
        if not name:
            QMessageBox.warning(self, "Invalid Name", "Please enter a valid database name.")
            return
        
        # Set file extension based on database type
        if db_type == "SQLite":
            extension = ".db"
            file_filter = "SQLite Database (*.db)"
        else:  # DuckDB
            extension = ".duckdb"
            file_filter = "DuckDB Database (*.duckdb)"
        
        # Get save location
        default_name = name + extension
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            f"Create New {db_type} Database", 
            default_name, 
            file_filter
        )
        
        if not file_path:
            return
        
        # Ensure correct extension
        if not file_path.lower().endswith(extension):
            file_path += extension
        
        try:
            # Create the database file
            if db_type == "SQLite":
                # Create SQLite database
                conn = sqlite3.connect(file_path)
                conn.execute("CREATE TABLE IF NOT EXISTS _metadata (created_at TEXT)")
                conn.execute("INSERT INTO _metadata (created_at) VALUES (?)", (datetime.now().isoformat(),))
                conn.commit()
                conn.close()
            else:  # DuckDB
                # Create DuckDB database
                conn = duckdb.connect(file_path)
                conn.execute("CREATE TABLE IF NOT EXISTS _metadata (created_at TEXT)")
                conn.execute("INSERT INTO _metadata (created_at) VALUES (?)", (datetime.now().isoformat(),))
                conn.commit()
                conn.close()
            
            # Set the file path in the dialog
            self.file_path_edit.setText(file_path)
            
            QMessageBox.information(
                self, 
                "Database Created", 
                f"Successfully created {db_type} database:\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Creation Error", 
                f"Failed to create database:\n{str(e)}"
            )
    
    def get_connection_info(self):
        return {
            "type": self.db_type_combo.currentText(),
            "file_path": self.file_path_edit.text()
        }

# Create Database dialog
class CreateDatabaseDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Create New Database")
        self.resize(400, 180)
        self.created_file_path = None
        
        self.layout = QVBoxLayout()
        
        # Database type selection
        self.db_type_layout = QHBoxLayout()
        self.db_type_label = QLabel("Database Type:")
        self.db_type_combo = QComboBox()
        self.db_type_combo.addItems(["SQLite", "DuckDB"])
        self.db_type_layout.addWidget(self.db_type_label)
        self.db_type_layout.addWidget(self.db_type_combo)
        self.layout.addLayout(self.db_type_layout)
        
        # Database name input
        self.name_layout = QHBoxLayout()
        self.name_label = QLabel("Database Name:")
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("Enter database name...")
        self.name_layout.addWidget(self.name_label)
        self.name_layout.addWidget(self.name_edit)
        self.layout.addLayout(self.name_layout)
        
        # Location selection
        self.location_layout = QHBoxLayout()
        self.location_label = QLabel("Save Location:")
        self.location_edit = QLineEdit()
        self.location_edit.setPlaceholderText("Choose location...")
        self.location_edit.setReadOnly(True)
        self.browse_location_button = QPushButton("Browse...")
        self.browse_location_button.clicked.connect(self.browse_location)
        self.location_layout.addWidget(self.location_label)
        self.location_layout.addWidget(self.location_edit)
        self.location_layout.addWidget(self.browse_location_button)
        self.layout.addLayout(self.location_layout)
        
        # Buttons
        self.button_layout = QHBoxLayout()
        self.create_button = QPushButton("Create Database")
        self.create_button.clicked.connect(self.create_database)
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.button_layout.addWidget(self.create_button)
        self.button_layout.addWidget(self.cancel_button)
        self.layout.addLayout(self.button_layout)
        
        self.setLayout(self.layout)
        
        # Set default location to user's Documents folder
        import os
        default_location = os.path.expanduser("~/Documents")
        if os.path.exists(default_location):
            self.location_edit.setText(default_location)
    
    def browse_location(self):
        directory = QFileDialog.getExistingDirectory(self, "Select Directory")
        if directory:
            self.location_edit.setText(directory)
    
    def create_database(self):
        name = self.name_edit.text().strip()
        location = self.location_edit.text().strip()
        db_type = self.db_type_combo.currentText()
        
        if not name:
            QMessageBox.warning(self, "Invalid Input", "Please enter a database name.")
            return
        
        if not location or not os.path.exists(location):
            QMessageBox.warning(self, "Invalid Location", "Please select a valid location.")
            return
        
        # Clean the name (remove invalid characters)
        name = re.sub(r'[<>:"/\\|?*]', '', name)
        if not name:
            QMessageBox.warning(self, "Invalid Name", "Please enter a valid database name.")
            return
        
        # Set file extension based on database type
        extension = ".db" if db_type == "SQLite" else ".duckdb"
        file_path = os.path.join(location, name + extension)
        
        # Check if file already exists
        if os.path.exists(file_path):
            reply = QMessageBox.question(
                self, 
                "File Exists", 
                f"A database with this name already exists:\n{file_path}\n\nDo you want to overwrite it?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return
        
        try:
            # Create the database file
            if db_type == "SQLite":
                conn = sqlite3.connect(file_path)
                conn.execute("CREATE TABLE IF NOT EXISTS _metadata (created_at TEXT, created_by TEXT)")
                conn.execute(
                    "INSERT INTO _metadata (created_at, created_by) VALUES (?, ?)", 
                    (datetime.now().isoformat(), "SQL Editor Application")
                )
                conn.commit()
                conn.close()
            else:  # DuckDB
                conn = duckdb.connect(file_path)
                conn.execute("CREATE TABLE IF NOT EXISTS _metadata (created_at TEXT, created_by TEXT)")
                conn.execute(
                    "INSERT INTO _metadata (created_at, created_by) VALUES (?, ?)", 
                    (datetime.now().isoformat(), "SQL Editor Application")
                )
                conn.commit()
                conn.close()
            
            self.created_file_path = file_path
            
            QMessageBox.information(
                self, 
                "Database Created", 
                f"Successfully created {db_type} database:\n{file_path}\n\nThe database will now be opened."
            )
            
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Creation Error", 
                f"Failed to create database:\n{str(e)}"
            )
    
    def get_connection_info(self):
        if self.created_file_path:
            return {
                "type": self.db_type_combo.currentText(),
                "file_path": self.created_file_path
            }
        return None

# Settings dialog for lazy loading configuration
class LazyLoadingSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Lazy Loading Settings")
        self.setModal(True)
        self.resize(500, 400)
        
        # Load current settings
        self.settings = QSettings('SQLEditor', 'QuerySettings')
        
        layout = QVBoxLayout(self)
        
        # Title
        title = QLabel("Query Result Loading Settings")
        title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # Description
        description = QLabel(
            "Configure how query results are loaded and displayed. "
            "Lazy loading helps handle massive datasets by loading data on-demand."
        )
        description.setWordWrap(True)
        description.setStyleSheet("margin: 10px 0px; color: #888;")
        layout.addWidget(description)
        
        # Settings form
        form_layout = QFormLayout()
        
        # Lazy loading threshold
        self.threshold_spin = QSpinBox()
        self.threshold_spin.setRange(1000, 10000000)
        self.threshold_spin.setValue(self.settings.value('lazy_loading_threshold', 100000, type=int))
        self.threshold_spin.setSuffix(" rows")
        self.threshold_spin.setToolTip("Queries returning more than this many rows will use lazy loading")
        form_layout.addRow("Lazy Loading Threshold:", self.threshold_spin)
        
        # Chunk size
        self.chunk_spin = QSpinBox()
        self.chunk_spin.setRange(100, 10000)
        self.chunk_spin.setValue(self.settings.value('chunk_size', 1000, type=int))
        self.chunk_spin.setSuffix(" rows")
        self.chunk_spin.setToolTip("Number of rows to load at once during lazy loading")
        form_layout.addRow("Chunk Size:", self.chunk_spin)
        
        # Cache size
        self.cache_spin = QSpinBox()
        self.cache_spin.setRange(10, 200)
        self.cache_spin.setValue(self.settings.value('max_cache_chunks', 50, type=int))
        self.cache_spin.setSuffix(" chunks")
        self.cache_spin.setToolTip("Maximum number of data chunks to keep in memory")
        form_layout.addRow("Cache Size:", self.cache_spin)
        
        # Enable/disable lazy loading
        self.enable_lazy = QCheckBox("Enable lazy loading for large datasets")
        self.enable_lazy.setChecked(self.settings.value('enable_lazy_loading', True, type=bool))
        self.enable_lazy.setToolTip("Uncheck to always load all data into memory (not recommended for large datasets)")
        form_layout.addRow("", self.enable_lazy)
        
        layout.addLayout(form_layout)
        
        # Performance info
        info_group = QGroupBox("Performance Information")
        info_layout = QVBoxLayout(info_group)
        
        info_text = QLabel(
            "• <b>Lazy Loading:</b> Loads data on-demand as you scroll. Ideal for datasets with millions/billions of rows.\n"
            "• <b>Regular Loading:</b> Loads all data into memory at once. Faster for small datasets but uses more memory.\n"
            "• <b>Chunk Size:</b> Smaller chunks = lower memory usage but more database queries.\n"
            "• <b>Cache Size:</b> Larger cache = less database queries but more memory usage."
        )
        info_text.setWordWrap(True)
        info_text.setStyleSheet("margin: 10px; color: #666;")
        info_layout.addWidget(info_text)
        
        layout.addWidget(info_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        reset_button = QPushButton("Reset to Defaults")
        reset_button.clicked.connect(self.reset_defaults)
        button_layout.addWidget(reset_button)
        
        button_layout.addStretch()
        
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        save_button = QPushButton("Save Settings")
        save_button.clicked.connect(self.save_settings)
        save_button.setDefault(True)
        button_layout.addWidget(save_button)
        
        layout.addLayout(button_layout)
        
        # Styling
        self.setStyleSheet("""
            QDialog {
                background-color: #2d2d2d;
                color: #f0f0f0;
            }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #555;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
            QPushButton {
                background-color: #4a4a4a;
                border: 1px solid #666;
                border-radius: 3px;
                padding: 8px 16px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #555;
            }
            QPushButton:pressed {
                background-color: #333;
            }
            QPushButton:default {
                background-color: #0078d4;
                border-color: #106ebe;
            }
        """)
    
    def reset_defaults(self):
        """Reset all settings to defaults"""
        self.threshold_spin.setValue(100000)
        self.chunk_spin.setValue(1000)
        self.cache_spin.setValue(50)
        self.enable_lazy.setChecked(True)
    
    def save_settings(self):
        """Save settings and close dialog"""
        self.settings.setValue('lazy_loading_threshold', self.threshold_spin.value())
        self.settings.setValue('chunk_size', self.chunk_spin.value())
        self.settings.setValue('max_cache_chunks', self.cache_spin.value())
        self.settings.setValue('enable_lazy_loading', self.enable_lazy.isChecked())
        self.accept()




class FolderImportDialog(QDialog):
    """Dialog for importing entire folders of CSV or Excel files"""
    
    def __init__(self, parent=None, connection=None, connection_info=None):
        super().__init__(parent)
        self.connection = connection
        self.connection_info = connection_info
        window_title = "Folder Import Tool"
        if POLARS_AVAILABLE:
            window_title += " (⚡ Polars Enhanced)"
        self.setWindowTitle(window_title)
        self.setModal(True)
        self.resize(700, 650)  # Slightly taller for new options
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Title
        title_text = "📁 Import Folder of Files"
        if POLARS_AVAILABLE:
            title_text += " ⚡"
        title = QLabel(title_text)
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet("color: #0078d4; margin-bottom: 15px;")
        layout.addWidget(title)
        
        # File type selection
        type_group = QGroupBox("File Type")
        type_layout = QVBoxLayout()
        
        self.file_type_combo = QComboBox()
        self.file_type_combo.addItems([
            "CSV Files (*.csv)",
            "Excel Files (*.xlsx, *.xls)"
        ])
        self.file_type_combo.currentTextChanged.connect(self.on_file_type_changed)
        type_layout.addWidget(self.file_type_combo)
        
        type_group.setLayout(type_layout)
        layout.addWidget(type_group)
        
        # Folder selection
        folder_group = QGroupBox("Folder Selection")
        folder_layout = QVBoxLayout()
        
        # Folder path
        folder_path_layout = QHBoxLayout()
        self.folder_path = QLineEdit()
        self.folder_path.setPlaceholderText("Select folder containing files...")
        self.folder_path.textChanged.connect(self.on_folder_changed)
        
        self.browse_folder_btn = QPushButton("Browse Folder")
        self.browse_folder_btn.clicked.connect(self.browse_folder)
        self.browse_folder_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #106ebe; }
        """)
        
        folder_path_layout.addWidget(self.folder_path)
        folder_path_layout.addWidget(self.browse_folder_btn)
        folder_layout.addLayout(folder_path_layout)
        
        # File pattern
        pattern_layout = QHBoxLayout()
        pattern_layout.addWidget(QLabel("File Pattern:"))
        self.file_pattern = QLineEdit("*.csv")
        self.file_pattern.setPlaceholderText("e.g., *.csv, data_*.csv, sales_2024_*.xlsx")
        self.file_pattern.textChanged.connect(self.on_pattern_changed)
        pattern_layout.addWidget(self.file_pattern)
        folder_layout.addLayout(pattern_layout)
        
        # Scan button and file list
        self.scan_btn = QPushButton("🔍 Scan Folder")
        self.scan_btn.clicked.connect(self.scan_folder)
        self.scan_btn.setEnabled(False)
        folder_layout.addWidget(self.scan_btn)
        
        # File list
        self.file_list = QListWidget()
        self.file_list.setMaximumHeight(120)
        folder_layout.addWidget(QLabel("Files found:"))
        folder_layout.addWidget(self.file_list)
        
        folder_group.setLayout(folder_layout)
        layout.addWidget(folder_group)
        
        # Import settings
        import_group = QGroupBox("Import Settings")
        import_layout = QVBoxLayout()
        
        # Table naming
        table_layout = QHBoxLayout()
        table_layout.addWidget(QLabel("Table Name Strategy:"))
        
        self.table_strategy = QComboBox()
        self.table_strategy.addItems([
            "Single Table (merge all files)",
            "Multiple Tables (one per file)"
        ])
        self.table_strategy.currentTextChanged.connect(self.on_strategy_changed)
        table_layout.addWidget(self.table_strategy)
        import_layout.addLayout(table_layout)
        
        # Single table name (shown when "Single Table" is selected)
        self.single_table_widget = QWidget()
        single_table_layout = QHBoxLayout()
        single_table_layout.setContentsMargins(0, 0, 0, 0)
        single_table_layout.addWidget(QLabel("Table Name:"))
        self.single_table_name = QLineEdit()
        self.single_table_name.setPlaceholderText("Enter table name for merged data...")
        self.single_table_name.textChanged.connect(self.update_import_button)
        single_table_layout.addWidget(self.single_table_name)
        self.single_table_widget.setLayout(single_table_layout)
        import_layout.addWidget(self.single_table_widget)
        
        # Table prefix (shown when "Multiple Tables" is selected)
        self.multi_table_widget = QWidget()
        multi_table_layout = QHBoxLayout()
        multi_table_layout.setContentsMargins(0, 0, 0, 0)
        multi_table_layout.addWidget(QLabel("Table Prefix:"))
        self.table_prefix = QLineEdit()
        self.table_prefix.setPlaceholderText("Optional prefix for table names...")
        multi_table_layout.addWidget(self.table_prefix)
        self.multi_table_widget.setLayout(multi_table_layout)
        import_layout.addWidget(self.multi_table_widget)
        self.multi_table_widget.hide()
        
        # Import mode
        mode_layout = QHBoxLayout()
        mode_layout.addWidget(QLabel("Import Mode:"))
        self.import_mode = QComboBox()
        self.import_mode.addItems(["Create New", "Replace Existing", "Append to Existing"])
        mode_layout.addWidget(self.import_mode)
        import_layout.addLayout(mode_layout)
        
        import_group.setLayout(import_layout)
        layout.addWidget(import_group)
        
        # Advanced options
        advanced_group = QGroupBox("Advanced Options")
        advanced_layout = QVBoxLayout()
        
        # CSV-specific options
        self.csv_options_widget = QWidget()
        csv_options_layout = QHBoxLayout()
        csv_options_layout.setContentsMargins(0, 0, 0, 0)
        
        csv_options_layout.addWidget(QLabel("Encoding:"))
        self.encoding_combo = QComboBox()
        self.encoding_combo.addItems(["utf-8", "cp1252", "iso-8859-1", "utf-16"])
        csv_options_layout.addWidget(self.encoding_combo)
        
        csv_options_layout.addWidget(QLabel("Delimiter:"))
        self.delimiter_combo = QComboBox()
        self.delimiter_combo.addItems(["Auto-detect", ",", ";", "\\t", "|"])
        csv_options_layout.addWidget(self.delimiter_combo)
        
        self.csv_options_widget.setLayout(csv_options_layout)
        advanced_layout.addWidget(self.csv_options_widget)
        
        # Excel-specific options
        self.excel_options_widget = QWidget()
        excel_options_layout = QVBoxLayout()
        excel_options_layout.setContentsMargins(0, 0, 0, 0)
        
        # Sheet handling row
        sheet_handling_layout = QHBoxLayout()
        sheet_handling_layout.addWidget(QLabel("Sheet Mode:"))
        self.sheet_handling = QComboBox()
        self.sheet_handling.addItems(["First Sheet Only", "All Sheets", "Specific Sheet"])
        self.sheet_handling.currentTextChanged.connect(self.on_sheet_mode_changed)
        sheet_handling_layout.addWidget(self.sheet_handling)
        excel_options_layout.addLayout(sheet_handling_layout)
        
        # Specific sheet selection row (initially hidden)
        self.specific_sheet_widget = QWidget()
        specific_sheet_layout = QHBoxLayout()
        specific_sheet_layout.setContentsMargins(0, 0, 0, 0)
        specific_sheet_layout.addWidget(QLabel("Sheet Name:"))
        self.specific_sheet_combo = QComboBox()
        self.specific_sheet_combo.setMinimumWidth(150)
        specific_sheet_layout.addWidget(self.specific_sheet_combo)
        
        self.scan_sheets_btn = QPushButton("🔍 Scan Sheets")
        self.scan_sheets_btn.clicked.connect(self.scan_excel_sheets)
        self.scan_sheets_btn.setEnabled(False)
        specific_sheet_layout.addWidget(self.scan_sheets_btn)
        
        self.specific_sheet_widget.setLayout(specific_sheet_layout)
        excel_options_layout.addWidget(self.specific_sheet_widget)
        self.specific_sheet_widget.hide()
        
        # Performance info
        performance_label = QLabel("⚡ Using Polars for ultra-fast processing" if POLARS_AVAILABLE else "📊 Using Pandas for processing")
        performance_label.setStyleSheet("font-size: 10px; color: #666; font-style: italic;")
        excel_options_layout.addWidget(performance_label)
        
        self.excel_options_widget.setLayout(excel_options_layout)
        advanced_layout.addWidget(self.excel_options_widget)
        self.excel_options_widget.hide()
        
        advanced_group.setLayout(advanced_layout)
        layout.addWidget(advanced_group)
        
        # Progress bar (initially hidden)
        self.progress_bar = QProgressBar()
        self.progress_bar.hide()
        layout.addWidget(self.progress_bar)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.import_btn = QPushButton("🚀 Import Files")
        self.import_btn.clicked.connect(self.start_import)
        self.import_btn.setEnabled(False)
        self.import_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover { background-color: #218838; }
            QPushButton:disabled { background-color: #cccccc; }
        """)
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.import_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
    def on_file_type_changed(self):
        """Handle file type change"""
        is_csv = "CSV" in self.file_type_combo.currentText()
        
        if is_csv:
            self.file_pattern.setText("*.csv")
            self.csv_options_widget.show()
            self.excel_options_widget.hide()
        else:
            self.file_pattern.setText("*.xlsx")
            self.csv_options_widget.hide()
            self.excel_options_widget.show()
            # Update scan sheets button state
            self.update_scan_sheets_button()
            
        self.scan_folder()
        
    def on_folder_changed(self):
        """Handle folder path change"""
        folder_exists = bool(self.folder_path.text().strip() and 
                           os.path.exists(self.folder_path.text().strip()))
        self.scan_btn.setEnabled(folder_exists)
        
        if folder_exists:
            # Auto-suggest table name
            folder_name = os.path.basename(self.folder_path.text().strip())
            if not self.single_table_name.text():
                self.single_table_name.setText(f"{folder_name}_data")
                
        self.update_import_button()
        self.update_scan_sheets_button()
        
    def on_pattern_changed(self):
        """Handle pattern change"""
        if self.folder_path.text().strip() and os.path.exists(self.folder_path.text().strip()):
            self.scan_folder()
            
    def on_strategy_changed(self):
        """Handle table strategy change"""
        is_single = "Single Table" in self.table_strategy.currentText()
        self.single_table_widget.setVisible(is_single)
        self.multi_table_widget.setVisible(not is_single)
        self.update_import_button()
        
    def on_sheet_mode_changed(self):
        """Handle sheet mode change for Excel files"""
        is_specific = "Specific Sheet" in self.sheet_handling.currentText()
        self.specific_sheet_widget.setVisible(is_specific)
        self.update_scan_sheets_button()
        
    def update_scan_sheets_button(self):
        """Update scan sheets button state"""
        folder_exists = bool(self.folder_path.text().strip() and 
                           os.path.exists(self.folder_path.text().strip()))
        is_excel = "Excel" in self.file_type_combo.currentText()
        is_specific = "Specific Sheet" in self.sheet_handling.currentText()
        
        self.scan_sheets_btn.setEnabled(folder_exists and is_excel and is_specific)
        
    def scan_excel_sheets(self):
        """Scan Excel files in folder to get available sheet names"""
        folder = self.folder_path.text().strip()
        pattern = self.file_pattern.text().strip()
        
        if not folder or not os.path.exists(folder) or not pattern:
            return
            
        try:
            # Get Excel files
            file_pattern = os.path.join(folder, pattern)
            files = glob.glob(file_pattern)
            
            if pattern == "*.xlsx":
                files.extend(glob.glob(os.path.join(folder, "*.xls")))
                
            if not files:
                QMessageBox.warning(self, "No Files", "No Excel files found to scan.")
                return
                
            # Collect all unique sheet names
            all_sheets = set()
            
            for file_path in files[:5]:  # Limit to first 5 files for performance
                try:
                    if POLARS_AVAILABLE:
                        # Use Polars to get sheet names (faster)
                        try:
                            # Polars doesn't have a direct way to list sheets, so we use pandas for this
                            excel_file = pd.ExcelFile(file_path)
                            sheets = excel_file.sheet_names
                        except:
                            # Fallback to pandas
                            excel_file = pd.ExcelFile(file_path)
                            sheets = excel_file.sheet_names
                    else:
                        # Use pandas
                        excel_file = pd.ExcelFile(file_path)
                        sheets = excel_file.sheet_names
                        
                    all_sheets.update(sheets)
                    
                except Exception as e:
                    continue  # Skip files that can't be read
                    
            if all_sheets:
                # Sort sheets and update combo box
                sorted_sheets = sorted(list(all_sheets))
                self.specific_sheet_combo.clear()
                self.specific_sheet_combo.addItems(sorted_sheets)
                
                QMessageBox.information(self, "Sheets Found", 
                    f"Found {len(sorted_sheets)} unique sheet names across {len(files)} Excel files:\n\n" + 
                    ", ".join(sorted_sheets[:10]) + ("..." if len(sorted_sheets) > 10 else ""))
            else:
                QMessageBox.warning(self, "No Sheets", "No readable sheet names found in Excel files.")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error scanning Excel sheets: {str(e)}")
        
    def browse_folder(self):
        """Browse for folder"""
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.folder_path.setText(folder)
            
    def scan_folder(self):
        """Scan folder for files matching pattern"""
        folder = self.folder_path.text().strip()
        pattern = self.file_pattern.text().strip()
        
        if not folder or not os.path.exists(folder) or not pattern:
            self.file_list.clear()
            self.update_import_button()
            return
            
        try:
            file_pattern = os.path.join(folder, pattern)
            files = glob.glob(file_pattern)
            
            # For Excel, also check .xls files if pattern is *.xlsx
            if "Excel" in self.file_type_combo.currentText() and pattern == "*.xlsx":
                xls_pattern = os.path.join(folder, "*.xls")
                files.extend(glob.glob(xls_pattern))
                
            files = sorted([os.path.basename(f) for f in files])
            
            self.file_list.clear()
            if files:
                for file in files:
                    self.file_list.addItem(f"📄 {file}")
            else:
                self.file_list.addItem("No files found matching pattern")
                
            self.update_import_button()
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error scanning folder: {str(e)}")
            
    def update_import_button(self):
        """Update import button state"""
        folder_ok = bool(self.folder_path.text().strip() and 
                        os.path.exists(self.folder_path.text().strip()))
        files_ok = (self.file_list.count() > 0 and 
                   self.file_list.item(0).text() != "No files found matching pattern")
        
        if "Single Table" in self.table_strategy.currentText():
            table_ok = bool(self.single_table_name.text().strip())
        else:
            table_ok = True  # Multiple tables always OK
            
        self.import_btn.setEnabled(folder_ok and files_ok and table_ok and 
                                 bool(self.connection))
        
    def start_import(self):
        """Start the import process"""
        try:
            folder = self.folder_path.text().strip()
            pattern = self.file_pattern.text().strip()
            is_csv = "CSV" in self.file_type_combo.currentText()
            is_single_table = "Single Table" in self.table_strategy.currentText()
            
            # Get file list
            file_pattern = os.path.join(folder, pattern)
            files = glob.glob(file_pattern)
            
            if "Excel" in self.file_type_combo.currentText() and pattern == "*.xlsx":
                files.extend(glob.glob(os.path.join(folder, "*.xls")))
                
            if not files:
                QMessageBox.warning(self, "No Files", "No files found to import.")
                return
                
            # Show progress
            self.progress_bar.show()
            self.progress_bar.setRange(0, len(files))
            self.progress_bar.setValue(0)
            self.import_btn.setEnabled(False)
            
            # Start import
            self.import_files(folder, files, is_csv, is_single_table)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error starting import: {str(e)}")
            
    def import_files(self, folder, files, is_csv, is_single_table):
        """Import the files"""
        try:
            mode_map = {
                "Create New": "create_new",
                "Replace Existing": "replace",
                "Append to Existing": "append"
            }
            mode = mode_map[self.import_mode.currentText()]
            
            successful_imports = 0
            failed_imports = []
            all_dataframes = []
            main_app = self.parent()
            
            # Validate database connection first
            if not main_app.current_connection:
                QMessageBox.critical(self, "No Database", "No database connection available.")
                return
            
            for i, file_path in enumerate(files):
                try:
                    # Update progress with file size info
                    file_size = os.path.getsize(file_path)
                    file_size_mb = file_size / (1024 * 1024)
                    
                    if file_size_mb < 1:
                        size_str = f"{file_size / 1024:.1f} KB"
                    else:
                        size_str = f"{file_size_mb:.1f} MB"
                    
                    self.progress_bar.setValue(i)
                    self.progress_bar.setFormat(f"Processing {os.path.basename(file_path)} ({size_str}) - {i+1}/{len(files)}")
                    QApplication.processEvents()
                    
                    # Read file
                    if is_csv:
                        # CSV import with optimized processing (with chunking for large files)
                        df = self.read_csv_optimized(file_path, f"Processing {os.path.basename(file_path)} ({size_str})")
                    else:
                        # Excel import using ultra-fast Polars when available
                        df = self.read_excel_optimized(file_path)
                    
                    if df.empty:
                        failed_imports.append(f"{os.path.basename(file_path)}: Empty file")
                        continue
                        
                    # Clean column names
                    df.columns = [self.clean_column_name(col) for col in df.columns]
                    
                    if is_single_table:
                        # Add source file column
                        df['_source_file'] = os.path.basename(file_path)
                        all_dataframes.append(df)
                    else:
                        # Import each file to separate table
                        base_name = os.path.splitext(os.path.basename(file_path))[0]
                        table_prefix = self.table_prefix.text().strip()
                        table_name = f"{table_prefix}{base_name}" if table_prefix else base_name
                        table_name = self.clean_table_name(table_name)
                        
                        # Validate table name
                        if not table_name or table_name.strip() == "":
                            failed_imports.append(f"{os.path.basename(file_path)}: Invalid table name")
                            continue
                        
                        # Import to database with better error handling
                        try:
                            success = self.safe_database_import(df, table_name, mode, main_app)
                            if success:
                                successful_imports += 1
                            else:
                                failed_imports.append(f"{os.path.basename(file_path)}: Database import failed")
                        except Exception as import_error:
                            failed_imports.append(f"{os.path.basename(file_path)}: {str(import_error)}")
                            
                except Exception as e:
                    failed_imports.append(f"{os.path.basename(file_path)}: {str(e)}")
                    
            # Handle single table import
            if is_single_table and all_dataframes:
                try:
                    # Combine all dataframes with proper alignment
                    combined_df = self.combine_dataframes_safely(all_dataframes)
                    table_name = self.single_table_name.text().strip()
                    
                    # Validate table name
                    if not table_name or table_name.strip() == "":
                        failed_imports.append("Combined table: Invalid table name")
                    else:
                        # Import to database with better error handling
                        try:
                            success = self.safe_database_import(combined_df, table_name, mode, main_app)
                            if success:
                                successful_imports = 1
                            else:
                                failed_imports.append(f"Combined table '{table_name}': Database import failed")
                        except Exception as import_error:
                            failed_imports.append(f"Combined table '{table_name}': {str(import_error)}")
                        
                except Exception as e:
                    failed_imports.append(f"Combined table: {str(e)}")
                    
            # Update progress
            self.progress_bar.setValue(len(files))
            
            # Show results with performance info
            if successful_imports > 0:
                # Calculate total processing stats
                total_files = len(files)
                total_size = sum(os.path.getsize(f) for f in files if os.path.exists(f))
                total_size_mb = total_size / (1024 * 1024)
                
                message = f"Successfully imported {successful_imports} "
                if is_single_table:
                    message += f"files into table '{self.single_table_name.text()}'"
                else:
                    message += "tables"
                    
                # Add performance stats
                if total_size_mb > 0:
                    message += f"\n\n📊 Processing Stats:"
                    message += f"\n• Total Files: {total_files}"
                    if total_size_mb < 1:
                        message += f"\n• Total Size: {total_size / 1024:.1f} KB"
                    else:
                        message += f"\n• Total Size: {total_size_mb:.1f} MB"
                    
                    if POLARS_AVAILABLE and not is_csv:
                        message += f"\n⚡ Enhanced with Polars for ultra-fast processing"
                    elif is_csv:
                        if total_size_mb > 50:
                            message += f"\n⚡ Used chunked processing for large CSV files"
                        else:
                            message += f"\n⚡ Used optimized CSV processing"
                    
                if failed_imports:
                    message += f"\n\n❌ Failed imports ({len(failed_imports)}):\n"
                    message += "\n".join(failed_imports[:10])  # Limit to first 10 errors
                    if len(failed_imports) > 10:
                        message += f"\n... and {len(failed_imports) - 10} more errors"
                    
                QMessageBox.information(self, "Import Complete", message)
                
                # Refresh schema browser
                if hasattr(main_app, 'refresh_schema_browser'):
                    main_app.refresh_schema_browser()
                    main_app.check_schema_changes()
                    
                self.accept()
            else:
                error_message = "No files were imported successfully."
                if failed_imports:
                    error_message += "\n\nErrors:\n" + "\n".join(failed_imports[:10])
                    if len(failed_imports) > 10:
                        error_message += f"\n... and {len(failed_imports) - 10} more errors"
                QMessageBox.critical(self, "Import Failed", error_message)
                
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Critical error during import: {str(e)}")
        finally:
            self.progress_bar.hide()
            self.import_btn.setEnabled(True)
            
    def combine_dataframes_safely(self, dataframes):
        """Safely combine dataframes with proper column alignment - new columns go to the end"""
        if not dataframes:
            return pd.DataFrame()
            
        if len(dataframes) == 1:
            return dataframes[0]
            
        try:
            # Get the first dataframe's columns as the base order
            base_columns = list(dataframes[0].columns)
            
            # Find all unique columns across all dataframes
            all_columns_set = set(base_columns)
            for df in dataframes[1:]:
                all_columns_set.update(df.columns)
            
            # Create final column order: base columns first, then new columns at the end
            new_columns = [col for col in all_columns_set if col not in base_columns]
            all_columns = base_columns + sorted(new_columns)  # Sort only the new columns
            
            # Align all dataframes to have the same columns in the correct order
            aligned_dfs = []
            for df in dataframes:
                aligned_df = df.copy()
                
                # Add missing columns with NaN at the end
                for col in all_columns:
                    if col not in aligned_df.columns:
                        aligned_df[col] = pd.NA
                        
                # Reorder columns: existing columns first, new columns at the end
                aligned_df = aligned_df.reindex(columns=all_columns)
                aligned_dfs.append(aligned_df)
                
            # Combine all dataframes
            return pd.concat(aligned_dfs, ignore_index=True)
            
        except Exception as e:
            # Fallback to simple concatenation without sorting columns
            return pd.concat(dataframes, ignore_index=True, sort=False)
            
    def safe_database_import(self, df, table_name, mode, main_app):
        """Safely import dataframe to database with proper error handling"""
        try:
            # Validate inputs
            if df is None or df.empty:
                raise ValueError("DataFrame is empty")
                
            if not table_name or table_name.strip() == "":
                raise ValueError("Table name is empty")
                
            if not main_app.current_connection:
                raise ValueError("No database connection")
            
            # Clean table name
            table_name = self.clean_table_name(table_name)
            
            # Get database type
            db_type = main_app.current_connection_info.get('type', '').lower()
            
            # Handle different import modes
            if mode == 'create_new':
                # For create new, check if table exists first
                if self.table_exists(table_name, main_app):
                    raise ValueError(f"Table '{table_name}' already exists. Use 'Replace' or 'Append' mode instead.")
                    
            elif mode == 'replace':
                # For replace, drop table if it exists
                try:
                    self.drop_table_if_exists(table_name, main_app)
                except:
                    pass  # Ignore errors if table doesn't exist
                    
            elif mode == 'append':
                # For append, table should exist
                if not self.table_exists(table_name, main_app):
                    # Table doesn't exist, create it instead
                    mode = 'create_new'
            
            # Import the dataframe
            if db_type == 'duckdb':
                return self.import_to_duckdb(df, table_name, mode, main_app)
            elif db_type == 'sqlite':
                return self.import_to_sqlite(df, table_name, mode, main_app)
            else:
                raise ValueError(f"Unsupported database type: {db_type}")
                
        except Exception as e:
            print(f"Database import error: {str(e)}")  # Debug logging
            raise e
            
    def table_exists(self, table_name, main_app):
        """Check if table exists in database"""
        try:
            db_type = main_app.current_connection_info.get('type', '').lower()
            
            if db_type == 'duckdb':
                result = main_app.current_connection.execute(f"""
                    SELECT COUNT(*) FROM information_schema.tables 
                    WHERE table_name = '{table_name}' AND table_schema = 'main'
                """).fetchone()
                return result[0] > 0
            elif db_type == 'sqlite':
                cursor = main_app.current_connection.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                result = cursor.fetchone()
                cursor.close()
                return result is not None
            else:
                return False
        except:
            return False
            
    def drop_table_if_exists(self, table_name, main_app):
        """Drop table if it exists"""
        try:
            db_type = main_app.current_connection_info.get('type', '').lower()
            
            if db_type == 'duckdb':
                main_app.current_connection.execute(f"DROP TABLE IF EXISTS \"{table_name}\"")
            elif db_type == 'sqlite':
                main_app.current_connection.execute(f"DROP TABLE IF EXISTS \"{table_name}\"")
                main_app.current_connection.commit()
        except Exception as e:
            print(f"Error dropping table {table_name}: {str(e)}")
            
    def import_to_duckdb(self, df, table_name, mode, main_app):
        """Import DataFrame to DuckDB"""
        try:
            # Register the DataFrame with DuckDB
            main_app.current_connection.register('temp_import_df', df)
            
            if mode == 'append' and self.table_exists(table_name, main_app):
                # Append to existing table
                main_app.current_connection.execute(f'INSERT INTO "{table_name}" SELECT * FROM temp_import_df')
            else:
                # Create new table
                main_app.current_connection.execute(f'CREATE TABLE "{table_name}" AS SELECT * FROM temp_import_df')
            
            # Unregister the temporary DataFrame
            main_app.current_connection.unregister('temp_import_df')
            return True
            
        except Exception as e:
            # Clean up temporary DataFrame
            try:
                main_app.current_connection.unregister('temp_import_df')
            except:
                pass
            raise e
            
    def import_to_sqlite(self, df, table_name, mode, main_app):
        """Import DataFrame to SQLite"""
        try:
            if mode == 'append':
                if_exists = 'append'
            else:
                if_exists = 'replace'
                
            df.to_sql(table_name, main_app.current_connection, if_exists=if_exists, index=False)
            main_app.current_connection.commit()
            return True
            
        except Exception as e:
            main_app.current_connection.rollback()
            raise e
            
    def clean_column_name(self, name):
        """Clean column name for SQL compatibility - capitalize and replace special chars with underscores"""
        import re
        name = str(name).strip()
        
        # Convert to uppercase
        name = name.upper()
        
        # Replace spaces and special characters with underscores
        name = re.sub(r'[^A-Z0-9_]', '_', name)  # Replace non-alphanumeric chars with underscore
        name = re.sub(r'_+', '_', name)          # Replace multiple underscores with single
        name = name.strip('_')                   # Remove leading/trailing underscores
        
        if name and name[0].isdigit():
            name = f"COL_{name}"
            
        return name or "UNNAMED_COLUMN"
        
    def clean_table_name(self, name):
        """Clean table name for SQL compatibility"""
        import re
        name = str(name).strip()
        name = re.sub(r'[^\w]', '_', name)
        name = re.sub(r'_+', '_', name)
        name = name.strip('_')
        
        if name and name[0].isdigit():
            name = f"table_{name}"
            
        return name or "imported_table"
        
    def read_excel_optimized(self, file_path):
        """Read Excel file using ultra-fast Polars or fallback to pandas"""
        sheet_handling = self.sheet_handling.currentText()
        
        try:
            if POLARS_AVAILABLE:
                # Use Polars for maximum speed
                try:
                    if sheet_handling == 'All Sheets':
                        # Read all sheets and combine
                        excel_file = pd.ExcelFile(file_path)  # Still need pandas to get sheet names
                        sheet_dfs = []
                        for sheet_name in excel_file.sheet_names:
                            try:
                                df_pl = pl.read_excel(file_path, sheet_name=sheet_name)
                                df = df_pl.to_pandas()
                                df['_sheet_name'] = sheet_name
                                sheet_dfs.append(df)
                            except:
                                # Fallback to pandas for this sheet
                                df = pd.read_excel(file_path, sheet_name=sheet_name)
                                df['_sheet_name'] = sheet_name
                                sheet_dfs.append(df)
                        
                        if sheet_dfs:
                            df = pd.concat(sheet_dfs, ignore_index=True)
                        else:
                            raise Exception("No readable sheets found")
                            
                    elif sheet_handling == 'Specific Sheet':
                        # Read specific sheet
                        sheet_name = self.specific_sheet_combo.currentText()
                        if sheet_name:
                            try:
                                df_pl = pl.read_excel(file_path, sheet_name=sheet_name)
                                df = df_pl.to_pandas()
                            except:
                                # Fallback to pandas
                                df = pd.read_excel(file_path, sheet_name=sheet_name)
                        else:
                            # No sheet selected, use first sheet
                            df_pl = pl.read_excel(file_path)
                            df = df_pl.to_pandas()
                    else:
                        # First sheet only
                        df_pl = pl.read_excel(file_path)
                        df = df_pl.to_pandas()
                        
                except Exception as polars_error:
                    # Fallback to pandas if Polars fails
                    df = self._read_excel_pandas_fallback(file_path, sheet_handling)
                    
            else:
                # Use pandas directly
                df = self._read_excel_pandas_fallback(file_path, sheet_handling)
                
            # Clean column names for SQL compatibility
            df.columns = [self.clean_column_name(col) for col in df.columns]
            return df
            
        except Exception as e:
            raise Exception(f"Failed to read Excel file {os.path.basename(file_path)}: {str(e)}")
            
    def _read_excel_pandas_fallback(self, file_path, sheet_handling):
        """Fallback Excel reading using pandas"""
        if sheet_handling == 'All Sheets':
            # Read all sheets and combine
            excel_file = pd.ExcelFile(file_path)
            sheet_dfs = []
            for sheet_name in excel_file.sheet_names:
                sheet_df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheet_df['_sheet_name'] = sheet_name
                sheet_dfs.append(sheet_df)
            
            if sheet_dfs:
                return pd.concat(sheet_dfs, ignore_index=True)
            else:
                raise Exception("No readable sheets found")
                
        elif sheet_handling == 'Specific Sheet':
            # Read specific sheet
            sheet_name = self.specific_sheet_combo.currentText()
            if sheet_name:
                return pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                # No sheet selected, use first sheet
                return pd.read_excel(file_path)
        else:
            # First sheet only
            return pd.read_excel(file_path)
            
    def read_csv_optimized(self, file_path, status_message="Processing CSV"):
        """Read CSV file with optimized chunking and memory management like automation"""
        try:
            # Get CSV options
            encoding = self.encoding_combo.currentText()
            delimiter = self.delimiter_combo.currentText()
            
            if delimiter == 'Auto-detect':
                delimiter = self.detect_csv_delimiter_fast(file_path)
            elif delimiter == '\\t':
                delimiter = '\t'
                
            # Get file size for optimization decisions
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)
            
            # Determine processing strategy based on file size
            if file_size_mb < 50:  # Small files - read directly
                return self.read_csv_direct(file_path, encoding, delimiter)
            elif file_size_mb < 500:  # Medium files - optimized reading
                return self.read_csv_optimized_medium(file_path, encoding, delimiter, status_message)
            else:  # Large files - chunked processing
                return self.read_csv_chunked(file_path, encoding, delimiter, status_message)
                
        except Exception as e:
            raise Exception(f"Failed to read CSV file {os.path.basename(file_path)}: {str(e)}")
            
    def detect_csv_delimiter_fast(self, file_path):
        """Fast CSV delimiter detection using sample"""
        try:
            # Read first few lines for delimiter detection
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                sample = f.read(8192)  # Read first 8KB
                
            # Try common delimiters
            delimiters = [',', ';', '\t', '|']
            delimiter_scores = {}
            
            for delimiter in delimiters:
                lines = sample.split('\n')[:10]  # Check first 10 lines
                if len(lines) > 1:
                    counts = [line.count(delimiter) for line in lines if line.strip()]
                    if counts and len(set(counts)) == 1 and counts[0] > 0:
                        delimiter_scores[delimiter] = counts[0]
                        
            if delimiter_scores:
                return max(delimiter_scores.items(), key=lambda x: x[1])[0]
            else:
                return ','  # Default fallback
                
        except:
            return ','  # Default fallback
            
    def read_csv_direct(self, file_path, encoding, delimiter):
        """Direct CSV reading for small files"""
        return pd.read_csv(
            file_path,
            encoding=encoding,
            sep=delimiter,
            low_memory=False,
            engine='c'  # Use C engine for speed
        )
        
    def read_csv_optimized_medium(self, file_path, encoding, delimiter, status_message):
        """Optimized CSV reading for medium files"""
        try:
            # Use optimized pandas parameters
            df = pd.read_csv(
                file_path,
                encoding=encoding,
                sep=delimiter,
                low_memory=False,
                engine='c',  # C engine is faster
                dtype_backend='numpy_nullable',  # Use nullable dtypes
                na_filter=True,
                skip_blank_lines=True
            )
            
            # Optimize data types
            df = self.optimize_dataframe_dtypes(df)
            return df
            
        except Exception as e:
            # Fallback to basic reading
            return pd.read_csv(file_path, encoding=encoding, sep=delimiter)
            
    def read_csv_chunked(self, file_path, encoding, delimiter, status_message):
        """Chunked CSV reading for large files with progress reporting"""
        try:
            # Determine optimal chunk size based on available memory
            chunk_size = self.calculate_optimal_chunk_size(file_path)
            
            # Read CSV in chunks
            chunks = []
            total_rows = 0
            chunk_count = 0
            
            csv_reader = pd.read_csv(
                file_path,
                encoding=encoding,
                sep=delimiter,
                chunksize=chunk_size,
                low_memory=False,
                engine='c',
                dtype_backend='numpy_nullable'
            )
            
            for chunk in csv_reader:
                try:
                    # Optimize chunk data types
                    chunk = self.optimize_dataframe_dtypes(chunk)
                    
                    # Clean chunk if needed
                    chunk = self.clean_dataframe_chunk(chunk)
                    
                    chunks.append(chunk)
                    total_rows += len(chunk)
                    chunk_count += 1
                    
                    # Update progress every 10 chunks
                    if chunk_count % 10 == 0:
                        self.progress_bar.setFormat(f"{status_message} - {total_rows:,} rows processed")
                        QApplication.processEvents()
                        
                except Exception as chunk_error:
                    print(f"Error processing chunk {chunk_count}: {str(chunk_error)}")
                    continue
                    
            if not chunks:
                raise Exception("No valid data chunks found")
                
            # Combine all chunks efficiently
            return self.combine_chunks_efficiently(chunks)
            
        except Exception as e:
            # Fallback to regular reading
            return pd.read_csv(file_path, encoding=encoding, sep=delimiter)
            
    def calculate_optimal_chunk_size(self, file_path):
        """Calculate optimal chunk size based on file size and available memory"""
        try:
            file_size = os.path.getsize(file_path)
            file_size_mb = file_size / (1024 * 1024)
            
            # Base chunk size on file size
            if file_size_mb < 100:
                return 50000
            elif file_size_mb < 500:
                return 100000
            elif file_size_mb < 1000:
                return 200000
            else:
                return 500000
                
        except:
            return 100000  # Default
            
    def optimize_dataframe_dtypes(self, df):
        """Optimize DataFrame data types for memory efficiency"""
        try:
            for col in df.columns:
                col_type = df[col].dtype
                
                if col_type == 'object':
                    # Try to convert to more efficient types
                    try:
                        # Try numeric conversion
                        df[col] = pd.to_numeric(df[col], errors='ignore')
                        if df[col].dtype == 'float64':
                            # Try to downcast to float32 if no precision loss
                            df[col] = pd.to_numeric(df[col], downcast='float')
                        elif df[col].dtype in ['int64', 'int32']:
                            # Try to downcast integers
                            df[col] = pd.to_numeric(df[col], downcast='integer')
                    except:
                        pass
                        
                elif col_type == 'float64':
                    # Try to downcast float64 to float32
                    try:
                        df[col] = pd.to_numeric(df[col], downcast='float')
                    except:
                        pass
                        
                elif col_type in ['int64', 'int32']:
                    # Try to downcast integers
                    try:
                        df[col] = pd.to_numeric(df[col], downcast='integer')
                    except:
                        pass
                        
            return df
            
        except Exception as e:
            # Return original DataFrame if optimization fails
            return df
            
    def clean_dataframe_chunk(self, chunk):
        """Clean and validate DataFrame chunk"""
        try:
            # Remove completely empty rows
            chunk = chunk.dropna(how='all')
            
            # Remove duplicate header rows (common in concatenated CSVs)
            if len(chunk) > 0:
                header_mask = chunk.astype(str).eq(chunk.columns.astype(str)).all(axis=1)
                chunk = chunk[~header_mask]
                
            return chunk
            
        except Exception as e:
            return chunk
            
    def combine_chunks_efficiently(self, chunks):
        """Efficiently combine DataFrame chunks"""
        try:
            if not chunks:
                return pd.DataFrame()
                
            if len(chunks) == 1:
                return chunks[0]
                
            # Use concat with optimized parameters
            combined = pd.concat(
                chunks,
                ignore_index=True,
                copy=False,  # Avoid unnecessary copying
                sort=False   # Don't sort columns
            )
            
            # Final cleanup and optimization
            combined = self.optimize_dataframe_dtypes(combined)
            
            # Reset progress bar format
            self.progress_bar.setFormat("")
            
            return combined
            
        except Exception as e:
            # Fallback to simple concatenation
            return pd.concat(chunks, ignore_index=True)


class DataImportDialog(QDialog):
    """Dialog for importing data from various file formats"""
    
    def __init__(self, parent=None, connection=None, connection_info=None):
        super().__init__(parent)
        self.connection = connection
        self.connection_info = connection_info
        self.setWindowTitle("Import Data")
        self.setModal(True)
        self.resize(850, 750)  # Made larger and resizable
        self.setMinimumSize(700, 600)  # Set minimum size
        self.init_ui()
    
    def init_ui(self):
        # Create main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Create scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        
        # Create content widget for scroll area
        content_widget = QWidget()
        layout = QVBoxLayout(content_widget)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title_label = QLabel("📁 Import Data to Database")
        title_label.setStyleSheet("font-size: 14pt; font-weight: bold; color: #0078d4; margin-bottom: 10px;")
        layout.addWidget(title_label)
        
        # File selection
        file_group = QGroupBox("Select Data File")
        file_layout = QVBoxLayout(file_group)
        
        file_path_layout = QHBoxLayout()
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("Select a data file...")
        self.file_path_edit.setReadOnly(True)
        
        self.browse_button = QPushButton("📁 Browse")
        self.browse_button.clicked.connect(self.browse_file)
        self.browse_button.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #106ebe; }
        """)
        
        file_path_layout.addWidget(self.file_path_edit)
        file_path_layout.addWidget(self.browse_button)
        file_layout.addLayout(file_path_layout)
        
        # File info
        self.file_info_label = QLabel("No file selected")
        self.file_info_label.setStyleSheet("color: #666; font-style: italic;")
        file_layout.addWidget(self.file_info_label)
        
        layout.addWidget(file_group)
        
        # Table options
        table_group = QGroupBox("Table Options")
        table_layout = QVBoxLayout(table_group)
        
        # Import mode
        mode_label = QLabel("Import Mode:")
        mode_label.setStyleSheet("font-weight: bold; margin-bottom: 5px;")
        table_layout.addWidget(mode_label)
        
        self.create_new_radio = QRadioButton("🆕 Create new table")
        self.create_new_radio.setChecked(True)
        self.create_new_radio.setToolTip("Create a new table (fails if table already exists)")
        
        self.append_radio = QRadioButton("➕ Append to existing table")
        self.append_radio.setToolTip("Add data to existing table (table must exist)")
        
        self.replace_radio = QRadioButton("🔄 Replace table (create if not exists)")
        self.replace_radio.setToolTip("Drop existing table and create new one with imported data, or create new table if it doesn't exist")
        
        table_layout.addWidget(self.create_new_radio)
        table_layout.addWidget(self.append_radio)
        table_layout.addWidget(self.replace_radio)
        
        # Table name input (for create mode)
        self.table_name_widget = QWidget()
        table_name_layout = QHBoxLayout(self.table_name_widget)
        table_name_layout.setContentsMargins(0, 5, 0, 0)
        table_name_layout.addWidget(QLabel("Table Name:"))
        self.table_name_edit = QLineEdit()
        self.table_name_edit.setPlaceholderText("Enter table name...")
        table_name_layout.addWidget(self.table_name_edit)
        table_layout.addWidget(self.table_name_widget)
        
        # Table selection dropdown (for append/replace modes)
        self.table_select_widget = QWidget()
        table_select_layout = QHBoxLayout(self.table_select_widget)
        table_select_layout.setContentsMargins(0, 5, 0, 0)
        table_select_layout.addWidget(QLabel("Select Table:"))
        self.table_select_combo = QComboBox()
        self.table_select_combo.setStyleSheet("""
            QComboBox {
                background-color: #f8f9fa;
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 5px;
                min-height: 20px;
            }
            QComboBox:focus {
                border-color: #0078d4;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #666;
                margin-right: 5px;
            }
        """)
        table_select_layout.addWidget(self.table_select_combo)
        table_layout.addWidget(self.table_select_widget)
        self.table_select_widget.hide()  # Initially hidden
        
        layout.addWidget(table_group)
        
        # Excel sheet selection (initially hidden)
        self.sheet_group = QGroupBox("Excel Sheet Selection")
        sheet_layout = QVBoxLayout(self.sheet_group)
        
        self.sheet_combo = QComboBox()
        sheet_layout.addWidget(self.sheet_combo)
        
        layout.addWidget(self.sheet_group)
        self.sheet_group.hide()
        
        # Advanced options
        advanced_group = QGroupBox("Advanced Options")
        advanced_layout = QVBoxLayout(advanced_group)
        
        # CSV options
        self.csv_options_widget = QWidget()
        csv_options_layout = QFormLayout(self.csv_options_widget)
        
        delimiter_layout = QHBoxLayout()
        
        # Delimiter dropdown with common options
        self.delimiter_combo = QComboBox()
        self.delimiter_combo.setMaximumWidth(120)
        self.delimiter_combo.addItems([
            "Comma (,)",
            "Semicolon (;)", 
            "Tab (\\t)",
            "Pipe (|)",
            "Space ( )",
            "Custom..."
        ])
        self.delimiter_combo.setCurrentText("Comma (,)")
        self.delimiter_combo.currentTextChanged.connect(self.on_delimiter_changed)
        
        # Custom delimiter input (initially hidden)
        self.delimiter_edit = QLineEdit(",")
        self.delimiter_edit.setMaximumWidth(60)
        self.delimiter_edit.setPlaceholderText("Custom")
        self.delimiter_edit.hide()
        
        self.auto_detect_button = QPushButton("🔍 Auto")
        self.auto_detect_button.setMaximumWidth(60)
        self.auto_detect_button.setToolTip("Auto-detect delimiter from file")
        self.auto_detect_button.clicked.connect(self.auto_detect_delimiter)
        self.auto_detect_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border: none;
                padding: 4px 8px;
                border-radius: 3px;
                font-size: 10px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #138496; }
        """)
        
        delimiter_layout.addWidget(self.delimiter_combo)
        delimiter_layout.addWidget(self.delimiter_edit)
        delimiter_layout.addWidget(self.auto_detect_button)
        delimiter_layout.addStretch()
        
        csv_options_layout.addRow("Delimiter:", delimiter_layout)
        
        self.encoding_combo = QComboBox()
        self.encoding_combo.addItems(["utf-8", "latin-1", "cp1252", "utf-16"])
        csv_options_layout.addRow("Encoding:", self.encoding_combo)
        
        self.header_checkbox = QCheckBox("First row contains headers")
        self.header_checkbox.setChecked(True)
        csv_options_layout.addRow("", self.header_checkbox)
        
        advanced_layout.addWidget(self.csv_options_widget)
        layout.addWidget(advanced_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.cancel_button = QPushButton("❌ Cancel")
        self.cancel_button.clicked.connect(self.reject)
        self.cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #5a6268; }
        """)
        
        self.import_button = QPushButton("📥 Import Data")
        self.import_button.clicked.connect(self.accept)
        self.import_button.setEnabled(False)
        self.import_button.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #218838; }
            QPushButton:disabled { background-color: #6c757d; opacity: 0.6; }
        """)
        
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.import_button)
        layout.addLayout(button_layout)
        
        # Connect signals
        self.file_path_edit.textChanged.connect(self.update_ui)
        self.table_name_edit.textChanged.connect(self.update_ui)
        self.table_select_combo.currentTextChanged.connect(self.update_ui)
        
        # Connect radio button signals to update table selection UI
        self.create_new_radio.toggled.connect(self.update_table_selection_ui)
        self.append_radio.toggled.connect(self.update_table_selection_ui)
        self.replace_radio.toggled.connect(self.update_table_selection_ui)
        
        # Load existing tables if connection is available
        self.load_existing_tables()
        
        # Add content widget to scroll area and scroll area to main layout
        scroll_area.setWidget(content_widget)
        main_layout.addWidget(scroll_area)
    
    def on_delimiter_changed(self):
        """Handle delimiter dropdown selection change"""
        selected = self.delimiter_combo.currentText()
        
        if selected == "Custom...":
            # Show custom input field
            self.delimiter_edit.show()
            self.delimiter_edit.setFocus()
        else:
            # Hide custom input and set predefined delimiter
            self.delimiter_edit.hide()
            
            # Map display text to actual delimiter
            delimiter_map = {
                "Comma (,)": ",",
                "Semicolon (;)": ";",
                "Tab (\\t)": "\t",
                "Pipe (|)": "|",
                "Space ( )": " "
            }
            
            actual_delimiter = delimiter_map.get(selected, ",")
            self.delimiter_edit.setText(actual_delimiter)
    
    def get_current_delimiter(self):
        """Get the currently selected delimiter value"""
        if self.delimiter_combo.currentText() == "Custom...":
            return self.delimiter_edit.text()
        else:
            delimiter_map = {
                "Comma (,)": ",",
                "Semicolon (;)": ";",
                "Tab (\\t)": "\t",
                "Pipe (|)": "|",
                "Space ( )": " "
            }
            return delimiter_map.get(self.delimiter_combo.currentText(), ",")
    
    def auto_detect_delimiter(self):
        """Auto-detect delimiter for the selected file"""
        file_path = self.file_path_edit.text()
        if not file_path or not os.path.exists(file_path):
            QMessageBox.warning(self, "No File", "Please select a file first.")
            return
        
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext not in ['.csv', '.txt', '.tsv']:
            QMessageBox.information(self, "Not Applicable", "Auto-detection only works for CSV, TSV, and TXT files.")
            return
        
        # Get the main app instance to use its detect_csv_delimiter method
        main_app = self.parent()
        while main_app and not hasattr(main_app, 'detect_csv_delimiter'):
            main_app = main_app.parent()
        
        if main_app and hasattr(main_app, 'detect_csv_delimiter'):
            detected_delimiter = main_app.detect_csv_delimiter(file_path)
            
            # Map detected delimiter to dropdown option
            delimiter_to_combo = {
                ',': "Comma (,)",
                ';': "Semicolon (;)",
                '\t': "Tab (\\t)",
                '|': "Pipe (|)",
                ' ': "Space ( )"
            }
            
            combo_option = delimiter_to_combo.get(detected_delimiter)
            
            if combo_option:
                # Set the dropdown to the detected delimiter
                self.delimiter_combo.setCurrentText(combo_option)
                self.delimiter_edit.hide()  # Hide custom input
            else:
                # Use custom option for unusual delimiters
                self.delimiter_combo.setCurrentText("Custom...")
                self.delimiter_edit.setText(detected_delimiter)
                self.delimiter_edit.show()
            
            # Show confirmation
            delimiter_name = {
                ',': 'comma',
                ';': 'semicolon', 
                '\t': 'tab',
                '|': 'pipe',
                ' ': 'space'
            }.get(detected_delimiter, f"'{detected_delimiter}'")
            
            QMessageBox.information(self, "Delimiter Detected", 
                                  f"Detected delimiter: {delimiter_name}")
        else:
            QMessageBox.warning(self, "Error", "Could not access delimiter detection functionality.")
    
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Data File",
            "",
            "All Supported (*.csv *.xlsx *.xls *.parquet *.json *.tsv *.txt);;"
            "CSV Files (*.csv);;"
            "Excel Files (*.xlsx *.xls);;"
            "Parquet Files (*.parquet);;"
            "JSON Files (*.json);;"
            "TSV Files (*.tsv);;"
            "Text Files (*.txt);;"
            "All Files (*)"
        )
        
        if file_path:
            self.file_path_edit.setText(file_path)
            self.analyze_file(file_path)
    
    def analyze_file(self, file_path):
        try:
            file_size = os.path.getsize(file_path) / 1024 / 1024  # Size in MB
            file_ext = os.path.splitext(file_path)[1].lower()
            
            # Update file info
            self.file_info_label.setText(f"File: {os.path.basename(file_path)} | Size: {file_size:.2f} MB | Type: {file_ext}")
            
            # Auto-suggest table name only if field is empty
            if not self.table_name_edit.text().strip():
                base_name = os.path.splitext(os.path.basename(file_path))[0]
                # Clean table name (remove special characters, replace with underscores)
                clean_name = re.sub(r'[^a-zA-Z0-9_]', '_', base_name).lower()
                self.table_name_edit.setText(clean_name)
            
            # Handle Excel files - show sheet selection
            if file_ext in ['.xlsx', '.xls']:
                self.load_excel_sheets(file_path)
                self.sheet_group.show()
                self.csv_options_widget.hide()
            else:
                self.sheet_group.hide()
                if file_ext in ['.csv', '.tsv', '.txt']:
                    self.csv_options_widget.show()
                    
                    # Auto-detect delimiter for CSV/TSV/TXT files
                    main_app = self.parent()
                    while main_app and not hasattr(main_app, 'detect_csv_delimiter'):
                        main_app = main_app.parent()
                    
                    if main_app and hasattr(main_app, 'detect_csv_delimiter'):
                        detected_delimiter = main_app.detect_csv_delimiter(file_path)
                        
                        # Map detected delimiter to dropdown option
                        delimiter_to_combo = {
                            ',': "Comma (,)",
                            ';': "Semicolon (;)",
                            '\t': "Tab (\\t)",
                            '|': "Pipe (|)",
                            ' ': "Space ( )"
                        }
                        
                        combo_option = delimiter_to_combo.get(detected_delimiter)
                        
                        if combo_option:
                            self.delimiter_combo.setCurrentText(combo_option)
                            self.delimiter_edit.hide()
                        else:
                            self.delimiter_combo.setCurrentText("Custom...")
                            self.delimiter_edit.setText(detected_delimiter)
                            self.delimiter_edit.show()
                    elif file_ext == '.tsv':
                        self.delimiter_combo.setCurrentText("Tab (\\t)")
                        self.delimiter_edit.hide()
                else:
                    self.csv_options_widget.hide()
            
        except Exception as e:
            self.file_info_label.setText(f"Error analyzing file: {str(e)}")
    
    def load_excel_sheets(self, file_path):
        try:
            # Read Excel file to get sheet names
            excel_file = pd.ExcelFile(file_path)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(excel_file.sheet_names)
        except Exception as e:
            QMessageBox.warning(self, "Excel Error", f"Could not read Excel sheets: {str(e)}")
    
    def load_existing_tables(self):
        """Load existing tables from the database into the dropdown"""
        self.table_select_combo.clear()
        
        if not self.connection or not self.connection_info:
            return
            
        try:
            if self.connection_info['type'].lower() == 'duckdb':
                tables_df = self.connection.execute("SHOW TABLES").fetchdf()
                existing_tables = tables_df['name'].tolist() if not tables_df.empty else []
            else:  # SQLite
                cursor = self.connection.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
                existing_tables = [row[0] for row in cursor.fetchall()]
            
            if existing_tables:
                # Add tables with icons
                for table in existing_tables:
                    self.table_select_combo.addItem(f"📊 {table}")
            else:
                self.table_select_combo.addItem("(No tables found)")
                self.table_select_combo.setEnabled(False)
                
        except Exception as e:
            self.table_select_combo.addItem(f"Error loading tables: {str(e)}")
            self.table_select_combo.setEnabled(False)
    
    def update_table_selection_ui(self):
        """Update the table selection UI based on the selected import mode"""
        if self.create_new_radio.isChecked():
            self.table_name_widget.show()
            self.table_select_widget.hide()
        else:  # append or replace mode
            self.table_name_widget.hide()
            self.table_select_widget.show()
            
            # Check if we have tables available
            if self.table_select_combo.count() == 0 or not self.table_select_combo.isEnabled():
                # Show warning if no tables available for append/replace
                if self.append_radio.isChecked():
                    QMessageBox.warning(
                        self, 
                        "No Tables Available", 
                        "No existing tables found in the database.\nPlease create a table first or use 'Create new table' mode."
                    )
                    self.create_new_radio.setChecked(True)
                    return
                elif self.replace_radio.isChecked() and self.table_select_combo.count() == 0:
                    QMessageBox.information(
                        self, 
                        "No Tables to Replace", 
                        "No existing tables found to replace.\nYou can use 'Create new table' mode instead."
                    )
                    self.create_new_radio.setChecked(True)
                    return
        
        self.update_ui()
    
    def update_ui(self):
        # Enable import button only if file and appropriate table selection are provided
        has_file = bool(self.file_path_edit.text().strip())
        
        if self.create_new_radio.isChecked():
            has_table_info = bool(self.table_name_edit.text().strip())
        else:  # append or replace mode
            current_text = self.table_select_combo.currentText()
            has_table_info = (current_text and 
                            current_text != "(No tables found)" and
                            not current_text.startswith("Error") and
                            (current_text.startswith("📊 ") or current_text not in ["(No tables found)"]))
        
        self.import_button.setEnabled(has_file and has_table_info)
    
    def get_import_info(self):
        file_path = self.file_path_edit.text()
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # Get table name based on mode
        if self.create_new_radio.isChecked():
            table_name = self.table_name_edit.text().strip()
        else:  # append or replace mode
            table_name = self.table_select_combo.currentText()
            # Remove icon prefix if present
            if table_name.startswith("📊 "):
                table_name = table_name[2:]  # Remove "📊 " prefix
        
        import_info = {
            'file_path': file_path,
            'table_name': table_name,
            'file_type': file_ext,
            'mode': 'create' if self.create_new_radio.isChecked() else 
                   'append' if self.append_radio.isChecked() else 'replace'
        }
        
        # Add Excel sheet if applicable
        if file_ext in ['.xlsx', '.xls'] and self.sheet_combo.currentText():
            import_info['sheet_name'] = self.sheet_combo.currentText()
        
        # Add CSV options if applicable
        if file_ext in ['.csv', '.tsv', '.txt']:
            import_info['delimiter'] = self.get_current_delimiter()
            import_info['encoding'] = self.encoding_combo.currentText()
            import_info['header'] = self.header_checkbox.isChecked()
        
        return import_info
class QueryTab(QWidget):
    schema_changed = pyqtSignal()  # Signal to notify when schema might have changed
    
    def __init__(self, parent=None, connection=None, connection_info=None):
        super().__init__(parent)
        self.connection = connection
        self.connection_info = connection_info
        self.query_worker = None
        self.table_names = []
        self.column_names = []
        
        # Fullscreen state
        self.is_fullscreen = False
        self.normal_geometry = None
        self.main_window = None
        
        # Layout
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        
        # Splitter for editor and results
        self.splitter = QSplitter(Qt.Orientation.Vertical)
        
        # Query editor with auto-completion
        self.editor = SQLTextEdit()
        
        # Set editor colors first
        editor_palette = self.editor.palette()
        editor_palette.setColor(QPalette.ColorRole.Base, ColorScheme.BACKGROUND)
        editor_palette.setColor(QPalette.ColorRole.Text, ColorScheme.TEXT)
        self.editor.setPalette(editor_palette)
        
        # Apply syntax highlighting after setting up the editor
        self.highlighter = SQLHighlighter(self.editor.document())
        
        # Force a refresh of syntax highlighting
        QTimer.singleShot(100, self.highlighter.rehighlight)
        
        # Results area
        self.results_widget = QWidget()
        self.results_layout = QVBoxLayout(self.results_widget)
        self.results_layout.setContentsMargins(0, 0, 0, 0)
        
        # Editor header with fullscreen button
        self.editor_header = QWidget()
        self.editor_header_layout = QHBoxLayout(self.editor_header)
        self.editor_header_layout.setContentsMargins(5, 5, 5, 5)
        
        self.editor_label = QLabel("Query Editor")
        self.editor_label.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        
        self.fullscreen_button = QPushButton("⛶")
        self.fullscreen_button.setMaximumSize(30, 25)
        self.fullscreen_button.setToolTip("Toggle fullscreen editor (F11)")
        self.fullscreen_button.clicked.connect(self.toggle_fullscreen)
        self.fullscreen_button.setStyleSheet("""
            QPushButton {
                background-color: #404040;
                color: #ffffff;
                border: 1px solid #606060;
                border-radius: 3px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #505050;
                border-color: #0078d4;
            }
            QPushButton:pressed {
                background-color: #0078d4;
            }
        """)
        
        self.editor_header_layout.addWidget(self.editor_label)
        self.editor_header_layout.addStretch()
        self.editor_header_layout.addWidget(self.fullscreen_button)
        
        # Results header
        self.results_header = QWidget()
        self.results_header_layout = QHBoxLayout(self.results_header)
        self.results_header_layout.setContentsMargins(5, 5, 5, 5)
        
        self.results_label = QLabel("Results")
        self.results_label.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        self.results_info = QLabel("")
        
        # Export button
        self.export_button = QPushButton("📤 Export")
        self.export_button.setMaximumSize(80, 25)
        self.export_button.setToolTip("Export results to file")
        self.export_button.setEnabled(False)  # Initially disabled
        self.export_button.clicked.connect(self.show_export_menu)
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: #ffffff;
                border: 1px solid #106ebe;
                border-radius: 3px;
                font-weight: bold;
                padding: 2px 8px;
            }
            QPushButton:hover {
                background-color: #106ebe;
                border-color: #005a9e;
            }
            QPushButton:pressed {
                background-color: #005a9e;
            }
            QPushButton:disabled {
                background-color: #404040;
                color: #808080;
                border-color: #606060;
            }
        """)
        
        self.results_header_layout.addWidget(self.results_label)
        self.results_header_layout.addWidget(self.results_info, 1)
        self.results_header_layout.addWidget(self.export_button)
        
        # Results table
        self.results_table = QTableView()
        self.results_table.setAlternatingRowColors(True)
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.verticalHeader().setVisible(True)
        self.results_table.setSortingEnabled(True)
        
        # Set table colors
        table_palette = self.results_table.palette()
        table_palette.setColor(QPalette.ColorRole.Base, QColor(55, 55, 55))
        table_palette.setColor(QPalette.ColorRole.AlternateBase, QColor(45, 45, 45))
        table_palette.setColor(QPalette.ColorRole.Text, ColorScheme.TEXT)
        self.results_table.setPalette(table_palette)
        
        # Create editor container with header
        self.editor_container = QWidget()
        self.editor_container_layout = QVBoxLayout(self.editor_container)
        self.editor_container_layout.setContentsMargins(0, 0, 0, 0)
        self.editor_container_layout.addWidget(self.editor_header)
        self.editor_container_layout.addWidget(self.editor)
        
        # Add widgets to layouts
        self.results_layout.addWidget(self.results_header)
        self.results_layout.addWidget(self.results_table)
        
        self.splitter.addWidget(self.editor_container)
        self.splitter.addWidget(self.results_widget)
        self.splitter.setSizes([200, 300])
        
        self.layout.addWidget(self.splitter)
    
    def execute_query(self):
        self.editor_header_layout.setContentsMargins(5, 5, 5, 5)
        
        self.editor_label = QLabel("Query Editor")
        self.editor_label.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        
        self.fullscreen_button = QPushButton("⛶")
        self.fullscreen_button.setMaximumSize(30, 25)
        self.fullscreen_button.setToolTip("Toggle fullscreen editor (F11)")
        self.fullscreen_button.clicked.connect(self.toggle_fullscreen)
        self.fullscreen_button.setStyleSheet("""
            QPushButton {
                background-color: #404040;
                color: #ffffff;
                border: 1px solid #606060;
                border-radius: 3px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #505050;
                border-color: #0078d4;
            }
            QPushButton:pressed {
                background-color: #0078d4;
            }
        """)
        
        self.editor_header_layout.addWidget(self.editor_label)
        self.editor_header_layout.addStretch()
        self.editor_header_layout.addWidget(self.fullscreen_button)
        
        # Results header
        self.results_header = QWidget()
        self.results_header_layout = QHBoxLayout(self.results_header)
        self.results_header_layout.setContentsMargins(5, 5, 5, 5)
        
        self.results_label = QLabel("Results")
        self.results_label.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        self.results_info = QLabel("")
        
        # Export button
        self.export_button = QPushButton("📤 Export")
        self.export_button.setMaximumSize(80, 25)
        self.export_button.setToolTip("Export results to file")
        self.export_button.setEnabled(False)  # Initially disabled
        self.export_button.clicked.connect(self.show_export_menu)
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #0078d4;
                color: #ffffff;
                border: 1px solid #106ebe;
                border-radius: 3px;
                font-weight: bold;
                padding: 2px 8px;
            }
            QPushButton:hover {
                background-color: #106ebe;
                border-color: #005a9e;
            }
            QPushButton:pressed {
                background-color: #005a9e;
            }
            QPushButton:disabled {
                background-color: #404040;
                color: #808080;
                border-color: #606060;
            }
        """)
        
        self.results_header_layout.addWidget(self.results_label)
        self.results_header_layout.addWidget(self.results_info, 1)
        self.results_header_layout.addWidget(self.export_button)
        
        # Results table
        self.results_table = QTableView()
        self.results_table.setAlternatingRowColors(True)
        self.results_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.verticalHeader().setVisible(True)
        self.results_table.setSortingEnabled(True)
        
        # Set table colors
        table_palette = self.results_table.palette()
        table_palette.setColor(QPalette.ColorRole.Base, QColor(55, 55, 55))
        table_palette.setColor(QPalette.ColorRole.AlternateBase, QColor(45, 45, 45))
        table_palette.setColor(QPalette.ColorRole.Text, ColorScheme.TEXT)
        self.results_table.setPalette(table_palette)
        
        # Create editor container with header
        self.editor_container = QWidget()
        self.editor_container_layout = QVBoxLayout(self.editor_container)
        self.editor_container_layout.setContentsMargins(0, 0, 0, 0)
        self.editor_container_layout.addWidget(self.editor_header)
        self.editor_container_layout.addWidget(self.editor)
        
        # Add widgets to layouts
        self.results_layout.addWidget(self.results_header)
        self.results_layout.addWidget(self.results_table)
        
        self.splitter.addWidget(self.editor_container)
        self.splitter.addWidget(self.results_widget)
        self.splitter.setSizes([200, 300])
        
        self.layout.addWidget(self.splitter)
        
    def execute_query(self):
        if not self.connection:
            QMessageBox.warning(self, "No Connection", "Please connect to a database first.")
            return
        
        # Check if there's a selection first
        cursor = self.editor.textCursor()
        if cursor.hasSelection():
            query = cursor.selectedText().strip()
            execution_type = "selection"
        else:
            query = self.editor.toPlainText().strip()
            execution_type = "full query"
            
        if not query:
            return
            
        # Show what's being executed
        query_preview = query[:100] + "..." if len(query) > 100 else query
        self.results_info.setText(f"Executing {execution_type}: {query_preview}")
        
        # Disable editor during execution
        self.editor.setReadOnly(True)
        
        # Execute query in a separate thread with enhanced worker
        # Check user preference for lazy loading
        settings = QSettings('SQLEditor', 'QuerySettings')
        lazy_enabled = settings.value('enable_lazy_loading', True, type=bool)
        lazy_threshold = settings.value('lazy_loading_threshold', 100000, type=int)
        
        self.query_worker = EnhancedQueryWorker(
            self.connection, 
            query, 
            use_lazy_loading=lazy_enabled,
            row_limit=lazy_threshold
        )
        self.query_worker.finished.connect(self.handle_query_results)
        self.query_worker.lazy_finished.connect(self.handle_lazy_query_results)
        self.query_worker.error.connect(self.handle_query_error)
        self.query_worker.start()
    
    def execute_selected_query(self):
        """Execute only the selected text as a query"""
        cursor = self.editor.textCursor()
        if cursor.hasSelection():
            self.execute_query()  # Will automatically detect selection
        else:
            QMessageBox.information(self, "No Selection", "Please select some text in the query editor first.")
    
    def handle_query_results(self, df, execution_time):
        """Handle regular query results (smaller datasets)"""
        # Update table model with results
        self.model = PandasTableModel(df)
        self.results_table.setModel(self.model)
        
        # Auto-resize columns for better visibility
        for i in range(self.model.columnCount()):
            self.results_table.setColumnWidth(i, 200)
        
        # Update results info
        row_count = len(df)
        self.results_info.setText(f"{row_count:,} {'row' if row_count == 1 else 'rows'} returned in {execution_time:.3f} seconds")
        
        # Enable export button if we have results
        self.export_button.setEnabled(row_count > 0)
        
        # Check if this was a DDL statement that might have changed the schema
        query = self.editor.toPlainText().strip().upper()
        ddl_keywords = ['CREATE TABLE', 'DROP TABLE', 'ALTER TABLE', 'CREATE VIEW', 'DROP VIEW', 'CREATE INDEX', 'DROP INDEX']
        if any(keyword in query for keyword in ddl_keywords):
            self.schema_changed.emit()
        
        # Re-enable editor
        self.editor.setReadOnly(False)
    
    def handle_lazy_query_results(self, lazy_model, execution_time):
        """Handle lazy query results (massive datasets)"""
        # Set the lazy loading model
        self.model = lazy_model
        self.results_table.setModel(self.model)
        
        # Auto-resize columns for better visibility
        for i in range(self.model.columnCount()):
            self.results_table.setColumnWidth(i, 200)
        
        # Update results info with lazy loading indicator
        row_count = lazy_model.total_rows
        if row_count > 1000000:  # Show formatted large numbers
            if row_count >= 1000000000:  # Billions
                display_count = f"{row_count/1000000000:.1f}B"
            elif row_count >= 1000000:  # Millions
                display_count = f"{row_count/1000000:.1f}M"
            else:
                display_count = f"{row_count:,}"
        else:
            display_count = f"{row_count:,}"
        
        self.results_info.setText(f"🚀 {display_count} rows (lazy loaded) • Query completed in {execution_time:.3f}s • Showing data on-demand")
        
        # Enable export button if we have results
        self.export_button.setEnabled(row_count > 0)
        
        # Check if this was a DDL statement that might have changed the schema
        query = self.editor.toPlainText().strip().upper()
        ddl_keywords = ['CREATE TABLE', 'DROP TABLE', 'ALTER TABLE', 'CREATE VIEW', 'DROP VIEW', 'CREATE INDEX', 'DROP INDEX']
        if any(keyword in query for keyword in ddl_keywords):
            self.schema_changed.emit()
        
        # Re-enable editor
        self.editor.setReadOnly(False)
    
    def handle_query_error(self, error_message):
        self.results_info.setText(f"Error: {error_message}")
        self.export_button.setEnabled(False)  # Disable export on error
        self.editor.setReadOnly(False)
        
    def update_schema_completions(self, table_names=None, column_names=None):
        """Update the auto-completion with table and column names from the schema"""
        if table_names is not None:
            self.table_names = table_names
        if column_names is not None:
            self.column_names = column_names
            
        # Update the editor's completer
        self.editor.update_completions(self.table_names, self.column_names)
    
    def toggle_fullscreen(self):
        """Toggle fullscreen mode for the query editor"""
        if not self.is_fullscreen:
            # Enter fullscreen mode
            self.enter_fullscreen()
        else:
            # Exit fullscreen mode
            self.exit_fullscreen()
    
    def enter_fullscreen(self):
        """Enter fullscreen mode"""
        try:
            # Find the main window
            self.main_window = self.window()
            
            # Create fullscreen window
            self.fullscreen_window = QWidget()
            self.fullscreen_window.setWindowTitle("SQL Editor - Fullscreen")
            self.fullscreen_window.setWindowFlags(Qt.WindowType.Window)
            
            # Set up fullscreen layout
            fullscreen_layout = QVBoxLayout(self.fullscreen_window)
            fullscreen_layout.setContentsMargins(10, 10, 10, 10)
            
            # Create fullscreen header
            header_widget = QWidget()
            header_layout = QHBoxLayout(header_widget)
            header_layout.setContentsMargins(0, 0, 0, 10)
            
            title_label = QLabel("SQL Editor - Fullscreen Mode")
            title_label.setFont(QFont("Segoe UI", 14, QFont.Weight.Bold))
            title_label.setStyleSheet(f"color: {ColorScheme.TEXT.name()};")
            
            exit_button = QPushButton("✕ Exit Fullscreen")
            exit_button.clicked.connect(self.exit_fullscreen)
            exit_button.setStyleSheet("""
                QPushButton {
                    background-color: #d32f2f;
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #b71c1c; }
            """)
            
            header_layout.addWidget(title_label)
            header_layout.addStretch()
            header_layout.addWidget(exit_button)
            
            # Move editor to fullscreen window
            self.editor.setParent(None)
            
            fullscreen_layout.addWidget(header_widget)
            fullscreen_layout.addWidget(self.editor)
            
            # Set dark theme for fullscreen window
            self.fullscreen_window.setStyleSheet(f"""
                QWidget {{
                    background-color: {ColorScheme.BACKGROUND.name()};
                    color: {ColorScheme.TEXT.name()};
                }}
            """)
            
            # Show fullscreen
            self.fullscreen_window.showFullScreen()
            self.editor.setFocus()
            
            # Update state
            self.is_fullscreen = True
            self.fullscreen_button.setText("⛶")
            self.fullscreen_button.setToolTip("Exit fullscreen editor (F11 or Esc)")
            
            # Install event filter for escape key
            self.fullscreen_window.installEventFilter(self)
            
        except Exception as e:
            print(f"Error entering fullscreen: {e}")
    
    def exit_fullscreen(self):
        """Exit fullscreen mode"""
        try:
            if hasattr(self, 'fullscreen_window') and self.fullscreen_window:
                # Move editor back to original container
                self.editor.setParent(None)
                self.editor_container_layout.addWidget(self.editor)
                
                # Close fullscreen window
                self.fullscreen_window.close()
                self.fullscreen_window = None
                
                # Update state
                self.is_fullscreen = False
                self.fullscreen_button.setText("⛶")
                self.fullscreen_button.setToolTip("Toggle fullscreen editor (F11)")
                
                # Focus back to editor
                self.editor.setFocus()
                
        except Exception as e:
            print(f"Error exiting fullscreen: {e}")
    
    def eventFilter(self, obj, event):
        """Handle events for fullscreen mode"""
        if (obj == self.fullscreen_window and 
            event.type() == event.Type.KeyPress):
            if event.key() == Qt.Key.Key_Escape:
                self.exit_fullscreen()
                return True
            elif event.key() == Qt.Key.Key_F11:
                self.exit_fullscreen()
                return True
        return super().eventFilter(obj, event)
    
    def keyPressEvent(self, event):
        """Handle key press events"""
        if event.key() == Qt.Key.Key_F11:
            self.toggle_fullscreen()
        else:
            super().keyPressEvent(event)
    
    def show_export_menu(self):
        """Show export options menu"""
        if not hasattr(self, 'model') or not self.model or self.model.rowCount() == 0:
            QMessageBox.information(self, "No Data", "No results to export.")
            return
        
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: #404040;
                color: #ffffff;
                border: 1px solid #606060;
                border-radius: 4px;
                padding: 4px;
            }
            QMenu::item {
                padding: 8px 16px;
                border-radius: 3px;
            }
            QMenu::item:selected {
                background-color: #0078d4;
            }
            QMenu::separator {
                height: 1px;
                background-color: #606060;
                margin: 4px 8px;
            }
        """)
        
        # Add export format actions
        csv_action = menu.addAction("📄 Export as CSV")
        csv_action.triggered.connect(lambda: self.export_data('csv'))
        
        excel_action = menu.addAction("📊 Export as Excel")
        excel_action.triggered.connect(lambda: self.export_data('excel'))
        
        json_action = menu.addAction("🔗 Export as JSON")
        json_action.triggered.connect(lambda: self.export_data('json'))
        
        parquet_action = menu.addAction("📦 Export as Parquet")
        parquet_action.triggered.connect(lambda: self.export_data('parquet'))
        
        menu.addSeparator()
        
        tsv_action = menu.addAction("📋 Export as TSV")
        tsv_action.triggered.connect(lambda: self.export_data('tsv'))
        
        html_action = menu.addAction("🌐 Export as HTML")
        html_action.triggered.connect(lambda: self.export_data('html'))
        
        xml_action = menu.addAction("📰 Export as XML")
        xml_action.triggered.connect(lambda: self.export_data('xml'))
        
        menu.addSeparator()
        
        clipboard_action = menu.addAction("📋 Copy to Clipboard")
        clipboard_action.triggered.connect(lambda: self.export_data('clipboard'))
        
        # Show menu at button position
        button_pos = self.export_button.mapToGlobal(self.export_button.rect().bottomLeft())
        menu.exec(button_pos)
    
    def export_data(self, format_type):
        """Export results data in the specified format"""
        try:
            if not hasattr(self, 'model') or not self.model or self.model.rowCount() == 0:
                QMessageBox.information(self, "No Data", "No results to export.")
                return
            
            # Get the dataframe from the model
            df = self.model._data.copy()
            
            if format_type == 'clipboard':
                # Copy to clipboard
                df.to_clipboard(index=False, sep='\t')
                QMessageBox.information(self, "Export Successful", 
                                      f"Results copied to clipboard!\n\n"
                                      f"📊 {len(df):,} rows × {len(df.columns)} columns")
                return
            
            # File export - get save location
            file_filters = {
                'csv': "CSV Files (*.csv)",
                'excel': "Excel Files (*.xlsx)",
                'json': "JSON Files (*.json)",
                'parquet': "Parquet Files (*.parquet)",
                'tsv': "TSV Files (*.tsv)",
                'html': "HTML Files (*.html)",
                'xml': "XML Files (*.xml)"
            }
            
            file_extensions = {
                'csv': '.csv',
                'excel': '.xlsx',
                'json': '.json',
                'parquet': '.parquet',
                'tsv': '.tsv',
                'html': '.html',
                'xml': '.xml'
            }
            
            default_name = f"query_results{file_extensions[format_type]}"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                f"Export Results as {format_type.upper()}",
                default_name,
                file_filters[format_type]
            )
            
            if not file_path:
                return  # User cancelled
            
            # Export based on format
            if format_type == 'csv':
                df.to_csv(file_path, index=False, encoding='utf-8')
                
            elif format_type == 'excel':
                if openpyxl is None:
                    QMessageBox.warning(self, "Excel Export Unavailable", 
                                      "Excel export requires the 'openpyxl' package.\n"
                                      "Please install it with: pip install openpyxl")
                    return
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Query Results', index=False)
                    
                    # Auto-adjust column widths
                    worksheet = writer.sheets['Query Results']
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
            elif format_type == 'json':
                df.to_json(file_path, orient='records', indent=2, date_format='iso')
                
            elif format_type == 'parquet':
                df.to_parquet(file_path, index=False)
                
            elif format_type == 'tsv':
                df.to_csv(file_path, index=False, sep='\t', encoding='utf-8')
                
            elif format_type == 'html':
                html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Query Results</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #f9f9f9; }}
        .info {{ margin-bottom: 20px; color: #666; }}
    </style>
</head>
<body>
    <h1>Query Results</h1>
    <div class="info">
        <p>Exported on: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p>Rows: {len(df):,} | Columns: {len(df.columns)}</p>
    </div>
    {df.to_html(index=False, escape=False, classes='results-table')}
</body>
</html>
                """
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(html_content)
                
            elif format_type == 'xml':
                xml_content = '<?xml version="1.0" encoding="UTF-8"?>\n'
                xml_content += '<query_results>\n'
                xml_content += f'  <metadata>\n'
                xml_content += f'    <export_date>{pd.Timestamp.now().isoformat()}</export_date>\n'
                xml_content += f'    <row_count>{len(df)}</row_count>\n'
                xml_content += f'    <column_count>{len(df.columns)}</column_count>\n'
                xml_content += f'  </metadata>\n'
                xml_content += '  <data>\n'
                
                for _, row in df.iterrows():
                    xml_content += '    <row>\n'
                    for col in df.columns:
                        # Clean column name for XML
                        clean_col = str(col).replace(' ', '_').replace('-', '_')
                        clean_col = ''.join(c for c in clean_col if c.isalnum() or c == '_')
                        value = str(row[col]) if pd.notna(row[col]) else ''
                        # Escape XML special characters
                        value = value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        xml_content += f'      <{clean_col}>{value}</{clean_col}>\n'
                    xml_content += '    </row>\n'
                
                xml_content += '  </data>\n'
                xml_content += '</query_results>'
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(xml_content)
            
            # Show success message
            file_size = os.path.getsize(file_path) / 1024  # Size in KB
            size_text = f"{file_size:.1f} KB" if file_size < 1024 else f"{file_size/1024:.1f} MB"
            
            QMessageBox.information(self, "Export Successful", 
                                  f"Results exported successfully!\n\n"
                                  f"📁 File: {os.path.basename(file_path)}\n"
                                  f"📊 Data: {len(df):,} rows × {len(df.columns)} columns\n"
                                  f"💾 Size: {size_text}\n"
                                  f"📍 Location: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", 
                               f"Failed to export results:\n\n{str(e)}")
            print(f"Export error: {e}")

# Database schema browser
class SchemaBrowser(QTreeWidget):
    # Signal to notify when schema data is updated (table_names, column_names)
    schema_data_updated = pyqtSignal(list, list)
    # Signal to notify when schema structure changes (table/column deleted)
    schema_changed = pyqtSignal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setHeaderLabels(["Database Objects"])
        self.setColumnCount(1)
        self.setAlternatingRowColors(True)
        self.setAnimated(True)
        
        # Store schema data for auto-completion
        self.table_names = []
        self.column_names = []
        
        # Store connection references for context menu operations
        self.connection = None
        self.connection_info = None
        
        # Enable context menu
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)
        
        # Set colors
        palette = self.palette()
        palette.setColor(QPalette.ColorRole.Base, ColorScheme.SIDEBAR_BG)
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor(40, 40, 40))
        palette.setColor(QPalette.ColorRole.Text, ColorScheme.TEXT)
        self.setPalette(palette)
        
        # Icons
        self.table_icon = qta.icon('fa5s.table', color=ColorScheme.ACCENT)
        self.view_icon = qta.icon('fa5s.eye', color=ColorScheme.ACCENT)
        self.index_icon = qta.icon('fa5s.key', color=ColorScheme.ACCENT)
        self.column_icon = qta.icon('fa5s.columns', color=ColorScheme.TEXT)
        self.pk_icon = qta.icon('fa5s.key', color=ColorScheme.SUCCESS)
        self.fk_icon = qta.icon('fa5s.link', color=ColorScheme.WARNING)
    
    def load_schema(self, connection, connection_info):
        """Load database schema with enhanced connection validation and recovery"""
        # Store connection references
        self.connection = connection
        self.connection_info = connection_info
        
        # Enhanced connection validation with multiple fallback strategies
        max_attempts = 3
        attempt = 0
        
        while attempt < max_attempts:
            try:
                # Test connection before proceeding
                test_result = connection.execute("SELECT 1 AS test").fetchone()
                if not test_result:
                    raise Exception("Connection test returned no result")
                
                # Connection is valid, proceed with schema loading
                break
                
            except Exception as e:
                print(f"Schema browser connection validation failed (attempt {attempt + 1}): {e}")
                attempt += 1
                
                if attempt < max_attempts:
                    # Try to get a fresh connection from the main app
                    try:
                        # Get the main app reference
                        main_app = self.parent()
                        while main_app and not hasattr(main_app, 'current_connection'):
                            main_app = main_app.parent()
                        
                        if main_app and hasattr(main_app, 'reconnect_current_database'):
                            print(f"Attempting to reconnect via main app...")
                            if main_app.reconnect_current_database():
                                connection = main_app.current_connection
                                connection_info = main_app.current_connection_info
                                self.connection = connection
                                self.connection_info = connection_info
                                print("Successfully got fresh connection from main app")
                                continue
                            else:
                                print("Main app reconnection failed")
                        
                        # Fallback: try to create a new connection directly
                        if connection_info and 'type' in connection_info:
                            db_type = connection_info['type'].lower()
                            file_path = connection_info.get('file_path') or connection_info.get('path')
                            
                            if file_path and os.path.exists(file_path):
                                if db_type in ['sqlite', 'sqlite3']:
                                    connection = sqlite3.connect(file_path)
                                elif db_type == 'duckdb':
                                    connection = duckdb.connect(file_path)
                                else:
                                    raise ValueError(f"Unsupported database type: {db_type}")
                                
                                self.connection = connection
                                print(f"Created new {db_type} connection directly")
                                continue
                            else:
                                print(f"Database file not found or no path: {file_path}")
                        
                    except Exception as conn_e:
                        print(f"Connection recovery failed: {conn_e}")
                
                # If this is the last attempt, clear the schema and return
                if attempt >= max_attempts:
                    print("All connection attempts failed, clearing schema browser")
                    self.clear()
                    return
        
        # Clear existing items
        self.clear()
        
        # Reset schema data collections
        self.table_names = []
        self.column_names = []
        
        # Load schema based on database type with error handling
        try:
            db_type = connection_info["type"].lower()
            if db_type in ["sqlite", "sqlite3"]:
                self.load_sqlite_schema(connection)
            elif db_type == "duckdb":
                self.load_duckdb_schema(connection)
            else:
                print(f"Unknown database type: {connection_info['type']}")
                return
                
        except Exception as e:
            print(f"Schema loading failed after successful connection: {e}")
            # Try one more time with a fresh connection
            try:
                main_app = self.parent()
                while main_app and not hasattr(main_app, 'current_connection'):
                    main_app = main_app.parent()
                
                if main_app and hasattr(main_app, 'reconnect_current_database'):
                    if main_app.reconnect_current_database():
                        fresh_connection = main_app.current_connection
                        if connection_info['type'].lower() in ['sqlite', 'sqlite3']:
                            self.load_sqlite_schema(fresh_connection)
                        elif connection_info['type'].lower() == 'duckdb':
                            self.load_duckdb_schema(fresh_connection)
            except Exception as retry_e:
                print(f"Schema loading retry also failed: {retry_e}")
                self.clear()
                return
    
    def load_sqlite_schema(self, connection):
        # Clear existing schema data
        self.table_names = []
        self.column_names = []
        
        # Create root item for database
        db_item = QTreeWidgetItem(self, [os.path.basename(connection.database)])
        db_item.setIcon(0, qta.icon('fa5s.database', color=ColorScheme.ACCENT))
        
        # Get list of tables
        cursor = connection.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = cursor.fetchall()
        
        # Tables group
        tables_item = QTreeWidgetItem(db_item, ["Tables"])
        tables_item.setIcon(0, qta.icon('fa5s.folder', color=ColorScheme.ACCENT))
        
        if not tables:
            # Show "No tables" message when database is empty
            no_tables_item = QTreeWidgetItem(tables_item, ["(No tables)"])
            no_tables_item.setIcon(0, qta.icon('fa5s.info-circle', color=ColorScheme.COMMENT))
        else:
            for table in tables:
                table_name = table[0]
                self.table_names.append(table_name)  # Collect table name for auto-completion
                
                table_item = QTreeWidgetItem(tables_item, [table_name])
                table_item.setIcon(0, self.table_icon)
                # Store metadata for context menu
                table_item.setData(0, Qt.ItemDataRole.UserRole, {'type': 'table', 'name': table_name})
                
                # Get columns for this table
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns = cursor.fetchall()
                
                for column in columns:
                    col_name = column[1]
                    col_type = column[2]
                    is_pk = column[5] == 1  # Primary key flag
                    
                    # Collect column name for auto-completion (avoid duplicates)
                    if col_name not in self.column_names:
                        self.column_names.append(col_name)
                    
                    column_text = f"{col_name} ({col_type})"
                    column_item = QTreeWidgetItem(table_item, [column_text])
                    column_item.setIcon(0, self.pk_icon if is_pk else self.column_icon)
                    # Store metadata for context menu
                    column_item.setData(0, Qt.ItemDataRole.UserRole, {'type': 'column', 'name': col_name, 'table': table_name, 'is_pk': is_pk})
        
        # Views group
        cursor.execute("SELECT name FROM sqlite_master WHERE type='view' ORDER BY name")
        views = cursor.fetchall()
        
        if views:
            views_item = QTreeWidgetItem(db_item, ["Views"])
            views_item.setIcon(0, qta.icon('fa5s.folder', color=ColorScheme.ACCENT))
            
            for view in views:
                view_name = view[0]
                self.table_names.append(view_name)  # Views can also be queried like tables
                view_item = QTreeWidgetItem(views_item, [view_name])
                view_item.setIcon(0, self.view_icon)
                # Store metadata for context menu
                view_item.setData(0, Qt.ItemDataRole.UserRole, {'type': 'view', 'name': view_name})
        
        # Emit signal with collected schema data
        self.schema_data_updated.emit(self.table_names, self.column_names)
        
        # Indexes group
        cursor.execute("SELECT name, tbl_name FROM sqlite_master WHERE type='index' AND name NOT LIKE 'sqlite_%' ORDER BY name")
        indexes = cursor.fetchall()
        
        if indexes:
            indexes_item = QTreeWidgetItem(db_item, ["Indexes"])
            indexes_item.setIcon(0, qta.icon('fa5s.folder', color=ColorScheme.ACCENT))
            
            for index in indexes:
                index_name = index[0]
                table_name = index[1]
                index_item = QTreeWidgetItem(indexes_item, [f"{index_name} (on {table_name})"])
                index_item.setIcon(0, self.index_icon)
        
        # Expand the database item
        db_item.setExpanded(True)
        tables_item.setExpanded(True)
    
    def load_duckdb_schema(self, connection):
        # Clear existing schema data
        self.table_names = []
        self.column_names = []
        
        # Create root item for database
        db_item = QTreeWidgetItem(self, ["DuckDB Database"])
        db_item.setIcon(0, qta.icon('fa5s.database', color=ColorScheme.ACCENT))
        
        # Get list of tables
        tables_df = connection.execute("SHOW TABLES").fetchdf()
        
        # Tables group
        tables_item = QTreeWidgetItem(db_item, ["Tables"])
        tables_item.setIcon(0, qta.icon('fa5s.folder', color=ColorScheme.ACCENT))
        
        if tables_df.empty:
            # Show "No tables" message when database is empty
            no_tables_item = QTreeWidgetItem(tables_item, ["(No tables)"])
            no_tables_item.setIcon(0, qta.icon('fa5s.info-circle', color=ColorScheme.COMMENT))
        else:
            for _, row in tables_df.iterrows():
                table_name = row['name']
                self.table_names.append(table_name)  # Collect table name for auto-completion
                
                table_item = QTreeWidgetItem(tables_item, [table_name])
                table_item.setIcon(0, self.table_icon)
                # Store metadata for context menu
                table_item.setData(0, Qt.ItemDataRole.UserRole, {'type': 'table', 'name': table_name})
                
                # Get columns for this table with robust connection handling
                try:
                    # Validate connection first
                    try:
                        connection.execute("SELECT 1").fetchone()
                    except Exception:
                        print(f"Connection issue detected for table {table_name}, skipping column loading")
                        continue
                    
                    # Try multiple approaches for getting column info
                    columns_data = None
                    
                    # Approach 1: Use PRAGMA table_info
                    try:
                        columns_df = connection.execute(f"PRAGMA table_info('{table_name}')").fetchdf()
                        if not columns_df.empty:
                            columns_data = [(row['name'], row['type'], row['pk'] == 1) for _, row in columns_df.iterrows()]
                    except Exception as e1:
                        print(f"PRAGMA table_info failed for {table_name}: {e1}")
                    
                    # Approach 2: Use information_schema if PRAGMA failed
                    if columns_data is None:
                        try:
                            result = connection.execute(f"""
                                SELECT column_name, data_type, 
                                       CASE WHEN is_nullable = 'NO' THEN 1 ELSE 0 END as is_pk
                                FROM information_schema.columns 
                                WHERE table_name = '{table_name}' AND table_schema = 'main'
                                ORDER BY ordinal_position
                            """).fetchall()
                            if result:
                                columns_data = [(row[0], row[1], row[2] == 1) for row in result]
                        except Exception as e2:
                            print(f"Information schema query failed for {table_name}: {e2}")
                    
                    # Approach 3: Try DESCRIBE as fallback
                    if columns_data is None:
                        try:
                            result = connection.execute(f"DESCRIBE {table_name}").fetchall()
                            if result:
                                columns_data = [(row[0], row[1], False) for row in result]
                        except Exception as e3:
                            print(f"DESCRIBE failed for {table_name}: {e3}")
                    
                    # Process columns if we got any data
                    if columns_data:
                        for col_name, col_type, is_pk in columns_data:
                            try:
                                # Collect column name for auto-completion (avoid duplicates)
                                if col_name not in self.column_names:
                                    self.column_names.append(col_name)
                                
                                column_text = f"{col_name} ({col_type})"
                                column_item = QTreeWidgetItem(table_item, [column_text])
                                column_item.setIcon(0, self.pk_icon if is_pk else self.column_icon)
                                # Store metadata for context menu
                                column_item.setData(0, Qt.ItemDataRole.UserRole, {'type': 'column', 'name': col_name, 'table': table_name, 'is_pk': is_pk})
                            except Exception as e:
                                print(f"Error processing column {col_name} for table {table_name}: {e}")
                                continue
                    else:
                        print(f"Could not load any column information for table {table_name}")
                        
                except Exception as e:
                    print(f"Error loading columns for table {table_name}: {e}")
                    # Continue processing other tables even if this one fails
        
        # Views group - DuckDB also supports views
        try:
            views_df = connection.execute("SELECT view_name FROM information_schema.views").fetchdf()
            
            if not views_df.empty:
                views_item = QTreeWidgetItem(db_item, ["Views"])
                views_item.setIcon(0, qta.icon('fa5s.folder', color=ColorScheme.ACCENT))
                
                for _, row in views_df.iterrows():
                    view_name = row['view_name']
                    self.table_names.append(view_name)  # Views can also be queried like tables
                    view_item = QTreeWidgetItem(views_item, [view_name])
                    view_item.setIcon(0, self.view_icon)
                    # Store metadata for context menu
                    view_item.setData(0, Qt.ItemDataRole.UserRole, {'type': 'view', 'name': view_name})
        except:
            # Some versions of DuckDB might not have this view
            pass
        
        # Emit signal with collected schema data
        self.schema_data_updated.emit(self.table_names, self.column_names)
        
        # Expand the database item
        db_item.setExpanded(True)
        tables_item.setExpanded(True)
    
    def show_context_menu(self, position):
        """Show context menu for the schema browser"""
        item = self.itemAt(position)
        if not item:
            return
        
        # Get item metadata
        item_data = item.data(0, Qt.ItemDataRole.UserRole)
        if not item_data:
            return
        
        # Create context menu
        menu = QMenu(self)
        
        # Add actions based on item type
        if item_data['type'] == 'table':
            delete_action = QAction(qta.icon('fa5s.trash', color=ColorScheme.ERROR), f"Delete Table '{item_data['name']}'", self)
            delete_action.triggered.connect(lambda: self.delete_table(item_data['name']))
            menu.addAction(delete_action)
            
        elif item_data['type'] == 'view':
            delete_action = QAction(qta.icon('fa5s.trash', color=ColorScheme.ERROR), f"Delete View '{item_data['name']}'", self)
            delete_action.triggered.connect(lambda: self.delete_view(item_data['name']))
            menu.addAction(delete_action)
            
        elif item_data['type'] == 'column':
            if not item_data.get('is_pk', False):  # Don't allow deleting primary keys
                delete_action = QAction(qta.icon('fa5s.trash', color=ColorScheme.ERROR), f"Delete Column '{item_data['name']}'", self)
                delete_action.triggered.connect(lambda: self.delete_column(item_data['table'], item_data['name']))
                menu.addAction(delete_action)
            else:
                info_action = QAction(qta.icon('fa5s.info-circle', color=ColorScheme.WARNING), "Cannot delete primary key column", self)
                info_action.setEnabled(False)
                menu.addAction(info_action)
        
        # Show the menu
        if menu.actions():
            menu.exec(self.mapToGlobal(position))
    
    def delete_table(self, table_name):
        """Delete a table with confirmation"""
        if not self.connection or not self.connection_info:
            QMessageBox.warning(self, "Error", "No database connection available.")
            return
        
        # Confirmation dialog
        reply = QMessageBox.question(
            self, 
            "Confirm Delete", 
            f"Are you sure you want to delete the table '{table_name}'?\n\nThis action cannot be undone and will permanently delete all data in the table.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Execute DROP TABLE with multiple fallback strategies
                success = False
                error_messages = []
                
                # Strategy 1: Try with double quotes
                try:
                    if self.connection_info['type'].lower() in ['sqlite', 'sqlite3']:
                        cursor = self.connection.cursor()
                        cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                        self.connection.commit()
                    else:  # DuckDB
                        self.connection.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                    success = True
                except Exception as e1:
                    error_messages.append(f"Double quotes: {str(e1)}")
                
                # Strategy 2: Try without quotes if first attempt failed
                if not success:
                    try:
                        if self.connection_info['type'].lower() in ['sqlite', 'sqlite3']:
                            cursor = self.connection.cursor()
                            cursor.execute(f'DROP TABLE IF EXISTS {table_name}')
                            self.connection.commit()
                        else:  # DuckDB
                            self.connection.execute(f'DROP TABLE IF EXISTS {table_name}')
                        success = True
                    except Exception as e2:
                        error_messages.append(f"No quotes: {str(e2)}")
                
                # Strategy 3: Try with square brackets (for some SQL dialects)
                if not success:
                    try:
                        if self.connection_info['type'].lower() in ['sqlite', 'sqlite3']:
                            cursor = self.connection.cursor()
                            cursor.execute(f'DROP TABLE IF EXISTS [{table_name}]')
                            self.connection.commit()
                        else:  # DuckDB
                            self.connection.execute(f'DROP TABLE IF EXISTS [{table_name}]')
                        success = True
                    except Exception as e3:
                        error_messages.append(f"Square brackets: {str(e3)}")
                
                if success:
                    # Refresh schema browser
                    self.load_schema(self.connection, self.connection_info)
                    
                    # Emit schema changed signal
                    self.schema_changed.emit()
                    
                    # Show success message
                    QMessageBox.information(self, "Success", f"Table '{table_name}' has been deleted successfully.")
                else:
                    # All strategies failed
                    error_detail = "\n".join(error_messages)
                    QMessageBox.critical(self, "Error", f"Failed to delete table '{table_name}' using all strategies:\n\n{error_detail}")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Unexpected error while deleting table '{table_name}':\n{str(e)}")
    
    def delete_view(self, view_name):
        """Delete a view with confirmation"""
        if not self.connection or not self.connection_info:
            QMessageBox.warning(self, "Error", "No database connection available.")
            return
        
        # Confirmation dialog
        reply = QMessageBox.question(
            self, 
            "Confirm Delete", 
            f"Are you sure you want to delete the view '{view_name}'?\n\nThis action cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Execute DROP VIEW with multiple fallback strategies
                success = False
                error_messages = []
                
                # Strategy 1: Try with double quotes
                try:
                    if self.connection_info['type'].lower() in ['sqlite', 'sqlite3']:
                        cursor = self.connection.cursor()
                        cursor.execute(f'DROP VIEW IF EXISTS "{view_name}"')
                        self.connection.commit()
                    else:  # DuckDB
                        self.connection.execute(f'DROP VIEW IF EXISTS "{view_name}"')
                    success = True
                except Exception as e1:
                    error_messages.append(f"Double quotes: {str(e1)}")
                
                # Strategy 2: Try without quotes if first attempt failed
                if not success:
                    try:
                        if self.connection_info['type'].lower() in ['sqlite', 'sqlite3']:
                            cursor = self.connection.cursor()
                            cursor.execute(f'DROP VIEW IF EXISTS {view_name}')
                            self.connection.commit()
                        else:  # DuckDB
                            self.connection.execute(f'DROP VIEW IF EXISTS {view_name}')
                        success = True
                    except Exception as e2:
                        error_messages.append(f"No quotes: {str(e2)}")
                
                # Strategy 3: Try with square brackets
                if not success:
                    try:
                        if self.connection_info['type'].lower() in ['sqlite', 'sqlite3']:
                            cursor = self.connection.cursor()
                            cursor.execute(f'DROP VIEW IF EXISTS [{view_name}]')
                            self.connection.commit()
                        else:  # DuckDB
                            self.connection.execute(f'DROP VIEW IF EXISTS [{view_name}]')
                        success = True
                    except Exception as e3:
                        error_messages.append(f"Square brackets: {str(e3)}")
                
                if success:
                    # Refresh schema browser
                    self.load_schema(self.connection, self.connection_info)
                    
                    # Emit schema changed signal
                    self.schema_changed.emit()
                    
                    # Show success message
                    QMessageBox.information(self, "Success", f"View '{view_name}' has been deleted successfully.")
                else:
                    # All strategies failed
                    error_detail = "\n".join(error_messages)
                    QMessageBox.critical(self, "Error", f"Failed to delete view '{view_name}' using all strategies:\n\n{error_detail}")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Unexpected error while deleting view '{view_name}':\n{str(e)}")
    
    def delete_column(self, table_name, column_name):
        """Delete a column with confirmation"""
        if not self.connection or not self.connection_info:
            QMessageBox.warning(self, "Error", "No database connection available.")
            return
        
        # Confirmation dialog
        reply = QMessageBox.question(
            self, 
            "Confirm Delete", 
            f"Are you sure you want to delete the column '{column_name}' from table '{table_name}'?\n\nThis action cannot be undone and will permanently delete all data in this column.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                # Enhanced ALTER TABLE support with better error handling
                success = False
                error_messages = []
                
                # First, check if the column exists
                try:
                    if self.connection_info['type'].lower() in ['sqlite', 'sqlite3']:
                        cursor = self.connection.cursor()
                        cursor.execute(f"PRAGMA table_info({table_name})")
                        columns = [row[1] for row in cursor.fetchall()]
                    else:  # DuckDB
                        result = self.connection.execute(f"DESCRIBE {table_name}").fetchall()
                        columns = [row[0] for row in result]
                    
                    if column_name not in columns:
                        QMessageBox.warning(self, "Column Not Found", f"Column '{column_name}' does not exist in table '{table_name}'.")
                        return
                        
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Failed to verify column existence: {str(e)}")
                    return
                
                if self.connection_info['type'].lower() in ['sqlite', 'sqlite3']:
                    # For SQLite, we need to recreate the table without the column
                    cursor = self.connection.cursor()
                    
                    # Get table schema with multiple strategies
                    columns = None
                    for quote_style in ['"', '', '`']:
                        try:
                            if quote_style:
                                cursor.execute(f'PRAGMA table_info({quote_style}{table_name}{quote_style})')
                            else:
                                cursor.execute(f'PRAGMA table_info({table_name})')
                            columns = cursor.fetchall()
                            break
                        except:
                            continue
                    
                    if not columns:
                        QMessageBox.critical(self, "Error", f"Could not retrieve table schema for '{table_name}'.")
                        return
                    
                    # Create new column list without the deleted column
                    remaining_columns = [col[1] for col in columns if col[1] != column_name]
                    
                    if len(remaining_columns) == 0:
                        QMessageBox.warning(self, "Error", "Cannot delete the last column of a table.")
                        return
                    
                    # Try multiple quoting strategies for SQLite
                    for quote_style in ['"', '', '`']:
                        try:
                            # Start transaction
                            cursor.execute("BEGIN TRANSACTION")
                            
                            try:
                                # Create temporary table with remaining columns
                                if quote_style:
                                    column_list = ", ".join([f'{quote_style}{col}{quote_style}' for col in remaining_columns])
                                    cursor.execute(f'CREATE TABLE {quote_style}{table_name}_temp{quote_style} AS SELECT {column_list} FROM {quote_style}{table_name}{quote_style}')
                                    
                                    # Drop original table
                                    cursor.execute(f'DROP TABLE {quote_style}{table_name}{quote_style}')
                                    
                                    # Rename temp table to original name
                                    cursor.execute(f'ALTER TABLE {quote_style}{table_name}_temp{quote_style} RENAME TO {quote_style}{table_name}{quote_style}')
                                else:
                                    column_list = ", ".join(remaining_columns)
                                    cursor.execute(f'CREATE TABLE {table_name}_temp AS SELECT {column_list} FROM {table_name}')
                                    
                                    # Drop original table
                                    cursor.execute(f'DROP TABLE {table_name}')
                                    
                                    # Rename temp table to original name
                                    cursor.execute(f'ALTER TABLE {table_name}_temp RENAME TO {table_name}')
                                
                                # Commit transaction
                                cursor.execute("COMMIT")
                                success = True
                                break
                                
                            except Exception as e:
                                # Rollback on error
                                cursor.execute("ROLLBACK")
                                error_messages.append(f"SQLite with {quote_style or 'no'} quotes: {str(e)}")
                                
                        except Exception as e:
                            error_messages.append(f"SQLite transaction with {quote_style or 'no'} quotes: {str(e)}")
                        
                else:  # DuckDB
                    # Try multiple quoting strategies for DuckDB
                    for quote_style in ['"', '', '`']:
                        try:
                            if quote_style:
                                self.connection.execute(f'ALTER TABLE {quote_style}{table_name}{quote_style} DROP COLUMN {quote_style}{column_name}{quote_style}')
                            else:
                                self.connection.execute(f'ALTER TABLE {table_name} DROP COLUMN {column_name}')
                            success = True
                            break
                        except Exception as e:
                            error_messages.append(f"DuckDB with {quote_style or 'no'} quotes: {str(e)}")
                
                if not success:
                    error_detail = "\n".join(error_messages)
                    QMessageBox.critical(self, "Error", f"Failed to delete column '{column_name}' from table '{table_name}' using all strategies:\n\n{error_detail}")
                    return
                
                # Only refresh if successful
                if success:
                    # Refresh schema browser
                    self.load_schema(self.connection, self.connection_info)
                    
                    # Emit schema changed signal
                    self.schema_changed.emit()
                    
                    # Show success message
                    QMessageBox.information(self, "Success", f"Column '{column_name}' has been deleted from table '{table_name}' successfully.")
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Unexpected error while deleting column '{column_name}' from table '{table_name}':\n{str(e)}")

# Main application window
class SQLEditorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SQL Editor")
        self.resize(1200, 800)
        self.connections = {}
        self.current_connection = None
        self.current_connection_info = None
        
        # Set application style
        self.setup_style()
        
        # Create central widget and layout
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QVBoxLayout(self.central_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Create main splitter
        self.main_splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Create schema browser
        self.schema_browser = SchemaBrowser()
        self.schema_browser.setMinimumWidth(250)
        
        # Connect schema browser signals
        self.schema_browser.schema_data_updated.connect(self.update_all_tabs_completions)
        self.schema_browser.schema_changed.connect(self.refresh_schema_browser)
        
        # Create tab widget for query editors
        self.tab_widget = QTabWidget()
        self.tab_widget.setTabsClosable(True)
        self.tab_widget.tabCloseRequested.connect(self.close_tab)
        
        # Add widgets to splitter
        self.main_splitter.addWidget(self.schema_browser)
        self.main_splitter.addWidget(self.tab_widget)
        self.main_splitter.setSizes([250, 950])
        
        # Add splitter to main layout
        self.main_layout.addWidget(self.main_splitter)
        
        # Create toolbar
        self.create_toolbar()
        
        # Create status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.connection_label = QLabel("Not connected")
        self.status_bar.addPermanentWidget(self.connection_label)
        
        # Create actions
        self.create_actions()
        
        # Create menus
        self.create_menus()
        
        # Add initial tab
        self.add_tab()
        
        # Auto-connect to main DuckDB database
        self.auto_connect_main_database()
        
        # Setup schema monitoring timer
        self.schema_timer = QTimer()
        self.schema_timer.timeout.connect(self.check_schema_changes)
        self.schema_timer.start(2000)  # Check every 2 seconds
        self.last_schema_hash = None
    
    def auto_connect_main_database(self):
        """Automatically connect to main.duckdb in the project folder"""
        try:
            # Get the current directory (where the SQL editor is running)
            current_dir = os.path.dirname(os.path.abspath(__file__))
            db_path = os.path.join(current_dir, "main.duckdb")
            
            # Create the database if it doesn't exist
            if not os.path.exists(db_path):
                # Create an empty DuckDB database
                temp_conn = duckdb.connect(db_path)
                temp_conn.close()
                self.status_bar.showMessage(f"Created new database: {db_path}")
            
            # Connection info for the main database
            connection_info = {
                'type': 'duckdb',
                'path': db_path,
                'name': 'main.duckdb'
            }
            
            # Connect to the database
            self.connect_to_database(connection_info)
            
            # Force an immediate schema refresh after connection
            QTimer.singleShot(100, self.refresh_schema_browser)
            
            self.status_bar.showMessage(f"Auto-connected to main.duckdb")
            
        except Exception as e:
            self.status_bar.showMessage(f"Failed to auto-connect: {str(e)}")
            print(f"Auto-connect error: {e}")
    
    def refresh_schema_browser(self):
        """Refresh the schema browser with comprehensive connection validation and recovery"""
        if not self.current_connection_info:
            print("No connection info available for schema refresh")
            return
            
        max_attempts = 3
        attempt = 0
        
        while attempt < max_attempts:
            try:
                # Validate current connection if it exists
                if self.current_connection:
                    try:
                        # Test the connection with a simple query
                        test_result = self.current_connection.execute("SELECT 1 AS test").fetchone()
                        if test_result:
                            # Connection is valid, proceed with schema loading
                            self.schema_browser.load_schema(self.current_connection, self.current_connection_info)
                            return
                    except Exception as e:
                        print(f"Current connection invalid (attempt {attempt + 1}): {e}")
                
                # Current connection is invalid, try to reconnect
                print(f"Attempting to reconnect (attempt {attempt + 1})...")
                if self.reconnect_current_database():
                    # Reconnection successful, try to load schema
                    self.schema_browser.load_schema(self.current_connection, self.current_connection_info)
                    return
                else:
                    print(f"Reconnection failed on attempt {attempt + 1}")
                
            except Exception as e:
                print(f"Schema refresh failed on attempt {attempt + 1}: {e}")
            
            attempt += 1
            time.sleep(0.5)  # Brief delay between attempts
        
        # All attempts failed
        print("All schema refresh attempts failed, clearing schema browser")
        self.schema_browser.clear()
        
        # Show a user-friendly message
        self.status_bar.showMessage("Connection lost - please reconnect to database", 5000)
    
    def reconnect_current_database(self):
        """Reconnect to the current database with enhanced error handling and reference updates"""
        if not self.current_connection_info:
            print("No connection info available for reconnection")
            return False
        
        try:
            # Close existing connection safely
            if self.current_connection:
                try:
                    self.current_connection.close()
                except Exception as e:
                    print(f"Error closing old connection (expected): {e}")
                finally:
                    self.current_connection = None
            
            # Extract connection details
            db_type = self.current_connection_info["type"]
            file_path = self.current_connection_info.get("file_path") or self.current_connection_info.get("path")
            
            if not file_path:
                print("No database file path available for reconnection")
                return False
            
            if not os.path.exists(file_path):
                print(f"Database file does not exist: {file_path}")
                return False
            
            # Create new connection based on database type
            print(f"Reconnecting to {db_type} database: {file_path}")
            
            if db_type.lower() in ["sqlite", "sqlite3"]:
                self.current_connection = sqlite3.connect(file_path)
            elif db_type.lower() == "duckdb":
                self.current_connection = duckdb.connect(file_path)
            else:
                print(f"Unsupported database type: {db_type}")
                return False
            
            # Test the new connection
            test_result = self.current_connection.execute("SELECT 1 AS test").fetchone()
            if not test_result:
                raise Exception("New connection test failed")
            
            # Update connection cache
            connection_key = f"{db_type}:{file_path}"
            self.connections[connection_key] = self.current_connection
            
            # Update all tabs with the new connection
            for i in range(self.tab_widget.count()):
                tab = self.tab_widget.widget(i)
                if hasattr(tab, 'connection'):
                    tab.connection = self.current_connection
                if hasattr(tab, 'connection_info'):
                    tab.connection_info = self.current_connection_info
            
            # Update window title
            self.setWindowTitle(f"SQL Editor - {os.path.basename(file_path)}")
            
            # Update status
            self.status_bar.showMessage(f"Successfully reconnected to {os.path.basename(file_path)}", 3000)
            
            print(f"Successfully reconnected to {file_path}")
            return True
            
        except Exception as e:
            print(f"Failed to reconnect to database: {e}")
            self.current_connection = None
            self.status_bar.showMessage(f"Reconnection failed: {str(e)}", 5000)
            return False
            
    def update_all_tabs_completions(self, table_names, column_names):
        """Update auto-completion for all query tabs when schema changes"""
        for i in range(self.tab_widget.count()):
            tab = self.tab_widget.widget(i)
            if isinstance(tab, QueryTab):
                tab.update_schema_completions(table_names, column_names)
    
    def check_schema_changes(self):
        """Check for schema changes and refresh if needed"""
        if not self.current_connection or not self.current_connection_info:
            return
        
        try:
            # Get current schema hash
            if self.current_connection_info['type'].lower() == 'duckdb':
                # Get table list for DuckDB
                result = self.current_connection.execute("SELECT table_name FROM information_schema.tables WHERE table_schema = 'main'").fetchall()
                schema_data = str(sorted([row[0] for row in result]))
            else:
                # Get table list for SQLite
                cursor = self.current_connection.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                schema_data = str(sorted([row[0] for row in cursor.fetchall()]))
            
            # Calculate hash
            import hashlib
            current_hash = hashlib.md5(schema_data.encode()).hexdigest()
            
            # Check if schema changed
            if self.last_schema_hash is not None and current_hash != self.last_schema_hash:
                self.refresh_schema_browser()
            
            self.last_schema_hash = current_hash
            
        except Exception as e:
            # Silently ignore errors in schema monitoring
            pass
    
    def setup_style(self):
        # Set dark theme
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, ColorScheme.BACKGROUND)
        palette.setColor(QPalette.ColorRole.WindowText, ColorScheme.TEXT)
        palette.setColor(QPalette.ColorRole.Base, ColorScheme.BACKGROUND)
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
        palette.setColor(QPalette.ColorRole.ToolTipBase, ColorScheme.BACKGROUND)
        palette.setColor(QPalette.ColorRole.ToolTipText, ColorScheme.TEXT)
        palette.setColor(QPalette.ColorRole.Text, ColorScheme.TEXT)
        palette.setColor(QPalette.ColorRole.Button, ColorScheme.BACKGROUND)
        palette.setColor(QPalette.ColorRole.ButtonText, ColorScheme.TEXT)
        palette.setColor(QPalette.ColorRole.BrightText, Qt.GlobalColor.red)
        palette.setColor(QPalette.ColorRole.Link, ColorScheme.ACCENT)
        palette.setColor(QPalette.ColorRole.Highlight, ColorScheme.HIGHLIGHT)
        palette.setColor(QPalette.ColorRole.HighlightedText, Qt.GlobalColor.black)
        
        QApplication.setPalette(palette)
    
    def create_toolbar(self):
        self.toolbar = QToolBar("Main Toolbar")
        self.toolbar.setIconSize(QSize(20, 20))
        self.toolbar.setMovable(False)
        self.addToolBar(self.toolbar)
        
        # Create Database button
        self.create_db_button = QToolButton()
        self.create_db_button.setIcon(qta.icon('fa5s.plus-circle', color=ColorScheme.SUCCESS))
        self.create_db_button.setText("Create DB")
        self.create_db_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.create_db_button.clicked.connect(self.show_create_database_dialog)
        self.toolbar.addWidget(self.create_db_button)
        
        # Connect button
        self.connect_button = QToolButton()
        self.connect_button.setIcon(qta.icon('fa5s.plug', color=ColorScheme.TEXT))
        self.connect_button.setText("Connect")
        self.connect_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.connect_button.clicked.connect(self.show_connection_dialog)
        self.toolbar.addWidget(self.connect_button)
        
        # Reconnect to Main DB button
        self.reconnect_main_button = QToolButton()
        self.reconnect_main_button.setIcon(qta.icon('fa5s.home', color=ColorScheme.ACCENT))
        self.reconnect_main_button.setText("Main DB")
        self.reconnect_main_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.reconnect_main_button.setToolTip("Reconnect to main.duckdb")
        self.reconnect_main_button.clicked.connect(self.auto_connect_main_database)
        self.toolbar.addWidget(self.reconnect_main_button)
        
        self.toolbar.addSeparator()
        
        # Execute button
        self.execute_button = QToolButton()
        self.execute_button.setIcon(qta.icon('fa5s.play', color=ColorScheme.SUCCESS))
        self.execute_button.setText("Execute")
        self.execute_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.execute_button.setToolTip("Execute entire query (F5)")
        self.execute_button.clicked.connect(self.execute_current_query)
        self.toolbar.addWidget(self.execute_button)
        
        # Execute Selection button
        self.execute_selection_button = QToolButton()
        self.execute_selection_button.setIcon(qta.icon('fa5s.play-circle', color=ColorScheme.ACCENT))
        self.execute_selection_button.setText("Execute Selection")
        self.execute_selection_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.execute_selection_button.setToolTip("Execute selected text only (Ctrl+E)")
        self.execute_selection_button.clicked.connect(self.execute_selected_query)
        self.toolbar.addWidget(self.execute_selection_button)
        
        # Export Results button
        self.export_results_button = QToolButton()
        self.export_results_button.setIcon(qta.icon('fa5s.download', color='#2196F3'))
        self.export_results_button.setText("Export Results")
        self.export_results_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.export_results_button.setToolTip("Export query results (Ctrl+Shift+E)")
        self.export_results_button.clicked.connect(self.export_current_results)
        self.toolbar.addWidget(self.export_results_button)
        
        # New tab button
        self.new_tab_button = QToolButton()
        self.new_tab_button.setIcon(qta.icon('fa5s.plus', color=ColorScheme.TEXT))
        self.new_tab_button.setText("New Tab")
        self.new_tab_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.new_tab_button.clicked.connect(self.add_tab)
        self.toolbar.addWidget(self.new_tab_button)
        
        self.toolbar.addSeparator()
        
        # Save button
        self.save_button = QToolButton()
        self.save_button.setIcon(qta.icon('fa5s.save', color=ColorScheme.TEXT))
        self.save_button.setText("Save")
        self.save_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.save_button.clicked.connect(self.save_query)
        self.toolbar.addWidget(self.save_button)
        
        # Open button
        self.open_button = QToolButton()
        self.open_button.setIcon(qta.icon('fa5s.folder-open', color=ColorScheme.TEXT))
        self.open_button.setText("Open")
        self.open_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.open_button.clicked.connect(self.open_query)
        self.toolbar.addWidget(self.open_button)
        
        self.toolbar.addSeparator()
        
        # Import Data button
        self.import_data_button = QToolButton()
        self.import_data_button.setIcon(qta.icon('fa5s.file-import', color=ColorScheme.SUCCESS))
        self.import_data_button.setText("Import Data")
        self.import_data_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.import_data_button.setToolTip("Import data from CSV, Excel, Parquet, JSON files")
        self.import_data_button.clicked.connect(self.show_import_dialog)
        self.toolbar.addWidget(self.import_data_button)
        
        # Folder Import button
        self.folder_import_button = QToolButton()
        self.folder_import_button.setIcon(qta.icon('fa5s.folder-plus', color=ColorScheme.WARNING))
        self.folder_import_button.setText("Folder Import")
        self.folder_import_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.folder_import_button.setToolTip("Import entire folders of CSV or Excel files (Ctrl+Shift+F)")
        self.folder_import_button.clicked.connect(self.show_folder_import_dialog)
        self.toolbar.addWidget(self.folder_import_button)
        

        
        # CSV Automation button
        self.csv_automation_button = QToolButton()
        self.csv_automation_button.setIcon(qta.icon('fa5s.cogs', color=ColorScheme.ACCENT))
        self.csv_automation_button.setText("CSV Automation")
        self.csv_automation_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextBesideIcon)
        self.csv_automation_button.setToolTip("Automate processing of multiple CSV folders with SQL (Ctrl+Alt+A)")
        self.csv_automation_button.clicked.connect(self.show_csv_automation_dialog)
        self.toolbar.addWidget(self.csv_automation_button)
    
    def create_actions(self):
        # File actions
        self.new_action = QAction(qta.icon('fa5s.file', color=ColorScheme.TEXT), "&New Query", self)
        self.new_action.setShortcut(QKeySequence.StandardKey.New)
        self.new_action.triggered.connect(self.add_tab)
        
        self.open_action = QAction(qta.icon('fa5s.folder-open', color=ColorScheme.TEXT), "&Open Query...", self)
        self.open_action.setShortcut(QKeySequence.StandardKey.Open)
        self.open_action.triggered.connect(self.open_query)
        
        self.save_action = QAction(qta.icon('fa5s.save', color=ColorScheme.TEXT), "&Save Query", self)
        self.save_action.setShortcut(QKeySequence.StandardKey.Save)
        self.save_action.triggered.connect(self.save_query)
        
        self.save_as_action = QAction("Save Query &As...", self)
        self.save_as_action.setShortcut(QKeySequence.StandardKey.SaveAs)
        self.save_as_action.triggered.connect(self.save_query_as)
        
        self.exit_action = QAction("E&xit", self)
        self.exit_action.setShortcut(QKeySequence.StandardKey.Quit)
        self.exit_action.triggered.connect(self.close)
        
        # Database actions
        self.connect_action = QAction(qta.icon('fa5s.plug', color=ColorScheme.TEXT), "&Connect to Database...", self)
        self.connect_action.triggered.connect(self.show_connection_dialog)
        
        self.disconnect_action = QAction(qta.icon('fa5s.power-off', color=ColorScheme.ERROR), "&Disconnect", self)
        self.disconnect_action.triggered.connect(self.disconnect_database)
        self.disconnect_action.setEnabled(False)
        
        self.reconnect_main_action = QAction(qta.icon('fa5s.home', color=ColorScheme.ACCENT), "Reconnect to &Main Database", self)
        self.reconnect_main_action.setShortcut("Ctrl+M")
        self.reconnect_main_action.triggered.connect(self.auto_connect_main_database)
        
        self.import_data_action = QAction(qta.icon('fa5s.file-import', color=ColorScheme.SUCCESS), "&Import Data...", self)
        self.import_data_action.setShortcut("Ctrl+I")
        self.import_data_action.triggered.connect(self.show_import_dialog)
        
        # Folder Import action
        self.folder_import_action = QAction(qta.icon('fa5s.folder-plus', color=ColorScheme.WARNING), "Folder &Import...", self)
        self.folder_import_action.setShortcut("Ctrl+Shift+F")
        self.folder_import_action.setStatusTip("Import entire folders of CSV or Excel files")
        self.folder_import_action.triggered.connect(self.show_folder_import_dialog)
        

        
        # CSV Automation action
        self.csv_automation_action = QAction(qta.icon('fa5s.cogs', color=ColorScheme.ACCENT), "CSV Automation...", self)
        self.csv_automation_action.setShortcut("Ctrl+Alt+A")
        self.csv_automation_action.setStatusTip("Automate processing of multiple CSV folder sources with SQL")
        self.csv_automation_action.triggered.connect(self.show_csv_automation_dialog)
        
        # Query actions
        self.execute_action = QAction(qta.icon('fa5s.play', color=ColorScheme.SUCCESS), "&Execute Query", self)
        self.execute_action.setShortcut("F5")
        self.execute_action.triggered.connect(self.execute_current_query)
        
        self.execute_selection_action = QAction(qta.icon('fa5s.play-circle', color=ColorScheme.ACCENT), "Execute &Selection", self)
        self.execute_selection_action.setShortcut("Ctrl+E")
        self.execute_selection_action.triggered.connect(self.execute_selected_query)
        
        # Export results action
        self.export_results_action = QAction(qta.icon('fa5s.download', color='#2196F3'), "Export &Results", self)
        self.export_results_action.setShortcut("Ctrl+Shift+E")
        self.export_results_action.triggered.connect(self.export_current_results)
        
        # Settings action
        self.settings_action = QAction(qta.icon('fa5s.cog'), "&Settings", self)
        self.settings_action.setStatusTip("Configure lazy loading and performance settings")
        self.settings_action.triggered.connect(self.show_settings_dialog)
        
        # Create keyboard shortcut for executing query with Ctrl+Enter
        self.execute_shortcut = QShortcut(QKeySequence("Ctrl+Return"), self)
        self.execute_shortcut.activated.connect(self.execute_current_query)
    
    def create_menus(self):
        # File menu
        self.file_menu = self.menuBar().addMenu("&File")
        self.file_menu.addAction(self.new_action)
        self.file_menu.addAction(self.open_action)
        self.file_menu.addSeparator()
        self.file_menu.addAction(self.save_action)
        self.file_menu.addAction(self.save_as_action)
        self.file_menu.addSeparator()
        self.file_menu.addAction(self.exit_action)
        
        # Database menu
        self.db_menu = self.menuBar().addMenu("&Database")
        
        # Create new database action
        self.create_db_action = QAction(qta.icon('fa5s.plus-circle'), "&Create New Database", self)
        self.create_db_action.setShortcut(QKeySequence("Ctrl+N"))
        self.create_db_action.triggered.connect(self.show_create_database_dialog)
        
        self.db_menu.addAction(self.create_db_action)
        self.db_menu.addSeparator()
        self.db_menu.addAction(self.connect_action)
        self.db_menu.addAction(self.reconnect_main_action)
        self.db_menu.addAction(self.disconnect_action)
        self.db_menu.addSeparator()
        self.db_menu.addAction(self.import_data_action)
        self.db_menu.addAction(self.folder_import_action)
        self.db_menu.addAction(self.csv_automation_action)
        
        # Query menu
        self.query_menu = self.menuBar().addMenu("&Query")
        self.query_menu.addAction(self.execute_action)
        self.query_menu.addAction(self.execute_selection_action)
        self.query_menu.addSeparator()
        self.query_menu.addAction(self.export_results_action)
        
        # Tools menu
        self.tools_menu = self.menuBar().addMenu("&Tools")
        self.tools_menu.addAction(self.settings_action)
    
    def add_tab(self):
        # Create new query tab
        tab = QueryTab(connection=self.current_connection, connection_info=self.current_connection_info)
        # Connect schema change signal
        tab.schema_changed.connect(self.refresh_schema_browser)
        tab_index = self.tab_widget.addTab(tab, f"Query {self.tab_widget.count() + 1}")
        self.tab_widget.setCurrentIndex(tab_index)
        tab.editor.setFocus()
    
    def close_tab(self, index):
        if self.tab_widget.count() > 1:
            self.tab_widget.removeTab(index)
        else:
            # If it's the last tab, clear it instead of closing
            tab = self.tab_widget.widget(0)
            tab.editor.clear()
    
    def show_connection_dialog(self):
        dialog = ConnectionDialog(self)
        if dialog.exec():
            connection_info = dialog.get_connection_info()
            self.connect_to_database(connection_info)
    
    def show_create_database_dialog(self):
        dialog = CreateDatabaseDialog(self)
        if dialog.exec():
            connection_info = dialog.get_connection_info()
            if connection_info and connection_info.get('file_path'):
                self.connect_to_database(connection_info)
    
    def show_import_dialog(self):
        if not self.current_connection:
            QMessageBox.warning(self, "No Connection", "Please connect to a database first.")
            return
        
        dialog = DataImportDialog(self, self.current_connection, self.current_connection_info)
        if dialog.exec():
            import_info = dialog.get_import_info()
            
            # For create mode, still ask for table name confirmation
            if import_info['mode'] == 'create':
                suggested_name = import_info['table_name'] or self.suggest_table_name(import_info['file_path'])
                table_name = self.show_table_name_dialog(suggested_name, import_info['file_path'], 'create')
                if table_name:
                    import_info['table_name'] = table_name
                else:
                    return  # User cancelled or didn't provide a name
            
            # For append and replace modes, table name is already selected from dropdown
            self.start_import_worker(import_info)
    
    def show_folder_import_dialog(self):
        """Show the folder import dialog"""
        if not self.current_connection:
            QMessageBox.warning(self, "No Connection", "Please connect to a database first.")
            return
            
        try:
            dialog = FolderImportDialog(self, self.current_connection, self.current_connection_info)
            dialog.exec()
            
            # Always refresh schema browser after dialog closes in case import occurred
            self.refresh_schema_browser()
            self.check_schema_changes()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open folder import dialog: {str(e)}")

    
    def show_csv_automation_dialog(self):
        """Show the CSV automation dialog"""
        if not self.current_connection:
            QMessageBox.warning(self, "No Connection", "Please connect to a database first.")
            return
            
        if not CSV_AUTOMATION_AVAILABLE:
            QMessageBox.warning(self, "Feature Not Available", 
                              "CSV automation feature is not available.\n"
                              "Please ensure csv_automation.py is in the same directory.")
            return
            
        try:
            # Get both result and dialog object to access automation info
            result, dialog = show_csv_automation_dialog(self, self.current_connection, self.current_connection_info)
            
            # Always refresh schema browser after dialog closes in case import occurred
            self.refresh_schema_browser()
            self.check_schema_changes()
            
            # Check if an automation with SQL query was executed successfully
            if result == QDialog.DialogCode.Accepted or dialog.automation_results:
                automation_info = dialog.get_executed_automation_info()
                if automation_info['has_sql_query']:
                    self.display_automation_sql_in_editor(automation_info)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open CSV automation dialog: {str(e)}")
    
    def show_settings_dialog(self):
        """Show the lazy loading settings dialog"""
        dialog = LazyLoadingSettingsDialog(self)
        if dialog.exec():
            # Settings were saved automatically in the dialog
            QMessageBox.information(
                self, 
                "Settings Saved", 
                "Lazy loading settings have been saved. Changes will apply to new queries."
            )
    
    def display_automation_sql_in_editor(self, automation_info):
        """Display the executed automation SQL query in the main query editor and execute it"""
        try:
            sql_query = automation_info['sql_query']
            output_table = automation_info['output_table']
            results = automation_info['results']
            
            if not sql_query or not output_table:
                return
            
            # Get current tab or create new one
            current_tab = self.tab_widget.currentWidget()
            if not current_tab or current_tab.editor.toPlainText().strip():
                # Create new tab if current one has content
                self.add_tab()
                current_tab = self.tab_widget.currentWidget()
            
            # Update tab title to reflect automation
            tab_index = self.tab_widget.currentIndex()
            self.tab_widget.setTabText(tab_index, f"Automation SQL - {output_table}")
            
            # Set the SQL query in the editor
            current_tab.editor.setPlainText(sql_query)
            
            # Add a comment header to the query for context
            header_comment = f"""-- CSV Automation Results
-- Output Table: {output_table}
-- Sources: {len(results.get('tables_created', []))} tables
-- Total Rows Processed: {results.get('total_rows', 0):,}
-- Execution Time: {results.get('execution_time', 0):.2f} seconds

"""
            
            # Prepend the header to the existing query
            full_query = header_comment + sql_query
            current_tab.editor.setPlainText(full_query)
            
            # Show status message
            self.statusBar().showMessage(
                f"🚀 Automation SQL loaded! Query creates '{output_table}' with {results.get('output_rows', 0):,} rows", 
                5000
            )
            
            # Auto-execute the query to show results
            QTimer.singleShot(500, current_tab.execute_query)
            
        except Exception as e:
            self.statusBar().showMessage(f"Failed to display automation SQL: {str(e)}", 3000)

    def start_import_worker(self, import_info):
        """Start the import worker thread with progress dialog"""
        # Create and show progress dialog
        self.progress_dialog = ProgressDialog(self, "Importing File...")
        
        # Create and start worker thread
        self.import_worker = ImportWorker(self, import_info)
        
        # Connect worker signals
        self.import_worker.progress.connect(self.progress_dialog.update_progress)
        self.import_worker.finished.connect(self.on_import_finished)
        self.import_worker.error.connect(self.on_import_error)
        
        # Connect progress dialog cancel to worker termination
        self.progress_dialog.rejected.connect(self.cancel_import)
        
        # Start the worker and show progress dialog
        self.import_worker.start()
        self.progress_dialog.exec()
    
    def on_import_finished(self, success, message):
        """Handle import completion"""
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.accept()
        
        if success:
            QMessageBox.information(self, "Import Successful", message)
            self.refresh_schema_browser()
            self.check_schema_changes()
        else:
            QMessageBox.warning(self, "Import Failed", message)
    
    def on_import_error(self, error_message):
        """Handle import error"""
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.reject()
        
        QMessageBox.critical(self, "Import Error", f"An error occurred during import:\n{error_message}")
    
    def cancel_import(self):
        """Cancel the running import"""
        if hasattr(self, 'import_worker') and self.import_worker and self.import_worker.isRunning():
            self.import_worker.cancel()  # Signal the worker to cancel gracefully
            self.import_worker.terminate()  # Force terminate if needed
            self.import_worker.wait()
            QMessageBox.information(self, "Import Cancelled", "Import operation was cancelled.")
    
    def show_table_name_dialog(self, suggested_name, file_path, mode='create'):
        """Show a custom dialog for table name input with better visibility"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Create New Table")
        dialog.setModal(True)
        dialog.resize(450, 200)
        
        # Set dialog style
        dialog.setStyleSheet(f"""
            QDialog {{
                background-color: {ColorScheme.BACKGROUND.name()};
                color: {ColorScheme.TEXT.name()};
            }}
            QLabel {{
                color: {ColorScheme.TEXT.name()};
                font-size: 12px;
                margin: 5px;
            }}
            QLineEdit {{
                background-color: {ColorScheme.SIDEBAR_BG.name()};
                color: {ColorScheme.TEXT.name()};
                border: 2px solid {ColorScheme.ACCENT.name()};
                border-radius: 5px;
                padding: 8px;
                font-size: 14px;
                font-weight: bold;
            }}
            QLineEdit:focus {{
                border-color: {ColorScheme.SUCCESS.name()};
            }}
            QPushButton {{
                background-color: {ColorScheme.ACCENT.name()};
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                font-size: 12px;
                font-weight: bold;
                min-width: 80px;
            }}
            QPushButton:hover {{
                background-color: {ColorScheme.SUCCESS.name()};
            }}
            QPushButton:pressed {{
                background-color: {ColorScheme.HIGHLIGHT.name()};
            }}
        """)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Title label
        title_label = QLabel("🆕 Create New Table")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #4FC3F7; margin-bottom: 10px;")
        layout.addWidget(title_label)
        
        # File info label
        file_info = QLabel(f"📁 Importing from: {os.path.basename(file_path)}")
        file_info.setStyleSheet("font-size: 11px; color: #B0BEC5; margin-bottom: 5px;")
        layout.addWidget(file_info)
        
        # Instruction label
        instruction_label = QLabel("Please enter a name for the new table:")
        instruction_label.setStyleSheet("font-size: 12px; margin-bottom: 5px;")
        layout.addWidget(instruction_label)
        
        # Table name input
        table_name_input = QLineEdit()
        table_name_input.setText(suggested_name)
        table_name_input.selectAll()  # Select all text for easy editing
        table_name_input.setPlaceholderText("Enter table name...")
        layout.addWidget(table_name_input)
        
        # Validation label
        validation_label = QLabel("✓ Table name looks good!")
        validation_label.setStyleSheet("font-size: 10px; color: #4CAF50; margin-top: 5px;")
        layout.addWidget(validation_label)
        
        # Button layout
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        cancel_button = QPushButton("Cancel")
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #757575;
            }
            QPushButton:hover {
                background-color: #9E9E9E;
            }
        """)
        create_button = QPushButton("Create Table")
        
        button_layout.addWidget(cancel_button)
        button_layout.addWidget(create_button)
        layout.addLayout(button_layout)
        
        # Validation function
        def validate_table_name():
            name = table_name_input.text().strip()
            if not name:
                validation_label.setText("⚠️ Please enter a table name")
                validation_label.setStyleSheet("font-size: 10px; color: #FF9800;")
                create_button.setEnabled(False)
            elif not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', name):
                validation_label.setText("⚠️ Use only letters, numbers, and underscores. Must start with letter or underscore.")
                validation_label.setStyleSheet("font-size: 10px; color: #FF5722;")
                create_button.setEnabled(False)
            else:
                validation_label.setText("✓ Table name looks good!")
                validation_label.setStyleSheet("font-size: 10px; color: #4CAF50;")
                create_button.setEnabled(True)
        
        # Connect validation
        table_name_input.textChanged.connect(validate_table_name)
        
        # Connect buttons
        cancel_button.clicked.connect(dialog.reject)
        create_button.clicked.connect(dialog.accept)
        
        # Allow Enter key to create table
        table_name_input.returnPressed.connect(lambda: dialog.accept() if create_button.isEnabled() else None)
        
        # Set focus and validate initially
        table_name_input.setFocus()
        validate_table_name()
        
        # Show dialog and return result
        if dialog.exec() == QDialog.DialogCode.Accepted:
            return table_name_input.text().strip()
        return None
    

    
    def suggest_table_name(self, file_path):
        """Suggest a table name based on the file path"""
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        # Clean table name (remove special characters, replace with underscores)
        clean_name = re.sub(r'[^a-zA-Z0-9_]', '_', base_name).lower()
        # Remove consecutive underscores
        clean_name = re.sub(r'_+', '_', clean_name)
        # Remove leading/trailing underscores
        clean_name = clean_name.strip('_')
        # Ensure it starts with a letter or underscore
        if clean_name and not clean_name[0].isalpha() and clean_name[0] != '_':
            clean_name = f"table_{clean_name}"
        # Handle empty names
        if not clean_name:
            clean_name = "new_table"
        return clean_name
    
    def connect_to_database(self, connection_info):
        try:
            db_type = connection_info["type"]
            # Handle both 'file_path' and 'path' keys for compatibility
            file_path = connection_info.get("file_path") or connection_info.get("path")
            
            if not file_path:
                QMessageBox.warning(self, "Connection Error", "Please specify a database file.")
                return
            
            # Check if we already have this connection
            connection_key = f"{db_type}:{file_path}"
            if connection_key in self.connections:
                self.current_connection = self.connections[connection_key]
                self.current_connection_info = connection_info
            else:
                # Create new connection
                if db_type.lower() in ["sqlite", "sqlite3"]:
                    connection = sqlite3.connect(file_path)
                elif db_type.lower() == "duckdb":
                    connection = duckdb.connect(file_path)
                else:
                    raise ValueError(f"Unsupported database type: {db_type}")
                
                self.connections[connection_key] = connection
                self.current_connection = connection
                self.current_connection_info = connection_info
            
            # Update UI - show special indicator for main database
            db_name = connection_info.get("name", os.path.basename(file_path))
            if db_name == "main.duckdb":
                self.connection_label.setText(f"🏠 {db_name} (Main Database)")
            else:
                self.connection_label.setText(f"Connected to {db_name} ({db_type})")
            
            self.disconnect_action.setEnabled(True)
            
            # Update schema browser
            self.schema_browser.load_schema(self.current_connection, self.current_connection_info)
            
            # Update all tabs with the new connection
            for i in range(self.tab_widget.count()):
                tab = self.tab_widget.widget(i)
                tab.connection = self.current_connection
                tab.connection_info = self.current_connection_info
                # Connect schema change signal if not already connected
                try:
                    tab.schema_changed.connect(self.refresh_schema_browser)
                except:
                    pass  # Signal might already be connected
            
            self.statusBar().showMessage(f"Connected to {file_path}", 3000)
            
        except Exception as e:
            QMessageBox.critical(self, "Connection Error", f"Failed to connect to database: {str(e)}")
    
    def disconnect_database(self):
        if self.current_connection:
            # Close the connection
            file_path = self.current_connection_info.get("file_path") or self.current_connection_info.get("path")
            connection_key = f"{self.current_connection_info['type']}:{file_path}"
            if connection_key in self.connections:
                try:
                    self.connections[connection_key].close()
                    del self.connections[connection_key]
                except:
                    pass
            
            # Update UI
            self.current_connection = None
            self.current_connection_info = None
            self.connection_label.setText("Not connected")
            self.disconnect_action.setEnabled(False)
            self.schema_browser.clear()
            
            # Update all tabs
            for i in range(self.tab_widget.count()):
                tab = self.tab_widget.widget(i)
                tab.connection = None
                tab.connection_info = None
            
            self.statusBar().showMessage("Disconnected from database", 3000)
    
    def import_data(self, import_info):
        """NEW: Completely rebuilt import system with guaranteed UI responsiveness"""
        try:
            # Always use the worker-based import for UI responsiveness
            self.start_import_worker(import_info, is_folder=False)
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to start import: {str(e)}")
    
    def import_data_optimized(self, import_info, worker=None):
        """NEW: Completely rebuilt optimized import with guaranteed UI responsiveness"""
        try:
            file_path = import_info['file_path']
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            
            if worker:
                worker.progress.emit(5, f"Starting import: {os.path.basename(file_path)} ({file_size_mb:.1f} MB)")
            
            # Use ultra-fast native import for maximum speed
            return self.ultra_fast_native_import(import_info, worker)
            
        except Exception as e:
            print(f"Import failed: {e}")
            if worker:
                worker.error.emit(f"Import failed: {str(e)}")
            return False
    
    def ultra_fast_native_import(self, import_info, worker=None):
        """ULTRA-FAST: Use native database import functions for maximum speed"""
        try:
            file_path = import_info['file_path']
            table_name = import_info['table_name']
            file_type = import_info['file_type']
            mode = import_info['mode']
            
            # DEBUG: Log import details
            print(f"\n=== IMPORT DEBUG INFO ===")
            print(f"File: {os.path.basename(file_path)}")
            print(f"Table: {table_name}")
            print(f"Mode: {mode}")
            print(f"File type: {file_type}")
            
            # Clean table name
            safe_table_name = self.clean_table_name(table_name)
            print(f"Safe table name: {safe_table_name}")
            
            if worker:
                worker.progress.emit(5, "Initializing ultra-fast native import...")
            
            # Check database type
            db_type = self.current_connection_info.get('type', '').lower()
            print(f"  Database type: {db_type}")
            
            if db_type == 'duckdb':
                print(f"  Using DuckDB native import for {file_type} in {mode} mode")
                return self.duckdb_native_import(file_path, file_type, safe_table_name, mode, import_info, worker)
            elif db_type == 'sqlite':
                print(f"  Using SQLite native import for {file_type} in {mode} mode")
                return self.sqlite_native_import(file_path, file_type, safe_table_name, mode, import_info, worker)
            else:
                print(f"  Falling back to streaming import for {file_type} in {mode} mode")
                # Fallback to streaming import
                return self.streaming_import_with_progress(import_info, worker)
                
        except Exception as e:
            print(f"Ultra-fast native import failed: {e}")
            if worker:
                worker.error.emit(f"Ultra-fast import failed: {str(e)}")
            return False
    
    def duckdb_native_import(self, file_path, file_type, table_name, mode, import_info, worker=None):
        """DuckDB native import - extremely fast for CSV/TSV/Parquet"""
        try:
            if worker:
                worker.progress.emit(10, "Using DuckDB native import for maximum speed...")
            
            # Drop table if create or replace mode
            if mode in ['create', 'replace']:
                self.drop_table_if_exists(table_name)
            
            # Check if we need to add filename column for folder imports
            add_filename_column = import_info.get('add_filename_column', False)
            source_filename = import_info.get('source_filename', '')
            
            # Build native import SQL based on file type
            if file_type in ['.csv', '.txt']:
                delimiter = import_info.get('delimiter', ',')
                encoding = import_info.get('encoding', 'utf-8')
                has_header = import_info.get('header', True)
                
                if worker:
                    worker.progress.emit(20, "Executing DuckDB native CSV import...")
                
                # DuckDB native CSV import with optional filename column
                if add_filename_column and source_filename:
                    base_select = f"SELECT *, '{source_filename}' AS _source_file FROM read_csv_auto("
                else:
                    base_select = "SELECT * FROM read_csv_auto("
                
                csv_params = f"""
                    '{file_path}',
                    delim='{delimiter}',
                    header={str(has_header).lower()},
                    all_varchar=true,
                    ignore_errors=true,
                    sample_size=100000
                )"""
                
                # For create/replace mode, use CREATE TABLE AS (table was already dropped)
                if mode in ['create', 'replace']:
                    sql = f"""
                    CREATE TABLE {table_name} AS 
                    {base_select}{csv_params}
                    """
                else:  # append mode
                    sql = f"""
                    INSERT INTO {table_name} 
                    {base_select}{csv_params}
                    """
                    
            elif file_type == '.tsv':
                has_header = import_info.get('header', True)
                
                if worker:
                    worker.progress.emit(20, "Executing DuckDB native TSV import...")
                
                # TSV import with optional filename column
                if add_filename_column and source_filename:
                    base_select = f"SELECT *, '{source_filename}' AS _source_file FROM read_csv_auto("
                else:
                    base_select = "SELECT * FROM read_csv_auto("
                
                tsv_params = f"""
                    '{file_path}',
                    delim='\\t',
                    header={str(has_header).lower()},
                    all_varchar=true,
                    ignore_errors=true,
                    sample_size=100000
                )"""
                
                # For create/replace mode, use CREATE TABLE AS (table was already dropped)
                if mode in ['create', 'replace']:
                    sql = f"""
                    CREATE TABLE {table_name} AS 
                    {base_select}{tsv_params}
                    """
                else:  # append mode
                    sql = f"""
                    INSERT INTO {table_name} 
                    {base_select}{tsv_params}
                    """
                    
            elif file_type == '.parquet':
                if worker:
                    worker.progress.emit(20, "Executing DuckDB native Parquet import...")
                
                # Enhanced parquet import with optional filename column
                if add_filename_column and source_filename:
                    base_select = f"SELECT *, '{source_filename}' AS _source_file FROM read_parquet('{file_path}')"
                else:
                    base_select = f"SELECT * FROM read_parquet('{file_path}')"
                
                # For create/replace mode, use CREATE TABLE AS (table was already dropped)
                if mode in ['create', 'replace']:
                    sql = f"""
                    CREATE TABLE {table_name} AS 
                    {base_select}
                    """
                else:  # append mode
                    sql = f"""
                    INSERT INTO {table_name} 
                    {base_select}
                    """
                    
            elif file_type == '.json':
                if worker:
                    worker.progress.emit(20, "Executing DuckDB native JSON import...")
                
                # Enhanced JSON import with optional filename column
                if add_filename_column and source_filename:
                    base_select = f"SELECT *, '{source_filename}' AS _source_file FROM read_json_auto('{file_path}')"
                else:
                    base_select = f"SELECT * FROM read_json_auto('{file_path}')"
                
                # For create/replace mode, use CREATE TABLE AS (table was already dropped)
                if mode in ['create', 'replace']:
                    sql = f"""
                    CREATE TABLE {table_name} AS 
                    {base_select}
                    """
                else:  # append mode
                    sql = f"""
                    INSERT INTO {table_name} 
                    {base_select}
                    """
            elif file_type in ['.xlsx', '.xls']:
                if worker:
                    worker.progress.emit(20, "Using streamlined Excel import...")
                
                # Use the streamlined import approach like other files
                return self.excel_streamlined_import(file_path, table_name, mode, import_info, worker)
            else:
                # Fallback for unsupported formats
                return self.streaming_import_with_progress({'file_path': file_path, 'table_name': table_name, 'file_type': file_type, 'mode': mode}, worker)
            
            if worker:
                worker.progress.emit(30, f"Executing ultra-fast {file_type.upper()} import...")
                worker.progress.emit(35, "Processing file structure...")
            
            # Execute the native import with enhanced error handling
            start_time = time.time()
            try:
                self.current_connection.execute(sql)
                if worker:
                    worker.progress.emit(70, "Import execution completed successfully...")
            except Exception as sql_error:
                if worker:
                    worker.progress.emit(40, f"Native import failed, trying alternative method...")
                print(f"Native SQL failed: {sql_error}")
                # Try alternative reading approaches for different file types
                if file_type == '.parquet':
                    return self.parquet_fallback_import(file_path, table_name, mode, worker)
                elif file_type == '.json':
                    return self.json_fallback_import(file_path, table_name, mode, worker)
                elif file_type in ['.xlsx', '.xls']:
                    return self.excel_fallback_import(file_path, table_name, mode, import_info, worker)
                else:
                    raise sql_error
            end_time = time.time()
            
            if worker:
                worker.progress.emit(80, "Counting imported rows...")
            
            # Get row count
            result = self.current_connection.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()
            row_count = result[0] if result else 0
            
            import_time = end_time - start_time
            speed = row_count / max(import_time, 0.001)
            
            if worker:
                worker.progress.emit(90, f"Import completed: {row_count:,} rows in {import_time:.2f}s ({speed:,.0f} rows/sec)")
            
            # Refresh schema browser with error handling
            try:
                self.refresh_schema_browser()
            except Exception as refresh_error:
                print(f"Schema refresh warning: {refresh_error}")
                # Import was successful, just refresh failed - continue
            
            print(f"Native import completed: {row_count:,} rows in {import_time:.2f}s ({speed:,.0f} rows/sec)")
            return True
            
        except Exception as e:
            print(f"DuckDB native import failed: {e}")
            if worker:
                worker.error.emit(f"DuckDB native import failed: {str(e)}")
            return False
    
    def drop_table_if_exists(self, table_name):
        """Helper to drop table if exists"""
        try:
            self.current_connection.execute(f"DROP TABLE IF EXISTS {table_name}")
        except Exception as e:
            print(f"Could not drop table {table_name}: {e}")
    
    def sqlite_native_import(self, file_path, file_type, table_name, mode, import_info, worker=None):
        """SQLite native import fallback - uses CSV import for CSV files"""
        try:
            if worker:
                worker.progress.emit(10, "Using SQLite optimized import...")
            
            # For SQLite, we'll use the streaming import as it doesn't have native CSV readers like DuckDB
            return self.streaming_import_with_progress(import_info, worker)
            
        except Exception as e:
            print(f"SQLite import failed: {e}")
            if worker:
                worker.error.emit(f"SQLite import failed: {str(e)}")
            return False
    
    def streaming_import_with_progress(self, import_info, worker=None):
        """Streaming import with progress for fallback cases"""
        try:
            if worker:
                worker.progress.emit(10, "Using streaming import fallback...")
            
            # Use the existing optimized import functions
            return self.import_small_file_fast(import_info, worker)
            
        except Exception as e:
            print(f"Streaming import failed: {e}")
            if worker:
                worker.error.emit(f"Streaming import failed: {str(e)}")
            return False
    
    def excel_fallback_import(self, file_path, table_name, mode, import_info, worker=None):
        """Excel fallback import method with optimized chunked reading"""
        try:
            if worker:
                worker.progress.emit(10, "Analyzing Excel file...")
            
            # Get file size to determine if we need chunked reading
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            
            safe_table_name = self.ensure_unique_table_name(table_name, mode)
            
            # For large Excel files (>50MB), use chunked reading
            if file_size_mb > 50:
                return self.excel_chunked_import(file_path, safe_table_name, mode, import_info, worker)
            else:
                return self.excel_direct_import(file_path, safe_table_name, mode, import_info, worker)
            
        except Exception as e:
            print(f"Excel fallback import failed: {e}")
            if worker:
                worker.error.emit(f"Excel import failed: {str(e)}")
            return False
    
    def excel_direct_import(self, file_path, table_name, mode, import_info, worker=None):
        """Direct Excel import for smaller files"""
        try:
            if worker:
                worker.progress.emit(20, "Loading Excel file...")
            
            # Use optimized Excel reading
            sheet_name = import_info.get('sheet_name', 0)
            df = pd.read_excel(
                file_path, 
                sheet_name=sheet_name,
                dtype=str,  # Read everything as string for consistency
                na_filter=False,  # Don't convert to NaN
                engine='openpyxl'
            )
            
            if df is None or df.empty:
                if worker:
                    worker.error.emit("No data found in Excel file")
                return False
            
            if worker:
                worker.progress.emit(50, f"Processing {len(df):,} rows from Excel...")
            
            # Process and import
            df = self.quick_process_dataframe(df)
            
            if worker:
                worker.progress.emit(70, f"Inserting {len(df):,} rows into database...")
            
            return self.fast_database_insert(df, table_name, mode, worker)
            
        except Exception as e:
            print(f"Excel direct import failed: {e}")
            if worker:
                worker.error.emit(f"Excel direct import failed: {str(e)}")
            return False
    
    def excel_chunked_import(self, file_path, table_name, mode, import_info, worker=None):
        """Chunked Excel import for large files"""
        try:
            if worker:
                worker.progress.emit(15, "Reading Excel file in chunks...")
            
            sheet_name = import_info.get('sheet_name', 0)
            
            # Read Excel file to get total rows first
            excel_file = pd.ExcelFile(file_path, engine='openpyxl')
            
            # Get sheet info
            if isinstance(sheet_name, int):
                actual_sheet = excel_file.sheet_names[sheet_name] if sheet_name < len(excel_file.sheet_names) else excel_file.sheet_names[0]
            else:
                actual_sheet = sheet_name if sheet_name in excel_file.sheet_names else excel_file.sheet_names[0]
            
            if worker:
                worker.progress.emit(25, f"Processing Excel sheet: {actual_sheet}")
            
            # Read in chunks using pandas chunking
            chunk_size = 10000  # Read 10k rows at a time
            table_created = False
            total_rows = 0
            
            # Use chunksize parameter if available (pandas 1.4+)
            try:
                chunk_reader = pd.read_excel(
                    file_path,
                    sheet_name=actual_sheet,
                    chunksize=chunk_size,
                    dtype=str,
                    na_filter=False,
                    engine='openpyxl'
                )
                
                for chunk_num, chunk_df in enumerate(chunk_reader):
                    if worker and worker.cancelled:
                        return False
                    
                    if worker:
                        progress = 30 + int((chunk_num * 50) / 100)  # Estimate progress
                        worker.progress.emit(progress, f"Processing chunk {chunk_num + 1}: {len(chunk_df):,} rows")
                    
                    # Process chunk
                    chunk_df = self.quick_process_dataframe(chunk_df)
                    
                    # Insert chunk
                    if not table_created:
                        success = self.fast_database_insert(chunk_df, table_name, mode, worker)
                        table_created = True
                    else:
                        success = self.fast_database_insert(chunk_df, table_name, 'append', worker)
                    
                    if not success:
                        if worker:
                            worker.error.emit(f"Failed to insert chunk {chunk_num + 1}")
                        return False
                    
                    total_rows += len(chunk_df)
                    
            except TypeError:
                # Fallback for older pandas versions without chunksize support for Excel
                if worker:
                    worker.progress.emit(30, "Reading entire Excel file (chunked reading not supported)...")
                
                df = pd.read_excel(
                    file_path,
                    sheet_name=actual_sheet,
                    dtype=str,
                    na_filter=False,
                    engine='openpyxl'
                )
                
                if df is None or df.empty:
                    if worker:
                        worker.error.emit("No data found in Excel file")
                    return False
                
                # Process in memory chunks
                chunk_size = 50000
                total_rows = len(df)
                
                for i in range(0, total_rows, chunk_size):
                    if worker and worker.cancelled:
                        return False
                    
                    chunk_df = df.iloc[i:i+chunk_size].copy()
                    chunk_num = i // chunk_size
                    
                    if worker:
                        progress = 40 + int((i / total_rows) * 50)
                        worker.progress.emit(progress, f"Processing chunk {chunk_num + 1}: {len(chunk_df):,} rows")
                    
                    # Process chunk
                    chunk_df = self.quick_process_dataframe(chunk_df)
                    
                    # Insert chunk
                    if not table_created:
                        success = self.fast_database_insert(chunk_df, table_name, mode, worker)
                        table_created = True
                    else:
                        success = self.fast_database_insert(chunk_df, table_name, 'append', worker)
                    
                    if not success:
                        if worker:
                            worker.error.emit(f"Failed to insert chunk {chunk_num + 1}")
                        return False
            
            if worker:
                worker.progress.emit(90, f"Excel import completed: {total_rows:,} rows")
            
            # Refresh schema browser
            self.refresh_schema_browser()
            
            print(f"Excel chunked import completed: {total_rows:,} rows imported to '{table_name}'")
            return True
            
        except Exception as e:
            print(f"Excel chunked import failed: {e}")
            if worker:
                worker.error.emit(f"Excel chunked import failed: {str(e)}")
            return False
    
    def excel_streamlined_import(self, file_path, table_name, mode, import_info, worker=None):
        """BLAZING-FAST Excel import - multi-engine optimization for maximum speed"""
        try:
            if worker:
                worker.progress.emit(2, "Selecting fastest Excel engine...")
            
            # Get sheet information
            sheet_name = import_info.get('sheet_name', 0)
            file_ext = os.path.splitext(file_path)[1].lower()
            
            start_total = time.time()
            
            # Try multiple engines in order of speed (fastest first)
            excel_data = None
            engine_used = "unknown"
            
            # Engine 1: Try python-calamine (Rust-based, fastest)
            try:
                import python_calamine
                if worker:
                    worker.progress.emit(8, "Using BLAZING-FAST Rust engine (calamine)...")
                excel_data, engine_used = self._read_excel_calamine(file_path, sheet_name)
            except ImportError:
                pass
            except Exception as e:
                print(f"Calamine engine failed: {e}")
            
            # Engine 2: Try polars (very fast)
            if excel_data is None:
                try:
                    import polars as pl
                    if worker:
                        worker.progress.emit(8, "Using ULTRA-FAST Polars engine...")
                    excel_data, engine_used = self._read_excel_polars(file_path, sheet_name)
                except ImportError:
                    pass
                except Exception as e:
                    print(f"Polars engine failed: {e}")
            
            # Engine 3: Try xlrd for .xls files (faster than openpyxl for old format)
            if excel_data is None and file_ext == '.xls':
                try:
                    import xlrd
                    if worker:
                        worker.progress.emit(8, "Using optimized XLS engine...")
                    excel_data, engine_used = self._read_excel_xlrd(file_path, sheet_name)
                except ImportError:
                    pass
                except Exception as e:
                    print(f"xlrd engine failed: {e}")
            
            # Engine 4: Fallback to optimized openpyxl
            if excel_data is None:
                if worker:
                    worker.progress.emit(8, "Using optimized OpenPyXL engine...")
                excel_data, engine_used = self._read_excel_openpyxl(file_path, sheet_name)
            
            if excel_data is None:
                if worker:
                    worker.error.emit("Failed to read Excel file with all available engines")
                return False
            
            read_time = time.time() - start_total
            rows_count = len(excel_data)
            
            if worker:
                worker.progress.emit(35, f"✨ {engine_used} read {rows_count:,} rows in {read_time:.2f}s")
            
            if rows_count == 0:
                if worker:
                    worker.error.emit("No data found in Excel file")
                return False
            
            # Convert to DataFrame for fast database insertion
            if worker:
                worker.progress.emit(45, "Converting to optimized format...")
            
            df = pd.DataFrame(excel_data[1:], columns=excel_data[0])  # First row as headers
            
            # Drop table if create mode
            if mode == 'create':
                self.drop_table_if_exists(table_name)
            
            if worker:
                worker.progress.emit(60, "Executing BLAZING database insert...")
            
            # Use the fastest database insertion method
            import_start = time.time()
            success = self.fast_database_insert(df, table_name, mode, worker)
            
            if not success:
                return False
            
            if worker:
                worker.progress.emit(85, "Verifying import...")
            
            # Get row count
            result = self.current_connection.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()
            actual_rows = result[0] if result else 0
            
            total_time = time.time() - start_total
            import_time = time.time() - import_start
            speed = actual_rows / max(total_time, 0.001)
            
            if worker:
                worker.progress.emit(95, f"🚀 BLAZING import: {actual_rows:,} rows in {total_time:.2f}s ({speed:,.0f} rows/sec)")
            
                            # Refresh schema browser with error handling
                try:
                    self.refresh_schema_browser()
                except Exception as refresh_error:
                    print(f"Schema refresh warning: {refresh_error}")
                    # Import was successful, just refresh failed - continue
                
                print(f"🚀 BLAZING Excel import ({engine_used}): {actual_rows:,} rows in {total_time:.2f}s ({speed:,.0f} rows/sec)")
                print(f"   📖 Excel read: {read_time:.2f}s | 💾 DB insert: {import_time:.2f}s")
                return True
            
        except Exception as e:
            print(f"Blazing Excel import failed: {e}")
            if worker:
                worker.error.emit(f"Excel import failed: {str(e)}")
            return False
    
    def _read_excel_calamine(self, file_path, sheet_name):
        """Read Excel using python-calamine (Rust-based, fastest)"""
        from python_calamine import CalamineWorkbook
        
        workbook = CalamineWorkbook.from_path(file_path)
        
        # Get sheet names and select the right one
        sheet_names = workbook.sheet_names
        if isinstance(sheet_name, int):
            actual_sheet = sheet_names[sheet_name] if sheet_name < len(sheet_names) else sheet_names[0]
        else:
            actual_sheet = sheet_name if sheet_name in sheet_names else sheet_names[0]
        
        # Read all data at once
        data = workbook.get_sheet_by_name(actual_sheet).to_python()
        return data, "Rust-Calamine"
    
    def _read_excel_polars(self, file_path, sheet_name):
        """Read Excel using polars (very fast)"""
        import polars as pl
        
        # Read with polars
        if isinstance(sheet_name, int):
            df = pl.read_excel(file_path, sheet_id=sheet_name + 1)  # polars uses 1-based indexing
        else:
            df = pl.read_excel(file_path, sheet_name=sheet_name)
        
        # Convert to list of lists
        headers = df.columns
        data = [headers] + df.to_pandas().values.tolist()  # Convert to compatible format
        return data, "Polars"
    
    def _read_excel_xlrd(self, file_path, sheet_name):
        """Read Excel using xlrd (good for .xls files)"""
        import xlrd
        
        workbook = xlrd.open_workbook(file_path)
        
        # Get the worksheet
        if isinstance(sheet_name, int):
            worksheet = workbook.sheet_by_index(sheet_name)
        else:
            worksheet = workbook.sheet_by_name(sheet_name)
        
        # Read all data
        data = []
        for row_idx in range(worksheet.nrows):
            row = [str(worksheet.cell_value(row_idx, col_idx)) for col_idx in range(worksheet.ncols)]
            data.append(row)
        
        return data, "xlrd"
    
    def _read_excel_openpyxl(self, file_path, sheet_name):
        """Read Excel using openpyxl (fallback, but optimized)"""
        from openpyxl import load_workbook
        
        # Load with maximum optimizations
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        
        # Get the worksheet
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name] if sheet_name < len(wb.worksheets) else wb.active
        else:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        
        # Read all data at once (faster than row-by-row)
        data = []
        for row in ws.iter_rows(values_only=True):
            clean_row = [str(cell) if cell is not None else '' for cell in row]
            data.append(clean_row)
        
        wb.close()
        return data, "OpenPyXL"
    
    def parquet_fallback_import(self, file_path, table_name, mode, worker=None):
        """Parquet fallback import method"""
        try:
            if worker:
                worker.progress.emit(20, "Loading Parquet file with pandas...")
            
            # Load with pandas
            df = pd.read_parquet(file_path)
            
            if df is None or df.empty:
                if worker:
                    worker.error.emit("No data found in Parquet file")
                return False
            
            if worker:
                worker.progress.emit(50, f"Processing {len(df):,} rows from Parquet...")
            
            # Process and import
            df = self.quick_process_dataframe(df)
            safe_table_name = self.ensure_unique_table_name(table_name, mode)
            
            return self.fast_database_insert(df, safe_table_name, mode, worker)
            
        except Exception as e:
            print(f"Parquet fallback import failed: {e}")
            if worker:
                worker.error.emit(f"Parquet import failed: {str(e)}")
            return False
    
    def json_fallback_import(self, file_path, table_name, mode, worker=None):
        """JSON fallback import method with improved error handling and chunking"""
        try:
            if worker:
                worker.progress.emit(10, "Analyzing JSON file...")
            
            # Get file size to determine approach
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            
            safe_table_name = self.ensure_unique_table_name(table_name, mode)
            
            # For large JSON files (>100MB), use chunked reading
            if file_size_mb > 100:
                return self.json_chunked_import(file_path, safe_table_name, mode, worker)
            else:
                return self.json_direct_import(file_path, safe_table_name, mode, worker)
            
        except Exception as e:
            print(f"JSON fallback import failed: {e}")
            if worker:
                worker.error.emit(f"JSON import failed: {str(e)}")
            return False
    
    def json_direct_import(self, file_path, table_name, mode, worker=None):
        """Direct JSON import for smaller files"""
        try:
            if worker:
                worker.progress.emit(20, "Loading JSON file...")
            
            # Try multiple approaches for loading JSON
            df = None
            
            # First try: pandas read_json (fastest)
            try:
                df = pd.read_json(file_path)
            except ValueError:
                # Second try: manual JSON parsing (more flexible)
                if worker:
                    worker.progress.emit(30, "Trying alternative JSON parsing...")
                
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                if isinstance(data, list):
                    df = pd.DataFrame(data)
                elif isinstance(data, dict):
                    df = pd.DataFrame([data])
                else:
                    if worker:
                        worker.error.emit("Unsupported JSON format")
                    return False
            
            if df is None or df.empty:
                if worker:
                    worker.error.emit("No data found in JSON file")
                return False
            
            if worker:
                worker.progress.emit(50, f"Processing {len(df):,} rows from JSON...")
            
            # Process and import
            df = self.quick_process_dataframe(df)
            
            if worker:
                worker.progress.emit(70, f"Inserting {len(df):,} rows into database...")
            
            return self.fast_database_insert(df, table_name, mode, worker)
            
        except Exception as e:
            print(f"JSON direct import failed: {e}")
            if worker:
                worker.error.emit(f"JSON direct import failed: {str(e)}")
            return False
    
    def json_chunked_import(self, file_path, table_name, mode, worker=None):
        """Chunked JSON import for large files"""
        try:
            if worker:
                worker.progress.emit(15, "Reading large JSON file in chunks...")
            
            # For very large JSON files, read and process in chunks
            try:
                # Try to load the JSON file
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                if not isinstance(data, list):
                    data = [data]
                
                total_records = len(data)
                if worker:
                    worker.progress.emit(25, f"Found {total_records:,} records in JSON file")
                
                # Process in chunks
                chunk_size = 50000  # Process 50k records at a time
                table_created = False
                
                for i in range(0, total_records, chunk_size):
                    if worker and worker.cancelled:
                        return False
                    
                    chunk_data = data[i:i+chunk_size]
                    chunk_num = i // chunk_size + 1
                    total_chunks = (total_records - 1) // chunk_size + 1
                    
                    if worker:
                        progress = 30 + int((i / total_records) * 60)
                        worker.progress.emit(progress, f"Processing chunk {chunk_num}/{total_chunks}: {len(chunk_data):,} records")
                    
                    # Convert chunk to DataFrame
                    chunk_df = pd.DataFrame(chunk_data)
                    
                    if chunk_df.empty:
                        continue
                    
                    # Process chunk
                    chunk_df = self.quick_process_dataframe(chunk_df)
                    
                    # Insert chunk
                    if not table_created:
                        success = self.fast_database_insert(chunk_df, table_name, mode, worker)
                        table_created = True
                    else:
                        success = self.fast_database_insert(chunk_df, table_name, 'append', worker)
                    
                    if not success:
                        if worker:
                            worker.error.emit(f"Failed to insert JSON chunk {chunk_num}")
                        return False
                
                if worker:
                    worker.progress.emit(90, f"JSON import completed: {total_records:,} records")
                
                # Refresh schema browser with error handling
                try:
                    self.refresh_schema_browser()
                except Exception as refresh_error:
                    print(f"Schema refresh warning: {refresh_error}")
                    # Import was successful, just refresh failed - continue
                
                print(f"JSON chunked import completed: {total_records:,} records imported to '{table_name}'")
                return True
                
            except json.JSONDecodeError as e:
                if worker:
                    worker.error.emit(f"Invalid JSON format: {str(e)}")
                return False
            except MemoryError:
                if worker:
                    worker.error.emit("JSON file too large to process - consider splitting it into smaller files")
                return False
            
        except Exception as e:
            print(f"JSON chunked import failed: {e}")
            if worker:
                worker.error.emit(f"JSON chunked import failed: {str(e)}")
            return False
    
    def import_small_file_fast(self, import_info, worker=None):
        """Fast import for smaller files using optimized pandas operations"""
        try:
            file_path = import_info['file_path']
            table_name = import_info['table_name']
            file_type = import_info['file_type']
            mode = import_info['mode']
            
            if worker:
                worker.progress.emit(20, "Loading file into memory...")
            
            # Load data with optimized settings
            df = self.safe_load_data_optimized(file_path, file_type, import_info)
            
            if df is None or df.empty:
                if worker:
                    worker.error.emit("No data found in the file.")
                return False
            
            if worker:
                worker.progress.emit(40, f"Processing {len(df):,} rows...")
            
            # Quick data processing
            df = self.quick_process_dataframe(df)
            
            if worker:
                worker.progress.emit(60, "Preparing database insert...")
            
            # Ensure unique table name
            safe_table_name = self.ensure_unique_table_name(table_name, mode)
            
            if worker:
                worker.progress.emit(80, f"Inserting data into '{safe_table_name}'...")
            
            # Fast database insert
            success = self.fast_database_insert(df, safe_table_name, mode, worker)
            
            if success:
                if worker:
                    worker.progress.emit(95, "Finalizing import...")
                
                # Update schema browser
                self.refresh_schema_browser()
                
                print(f"Fast import completed: {len(df):,} rows imported to '{safe_table_name}'")
                return True
            else:
                return False
                
        except Exception as e:
            print(f"Fast import failed: {e}")
            if worker:
                worker.error.emit(f"Fast import failed: {str(e)}")
            return False
    
    def import_large_file_chunked(self, import_info, worker=None):
        """Memory-efficient chunked import for large files"""
        try:
            file_path = import_info['file_path']
            table_name = import_info['table_name']
            file_type = import_info['file_type']
            mode = import_info['mode']
            
            # Determine optimal chunk size based on file size
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            if file_size_mb > 500:
                chunk_size = 50000  # 50K rows for very large files
            elif file_size_mb > 100:
                chunk_size = 100000  # 100K rows for large files
            else:
                chunk_size = 200000  # 200K rows for medium files
            
            if worker:
                worker.progress.emit(15, f"Processing large file in chunks of {chunk_size:,} rows...")
            
            # Ensure unique table name
            safe_table_name = self.ensure_unique_table_name(table_name, mode)
            
            # Handle table creation/replacement
            if mode == 'replace':
                try:
                    if self.current_connection_info['type'].lower() == 'duckdb':
                        self.current_connection.execute(f"DROP TABLE IF EXISTS {safe_table_name}")
                    else:  # SQLite
                        self.current_connection.execute(f"DROP TABLE IF EXISTS {safe_table_name}")
                        self.current_connection.commit()
                except:
                    pass
            
            total_rows = 0
            chunk_num = 0
            table_created = False
            
            # Process file in chunks
            for chunk_df in self.read_file_chunks(file_path, file_type, import_info, chunk_size):
                if worker and worker.cancelled:
                    return False
                
                chunk_num += 1
                chunk_rows = len(chunk_df)
                total_rows += chunk_rows
                
                if worker:
                    progress = min(90, 15 + (chunk_num * 5))  # Gradual progress
                    worker.progress.emit(progress, f"Processing chunk {chunk_num}: {chunk_rows:,} rows (Total: {total_rows:,})")
                
                # Quick process chunk
                chunk_df = self.quick_process_dataframe(chunk_df)
                
                # Insert chunk
                if not table_created:
                    # First chunk creates the table
                    success = self.fast_database_insert(chunk_df, safe_table_name, mode, worker)
                    table_created = True
                else:
                    # Subsequent chunks append
                    success = self.fast_database_insert(chunk_df, safe_table_name, 'append', worker)
                
                if not success:
                    print(f"Failed to insert chunk {chunk_num}")
                    return False
                
                # Clear memory
                del chunk_df
            
            if worker:
                worker.progress.emit(95, "Finalizing large file import...")
            
            # Update schema browser
            self.refresh_schema_browser()
            
            print(f"Chunked import completed: {total_rows:,} rows imported to '{safe_table_name}'")
            return True
            
        except Exception as e:
            print(f"Chunked import failed: {e}")
            if worker:
                worker.error.emit(f"Chunked import failed: {str(e)}")
            return False
    

    

    
    def load_excel_all_sheets(self, file_path):
        """Load all sheets from an Excel file and combine them"""
        try:
            excel_file = pd.ExcelFile(file_path)
            all_sheets = []
            
            for sheet_name in excel_file.sheet_names:
                try:
                    sheet_df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
                    if not sheet_df.empty:
                        # Add sheet name column
                        sheet_df['_sheet_name'] = sheet_name
                        all_sheets.append(sheet_df)
                except Exception as e:
                    print(f"Failed to read sheet '{sheet_name}': {e}")
                    continue
            
            if all_sheets:
                # Combine all sheets
                combined_df = pd.concat(all_sheets, ignore_index=True, sort=False)
                return combined_df
            else:
                return None
                
        except Exception as e:
            print(f"Failed to load Excel file '{file_path}': {e}")
            return None
    
    def load_excel_with_dialog(self, file_path):
        """Load Excel file with user sheet selection"""
        try:
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            
            if len(sheet_names) == 1:
                # Only one sheet, just load it
                return pd.read_excel(file_path, sheet_name=0, dtype=str)
            
            # Multiple sheets - ask user to select
            sheet_name, ok = QInputDialog.getItem(
                self, 
                f"Select Sheet - {os.path.basename(file_path)}", 
                "Choose which sheet to import:",
                sheet_names, 
                0, 
                False
            )
            
            if ok and sheet_name:
                return pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
            else:
                # User cancelled - skip this file
                return None
                
        except Exception as e:
            print(f"Failed to load Excel file '{file_path}': {e}")
            return None
    
    def safe_load_data(self, file_path, file_type, import_info):
        """Safely load data from any file type with robust error handling"""
        df = None
        
        try:
            if file_type == '.csv':
                # Try multiple encoding strategies for CSV
                encodings = [import_info.get('encoding', 'utf-8'), 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                for encoding in encodings:
                    try:
                        df = pd.read_csv(
                            file_path,
                            delimiter=import_info.get('delimiter', ','),
                            encoding=encoding,
                            header=0 if import_info.get('header', True) else None,
                            on_bad_lines='skip',  # Skip bad lines instead of failing
                            low_memory=False,     # Read entire file to avoid dtype warnings
                            dtype=str             # Read everything as string initially
                        )
                        print(f"Successfully loaded CSV with encoding: {encoding}")
                        break
                    except UnicodeDecodeError:
                        continue
                    except Exception as e:
                        if encoding == encodings[-1]:  # Last encoding attempt
                            print(f"CSV load failed with all encodings: {e}")
                            # Try to load with different parameters
                            try:
                                df = pd.read_csv(file_path, delimiter=',', encoding='utf-8', header=None, on_bad_lines='skip', dtype=str)
                            except:
                                pass
            
            elif file_type == '.tsv':
                try:
                    df = pd.read_csv(
                        file_path,
                        delimiter=import_info.get('delimiter', '\t'),
                        encoding=import_info.get('encoding', 'utf-8'),
                        header=0 if import_info.get('header', True) else None,
                        on_bad_lines='skip',
                        dtype=str
                    )
                except:
                    # Fallback
                    df = pd.read_csv(file_path, delimiter='\t', encoding='utf-8', header=None, on_bad_lines='skip', dtype=str)
            
            elif file_type == '.txt':
                try:
                    df = pd.read_csv(
                        file_path,
                        delimiter=import_info.get('delimiter', ','),
                        encoding=import_info.get('encoding', 'utf-8'),
                        header=0 if import_info.get('header', True) else None,
                        on_bad_lines='skip',
                        dtype=str
                    )
                except:
                    # Fallback
                    df = pd.read_csv(file_path, delimiter=',', encoding='utf-8', header=None, on_bad_lines='skip', dtype=str)
            
            elif file_type in ['.xlsx', '.xls']:
                try:
                    sheet_name = import_info.get('sheet_name', 0)
                    df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
                except Exception as e:
                    print(f"Excel load failed: {e}")
                    # Try reading first sheet as fallback
                    try:
                        df = pd.read_excel(file_path, sheet_name=0, dtype=str)
                    except:
                        pass
            
            elif file_type == '.parquet':
                try:
                    df = pd.read_parquet(file_path)
                    # Convert to string to avoid type issues
                    df = df.astype(str)
                except Exception as e:
                    print(f"Parquet load failed: {e}")
            
            elif file_type == '.json':
                try:
                    df = pd.read_json(file_path)
                    # Convert to string to avoid type issues
                    df = df.astype(str)
                except Exception as e:
                    print(f"JSON load failed: {e}")
                    # Try alternative JSON loading
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            import json
                            data = json.load(f)
                            if isinstance(data, list):
                                df = pd.DataFrame(data)
                            elif isinstance(data, dict):
                                df = pd.DataFrame([data])
                            df = df.astype(str)
                    except:
                        pass
            
        except Exception as e:
            print(f"Failed to load {file_type} file: {e}")
            
        return df
    
    def safe_load_data_optimized(self, file_path, file_type, import_info):
        """Optimized data loading with performance enhancements"""
        try:
            if file_type == '.csv' or file_type == '.txt':
                # Optimized CSV/TXT loading
                return pd.read_csv(
                    file_path,
                    delimiter=import_info.get('delimiter', ','),
                    encoding=import_info.get('encoding', 'utf-8'),
                    header=0 if import_info.get('header', True) else None,
                    on_bad_lines='skip',
                    low_memory=False,
                    dtype=str,  # Read as strings to avoid type inference overhead
                    engine='c',  # Use faster C engine
                    na_filter=False  # Don't convert to NaN, keep as strings
                )
            
            elif file_type == '.tsv':
                # TSV loading
                return pd.read_csv(
                    file_path,
                    delimiter='\t',
                    encoding=import_info.get('encoding', 'utf-8'),
                    header=0 if import_info.get('header', True) else None,
                    on_bad_lines='skip',
                    low_memory=False,
                    dtype=str,
                    engine='c',
                    na_filter=False
                )
            
            elif file_type in ['.xlsx', '.xls']:
                # Enhanced Excel loading with better error handling
                print(f"Attempting to load Excel file: {file_path} (type: {file_type})")
                try:
                    # Try with specified sheet first
                    sheet_name = import_info.get('sheet_name', 0)
                    print(f"Trying to load sheet: {sheet_name}")
                    df = pd.read_excel(
                        file_path,
                        sheet_name=sheet_name,
                        dtype=str,
                        na_filter=False,
                        engine='openpyxl' if file_type == '.xlsx' else None
                    )
                    print(f"Successfully loaded Excel file {os.path.basename(file_path)} with {len(df)} rows and {len(df.columns)} columns")
                    if len(df) == 0:
                        print(f"WARNING: Excel file {os.path.basename(file_path)} has 0 rows")
                    return df
                except Exception as e:
                    print(f"Failed to load Excel with sheet {sheet_name}, trying first sheet: {e}")
                    # Fallback to first sheet
                    try:
                        print("Trying to load first sheet (index 0)")
                        df = pd.read_excel(
                            file_path,
                            sheet_name=0,
                            dtype=str,
                            na_filter=False,
                            engine='openpyxl' if file_type == '.xlsx' else None
                        )
                        print(f"Successfully loaded Excel file {os.path.basename(file_path)} (first sheet) with {len(df)} rows and {len(df.columns)} columns")
                        if len(df) == 0:
                            print(f"WARNING: Excel file {os.path.basename(file_path)} first sheet has 0 rows")
                        return df
                    except Exception as e2:
                        print(f"Failed to load Excel file {file_path}: {e2}")
                        # Try without specifying engine
                        try:
                            print("Trying to load Excel without specifying engine")
                            df = pd.read_excel(file_path, sheet_name=0, dtype=str, na_filter=False)
                            print(f"Successfully loaded Excel file {os.path.basename(file_path)} (no engine) with {len(df)} rows and {len(df.columns)} columns")
                            if len(df) == 0:
                                print(f"WARNING: Excel file {os.path.basename(file_path)} (no engine) has 0 rows")
                            return df
                        except Exception as e3:
                            print(f"All Excel loading methods failed for {file_path}: {e3}")
                            return None
            
            elif file_type == '.parquet':
                # Parquet loading
                try:
                    df = pd.read_parquet(file_path)
                    df = df.astype(str)
                    print(f"Successfully loaded Parquet file {os.path.basename(file_path)} with {len(df)} rows")
                    return df
                except Exception as e:
                    print(f"Failed to load Parquet file {file_path}: {e}")
                    return None
            
            elif file_type == '.json':
                # JSON loading with multiple strategies
                try:
                    # Try pandas read_json first
                    if file_path.endswith('.jsonl'):
                        df = pd.read_json(file_path, lines=True)
                    else:
                        df = pd.read_json(file_path)
                    df = df.astype(str)
                    print(f"Successfully loaded JSON file {os.path.basename(file_path)} with {len(df)} rows")
                    return df
                except Exception as e:
                    print(f"Pandas JSON load failed, trying manual parsing: {e}")
                    # Fallback to manual JSON parsing
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            import json
                            data = json.load(f)
                            if isinstance(data, list):
                                df = pd.DataFrame(data)
                            elif isinstance(data, dict):
                                df = pd.DataFrame([data])
                            else:
                                print(f"Unsupported JSON structure in {file_path}")
                                return None
                            df = df.astype(str)
                            print(f"Successfully loaded JSON file {os.path.basename(file_path)} (manual) with {len(df)} rows")
                            return df
                    except Exception as e2:
                        print(f"Failed to load JSON file {file_path}: {e2}")
                        return None
            
            else:
                # Fallback to regular loading
                print(f"Using fallback loading for file type {file_type}")
                return self.safe_load_data(file_path, file_type, import_info)
                
        except Exception as e:
            print(f"Optimized loading failed for {file_path}, falling back to safe loading: {e}")
            return self.safe_load_data(file_path, file_type, import_info)
    
    def read_file_chunks(self, file_path, file_type, import_info, chunk_size):
        """Generator that yields chunks of data from large files"""
        try:
            if file_type == '.csv':
                # CSV chunked reading
                chunk_reader = pd.read_csv(
                    file_path,
                    delimiter=import_info.get('delimiter', ','),
                    encoding=import_info.get('encoding', 'utf-8'),
                    header=0 if import_info.get('header', True) else None,
                    on_bad_lines='skip',
                    dtype=str,
                    engine='c',
                    na_filter=False,
                    chunksize=chunk_size
                )
                
                for chunk in chunk_reader:
                    yield chunk
            
            elif file_type in ['.xlsx', '.xls']:
                # For Excel, we can't easily chunk, so load and split
                df = pd.read_excel(
                    file_path,
                    sheet_name=import_info.get('sheet_name', 0),
                    dtype=str,
                    na_filter=False
                )
                
                # Split into chunks
                for i in range(0, len(df), chunk_size):
                    yield df.iloc[i:i + chunk_size]
            
            else:
                # For other formats, load and split
                df = self.safe_load_data_optimized(file_path, file_type, import_info)
                if df is not None and not df.empty:
                    for i in range(0, len(df), chunk_size):
                        yield df.iloc[i:i + chunk_size]
                        
        except Exception as e:
            print(f"Chunked reading failed: {e}")
            # Fallback: try to load entire file and split
            try:
                df = self.safe_load_data_optimized(file_path, file_type, import_info)
                if df is not None and not df.empty:
                    for i in range(0, len(df), chunk_size):
                        yield df.iloc[i:i + chunk_size]
            except:
                pass
    
    def quick_process_dataframe(self, df):
        """Quick dataframe processing for performance"""
        try:
            # Basic cleaning without heavy sanitization
            if df is None or df.empty:
                return df
            
            # Clean column names quickly
            df.columns = [str(col).strip() if col is not None else f"col_{i}" 
                         for i, col in enumerate(df.columns)]
            
            # Handle duplicate column names
            cols = df.columns.tolist()
            seen = set()
            unique_cols = []
            for col in cols:
                original_col = col
                counter = 1
                while col in seen:
                    col = f"{original_col}_{counter}"
                    counter += 1
                seen.add(col)
                unique_cols.append(col)
            df.columns = unique_cols
            
            # Remove completely empty rows
            df = df.dropna(how='all')
            
            return df
            
        except Exception as e:
            print(f"Quick processing failed: {e}")
            return df
    
    def fast_database_insert(self, df, table_name, mode, worker=None):
        """Optimized database insertion with bulk operations"""
        try:
            if df is None or df.empty:
                return False
            
            db_type = self.current_connection_info['type'].lower()
            
            if db_type == 'duckdb':
                return self.fast_duckdb_insert(df, table_name, mode)
            else:
                return self.fast_sqlite_insert(df, table_name, mode)
                
        except Exception as e:
            print(f"Fast database insert failed: {e}")
            if worker:
                worker.error.emit(f"Database insert failed: {str(e)}")
            return False
    
    def fast_duckdb_insert(self, df, table_name, mode):
        """Optimized DuckDB insertion using native methods"""
        try:
            # Use DuckDB's native pandas integration for maximum speed
            if mode == 'replace':
                # Drop table if exists
                try:
                    self.current_connection.execute(f"DROP TABLE IF EXISTS {table_name}")
                except:
                    pass
                
                # Create and insert in one operation
                self.current_connection.execute(f"CREATE TABLE {table_name} AS SELECT * FROM df")
            
            elif mode == 'append':
                # Check if table exists
                try:
                    result = self.current_connection.execute(f"SELECT 1 FROM {table_name} LIMIT 1").fetchone()
                    table_exists = True
                except:
                    table_exists = False
                
                if not table_exists:
                    # Create table
                    self.current_connection.execute(f"CREATE TABLE {table_name} AS SELECT * FROM df")
                else:
                    # Insert data
                    self.current_connection.execute(f"INSERT INTO {table_name} SELECT * FROM df")
            
            else:  # create mode
                # Create new table
                self.current_connection.execute(f"CREATE TABLE {table_name} AS SELECT * FROM df")
            
            return True
            
        except Exception as e:
            print(f"Fast DuckDB insert failed: {e}")
            # Fallback to regular method
            return self.duckdb_safe_import(df, table_name, mode)
    
    def fast_sqlite_insert(self, df, table_name, mode):
        """Optimized SQLite insertion using bulk operations"""
        try:
            if mode == 'replace':
                # Use pandas to_sql with replace
                df.to_sql(table_name, self.current_connection, if_exists='replace', index=False, method='multi')
            elif mode == 'append':
                # Use pandas to_sql with append
                df.to_sql(table_name, self.current_connection, if_exists='append', index=False, method='multi')
            else:  # create mode
                # Use pandas to_sql with fail (will create new table)
                df.to_sql(table_name, self.current_connection, if_exists='fail', index=False, method='multi')
            
            self.current_connection.commit()
            return True
            
        except Exception as e:
            print(f"Fast SQLite insert failed: {e}")
            # Fallback to regular method
            return self.safe_import_to_database(df, table_name, mode)
    
    def combine_dataframes_with_alignment(self, dataframes):
        """Combine multiple dataframes with proper column alignment - new columns go to the end"""
        if not dataframes:
            return pd.DataFrame()
        
        if len(dataframes) == 1:
            return dataframes[0]
        
        try:
            # Get the first dataframe's columns as the base order
            base_columns = list(dataframes[0].columns)
            
            # Find all unique column names across all dataframes
            all_columns_set = set(base_columns)
            for df in dataframes[1:]:
                all_columns_set.update(df.columns)
            
            # Create final column order: base columns first, then new columns at the end
            new_columns = [col for col in all_columns_set if col not in base_columns]
            all_columns = base_columns + sorted(new_columns)  # Sort only the new columns
            
            print(f"Found {len(all_columns)} unique columns across all files")
            print(f"Base columns ({len(base_columns)}): {base_columns}")
            if new_columns:
                print(f"New columns added at end ({len(new_columns)}): {sorted(new_columns)}")
            
            # Align all dataframes to have the same columns
            aligned_dataframes = []
            for i, df in enumerate(dataframes):
                # Create a new dataframe with all columns in correct order
                aligned_df = pd.DataFrame(index=df.index)
                
                # Copy existing columns in their original order
                for col in df.columns:
                    aligned_df[col] = df[col]
                
                # Add missing columns with None values at the end
                for col in all_columns:
                    if col not in aligned_df.columns:
                        aligned_df[col] = None
                
                # Reorder columns: existing first, new at the end
                aligned_df = aligned_df[all_columns]
                aligned_dataframes.append(aligned_df)
                print(f"Aligned dataframe {i+1}: {len(aligned_df)} rows, {len(aligned_df.columns)} columns")
            
            # Combine all aligned dataframes
            combined_df = pd.concat(aligned_dataframes, ignore_index=True, sort=False)
            
            # Convert all data to strings for database compatibility
            for col in combined_df.columns:
                combined_df[col] = combined_df[col].astype(str)
                # Replace 'None' strings with actual None
                combined_df[col] = combined_df[col].replace('None', None)
            
            print(f"Successfully combined {len(dataframes)} dataframes into {len(combined_df)} rows")
            return combined_df
            
        except Exception as e:
            print(f"Error combining dataframes: {e}")
            # Fallback: try simple concatenation without sorting columns
            try:
                combined_df = pd.concat(dataframes, ignore_index=True, sort=False)
                # Fill missing values
                combined_df = combined_df.fillna('')
                # Convert to strings
                for col in combined_df.columns:
                    combined_df[col] = combined_df[col].astype(str)
                return combined_df
            except Exception as e2:
                print(f"Fallback concatenation also failed: {e2}")
                raise e
    
    def sanitize_dataframe(self, df):
        """Sanitize dataframe to prevent any import errors by converting problematic data to text"""
        try:
            print(f"Sanitizing dataframe with {len(df)} rows and {len(df.columns)} columns")
            
            # Create a copy to avoid modifying the original
            df_clean = df.copy()
            
            # Handle column names - ensure they're clean
            df_clean.columns = [str(col).strip() if col is not None else f"col_{i}" for i, col in enumerate(df_clean.columns)]
            
            # Process each column
            for col in df_clean.columns:
                try:
                    # Convert the entire column to string initially
                    df_clean[col] = df_clean[col].astype(str)
                    
                    # Handle specific problematic values
                    df_clean[col] = df_clean[col].replace({
                        'nan': None,
                        'NaN': None,
                        'None': None,
                        'null': None,
                        'NULL': None,
                        '': None,
                        'inf': 'infinity',
                        '-inf': '-infinity',
                        'Infinity': 'infinity',
                        '-Infinity': '-infinity'
                    })
                    
                    # Remove any remaining problematic characters
                    df_clean[col] = df_clean[col].apply(lambda x: self.safe_string_convert(x) if x is not None else None)
                    
                except Exception as e:
                    print(f"Error processing column {col}: {e}")
                    # If anything fails, convert entire column to safe strings
                    df_clean[col] = df_clean[col].apply(lambda x: str(x) if x is not None else None)
            
            # Remove completely empty rows
            df_clean = df_clean.dropna(how='all')
            
            # Ensure no column names are duplicated
            cols = df_clean.columns.tolist()
            seen = set()
            unique_cols = []
            for col in cols:
                original_col = col
                counter = 1
                while col in seen:
                    col = f"{original_col}_{counter}"
                    counter += 1
                seen.add(col)
                unique_cols.append(col)
            df_clean.columns = unique_cols
            
            print(f"Sanitization complete: {len(df_clean)} rows × {len(df_clean.columns)} columns")
            return df_clean
            
        except Exception as e:
            print(f"Sanitization failed: {e}")
            # Last resort: convert everything to string
            try:
                df_safe = pd.DataFrame()
                for i, col in enumerate(df.columns):
                    col_name = f"column_{i}" if col is None or str(col).strip() == '' else str(col)
                    df_safe[col_name] = df.iloc[:, i].apply(lambda x: str(x) if x is not None else None)
                return df_safe
            except:
                # Ultimate fallback: return empty dataframe with at least one column
                return pd.DataFrame({'data': ['No data could be imported']})
    
    def detect_csv_delimiter(self, file_path, sample_size=1024):
        """Automatically detect the delimiter used in a CSV file"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                sample = file.read(sample_size)
                
            # Use csv.Sniffer to detect delimiter
            if csv:
                sniffer = csv.Sniffer()
                try:
                    dialect = sniffer.sniff(sample, delimiters=',;\t|')
                    detected_delimiter = dialect.delimiter
                    print(f"Auto-detected delimiter: '{detected_delimiter}' for file: {os.path.basename(file_path)}")
                    return detected_delimiter
                except:
                    pass
            
            # Fallback: count common delimiters and choose the most frequent
            delimiters = {',': 0, ';': 0, '\t': 0, '|': 0}
            for delimiter in delimiters:
                delimiters[delimiter] = sample.count(delimiter)
            
            # Choose the delimiter with the highest count (but at least 1)
            best_delimiter = max(delimiters, key=delimiters.get)
            if delimiters[best_delimiter] > 0:
                print(f"Fallback delimiter detection: '{best_delimiter}' for file: {os.path.basename(file_path)}")
                return best_delimiter
            
            # Default to comma if no delimiter found
            print(f"No delimiter detected, defaulting to comma for file: {os.path.basename(file_path)}")
            return ','
            
        except Exception as e:
            print(f"Error detecting delimiter for {file_path}: {e}")
            return ','
    
    def safe_string_convert(self, value):
        """Safely convert any value to a database-compatible string"""
        try:
            if value is None or pd.isna(value):
                return None
            
            # Handle different types
            if isinstance(value, (int, float)):
                if pd.isna(value) or pd.isinf(value):
                    return None
                return str(value)
            
            # Convert to string and handle encoding issues
            str_val = str(value)
            
            # Remove or replace problematic characters
            str_val = str_val.replace('\x00', '')  # Remove null bytes
            str_val = str_val.replace('\r\n', '\n')  # Normalize line endings
            str_val = str_val.replace('\r', '\n')
            
            # Limit string length to prevent database issues
            if len(str_val) > 10000:  # Reasonable limit
                str_val = str_val[:10000] + "...[truncated]"
            
            return str_val
            
        except Exception:
            return "conversion_error"
    
    def safe_import_to_database(self, df, table_name, mode):
        """Safely import dataframe to database with multiple fallback strategies"""
        try:
            db_type = self.current_connection_info['type'].lower()
            
            # Strategy 1: Try database-specific import
            try:
                if db_type == 'duckdb':
                    # Use DuckDB-specific import method to avoid pandas to_sql issues
                    return self.duckdb_safe_import(df, table_name, mode)
                else:
                    # SQLite import - use safer method without 'multi'
                    if mode == 'create':
                        # Check if table exists for create mode
                        cursor = self.current_connection.cursor()
                        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                        if cursor.fetchone():
                            raise ValueError(f"Table '{table_name}' already exists. Use 'Replace' mode to overwrite or 'Append' to add data.")
                        df.to_sql(table_name, self.current_connection, if_exists='fail', index=False)
                    elif mode == 'append':
                        self.flexible_append_data(df, table_name, 'sqlite')
                    else:  # replace
                        df.to_sql(table_name, self.current_connection, if_exists='replace', index=False)
                    self.current_connection.commit()
                
                print("Normal import successful")
                return True
                
            except Exception as e:
                print(f"Normal import failed: {e}")
                
                # Strategy 2: Force all columns to TEXT type
                try:
                    print("Trying with all TEXT columns...")
                    
                    # Create table manually with all TEXT columns
                    columns_sql = ", ".join([f'"{col}" TEXT' for col in df.columns])
                    
                    if mode == 'replace':
                        self.current_connection.execute(f"DROP TABLE IF EXISTS {table_name}")
                    
                    create_sql = f"CREATE TABLE IF NOT EXISTS {table_name} ({columns_sql})"
                    self.current_connection.execute(create_sql)
                    
                    if db_type == 'sqlite':
                        self.current_connection.commit()
                    
                    # Insert data row by row if needed
                    df.to_sql(table_name, self.current_connection, if_exists='append', index=False)
                    
                    if db_type == 'sqlite':
                        self.current_connection.commit()
                    
                    print("TEXT columns import successful")
                    return True
                    
                except Exception as e2:
                    print(f"TEXT columns import failed: {e2}")
                    
                    # Strategy 3: Row-by-row insert with error handling
                    try:
                        print("Trying row-by-row insert...")
                        
                        # Ensure table exists
                        if mode == 'replace':
                            self.current_connection.execute(f"DROP TABLE IF EXISTS {table_name}")
                        
                        # Create table with generic structure
                        columns_sql = ", ".join([f'"{col}" TEXT' for col in df.columns])
                        create_sql = f"CREATE TABLE IF NOT EXISTS {table_name} ({columns_sql})"
                        self.current_connection.execute(create_sql)
                        
                        if db_type == 'sqlite':
                            self.current_connection.commit()
                        
                        # Insert rows one by one, skipping problematic ones
                        successful_rows = 0
                        placeholders = ", ".join(["?" for _ in df.columns])
                        insert_sql = f'INSERT INTO {table_name} VALUES ({placeholders})'
                        
                        for idx, row in df.iterrows():
                            try:
                                values = [self.safe_string_convert(val) for val in row.values]
                                self.current_connection.execute(insert_sql, values)
                                successful_rows += 1
                                
                                if successful_rows % 1000 == 0:
                                    if db_type == 'sqlite':
                                        self.current_connection.commit()
                                    print(f"Inserted {successful_rows} rows...")
                                    
                            except Exception as row_error:
                                print(f"Skipped row {idx}: {row_error}")
                                continue
                        
                        if db_type == 'sqlite':
                            self.current_connection.commit()
                        
                        print(f"Row-by-row insert completed: {successful_rows} rows inserted")
                        return True
                        
                    except Exception as e3:
                        print(f"Row-by-row insert failed: {e3}")
                        return False
        
        except Exception as e:
            print(f"Safe import completely failed: {e}")
            return False
    
    def duckdb_safe_import(self, df, table_name, mode):
        """DuckDB-specific import method to handle transactions properly"""
        try:
            # For DuckDB, avoid pandas to_sql which can cause transaction issues
            # Instead, create table manually and use INSERT statements
            
            if mode == 'create':
                # Check if table exists for create mode
                try:
                    result = self.current_connection.execute(f"SELECT 1 FROM {table_name} LIMIT 1").fetchone()
                    if result is not None:
                        raise ValueError(f"Table '{table_name}' already exists. Use 'Replace' mode to overwrite or 'Append' to add data.")
                except Exception as e:
                    if "does not exist" not in str(e).lower() and "no such table" not in str(e).lower():
                        raise e  # Re-raise if it's not a "table doesn't exist" error
                    # Table doesn't exist, which is what we want for create mode
            elif mode == 'replace':
                try:
                    self.current_connection.execute(f"DROP TABLE IF EXISTS {table_name}")
                except:
                    pass
            
            # Create table with all TEXT columns for safety
            columns_sql = ", ".join([f'"{col}" TEXT' for col in df.columns])
            
            if mode in ['create', 'replace']:
                create_sql = f"CREATE TABLE {table_name} ({columns_sql})"
                self.current_connection.execute(create_sql)
            elif mode == 'append':
                # For append mode, table should already exist
                try:
                    result = self.current_connection.execute(f"SELECT 1 FROM {table_name} LIMIT 1").fetchone()
                except Exception as e:
                    if "does not exist" in str(e).lower() or "no such table" in str(e).lower():
                        raise ValueError(f"Table '{table_name}' does not exist. Use 'Create' mode to create a new table.")
                    raise e
            
            # Insert data using DuckDB's efficient INSERT
            placeholders = ", ".join(["?" for _ in df.columns])
            insert_sql = f'INSERT INTO {table_name} VALUES ({placeholders})'
            
            # Convert all data to strings to avoid type issues
            data_rows = []
            for _, row in df.iterrows():
                converted_row = [self.safe_string_convert(val) for val in row.values]
                data_rows.append(converted_row)
            
            # Batch insert for efficiency
            batch_size = 1000
            for i in range(0, len(data_rows), batch_size):
                batch = data_rows[i:i + batch_size]
                try:
                    self.current_connection.executemany(insert_sql, batch)
                except Exception as batch_error:
                    print(f"Batch insert failed, trying row by row: {batch_error}")
                    # Fall back to individual inserts for this batch
                    for row in batch:
                        try:
                            self.current_connection.execute(insert_sql, row)
                        except Exception as row_error:
                            print(f"Skipped problematic row: {row_error}")
                            continue
            
            print(f"DuckDB import successful: {len(df)} rows")
            return True
            
        except Exception as e:
            print(f"DuckDB import failed: {e}")
            return self.row_by_row_insert_fallback(df, table_name, mode)
    
    def row_by_row_insert_fallback(self, df, table_name, mode):
        """Final fallback method for problematic imports"""
        try:
            db_type = self.current_connection_info['type'].lower()
            
            # Drop and recreate table for replace mode
            if mode == 'replace':
                try:
                    self.current_connection.execute(f"DROP TABLE IF EXISTS {table_name}")
                    if db_type == 'sqlite':
                        self.current_connection.commit()
                except:
                    pass
            
            # Create table with all TEXT columns
            columns_sql = ", ".join([f'"{col}" TEXT' for col in df.columns])
            create_sql = f"CREATE TABLE IF NOT EXISTS {table_name} ({columns_sql})"
            self.current_connection.execute(create_sql)
            
            if db_type == 'sqlite':
                self.current_connection.commit()
            
            # Insert rows one by one with error handling
            successful_rows = 0
            placeholders = ", ".join(["?" for _ in df.columns])
            insert_sql = f'INSERT INTO {table_name} VALUES ({placeholders})'
            
            for idx, row in df.iterrows():
                try:
                    values = [self.safe_string_convert(val) for val in row.values]
                    self.current_connection.execute(insert_sql, values)
                    successful_rows += 1
                    
                    # Commit periodically for SQLite
                    if successful_rows % 500 == 0 and db_type == 'sqlite':
                        self.current_connection.commit()
                        
                except Exception as row_error:
                    print(f"Skipped row {idx}: {row_error}")
                    continue
            
            if db_type == 'sqlite':
                self.current_connection.commit()
            
            print(f"Fallback import completed: {successful_rows}/{len(df)} rows inserted")
            return successful_rows > 0
            
        except Exception as e:
            print(f"Fallback import failed: {e}")
            return False
    
    def flexible_append_data(self, df, table_name, db_type):
        """Append data with flexible column handling - adds missing columns automatically and handles all errors"""
        try:
            # Get existing table schema with error handling
            existing_columns = set()
            table_exists = False
            
            try:
                if db_type == 'duckdb':
                    existing_columns_df = self.current_connection.execute(f"PRAGMA table_info('{table_name}')").fetchdf()
                    existing_columns = set(existing_columns_df['name'].tolist())
                    table_exists = True
                else:  # sqlite
                    cursor = self.current_connection.cursor()
                    cursor.execute(f"PRAGMA table_info({table_name})")
                    existing_columns = set([row[1] for row in cursor.fetchall()])
                    table_exists = True
            except:
                table_exists = False
            
            if not table_exists:
                # Table doesn't exist, create it with all TEXT columns for safety
                try:
                    columns_sql = ", ".join([f'"{col}" TEXT' for col in df.columns])
                    create_sql = f"CREATE TABLE {table_name} ({columns_sql})"
                    self.current_connection.execute(create_sql)
                    
                    if db_type == 'sqlite':
                        self.current_connection.commit()
                    
                    # Now insert the data
                    df.to_sql(table_name, self.current_connection, if_exists='append', index=False)
                    
                    if db_type == 'sqlite':
                        self.current_connection.commit()
                    return
                except Exception as e:
                    print(f"Failed to create new table: {e}")
                    # Use row-by-row insert as fallback
                    return self.row_by_row_insert_fallback(df, table_name, 'create')
            
            # Get new columns from the dataframe
            new_columns = set(df.columns)
            
            # Find columns that need to be added to the existing table
            missing_columns = new_columns - existing_columns
            
            # Add missing columns to the existing table (always as TEXT for safety)
            if missing_columns:
                print(f"Adding new columns to table '{table_name}': {', '.join(missing_columns)}")
                for col in missing_columns:
                    try:
                        # Always use TEXT type to avoid conflicts
                        alter_sql = f'ALTER TABLE {table_name} ADD COLUMN "{col}" TEXT'
                        self.current_connection.execute(alter_sql)
                    except Exception as alter_error:
                        print(f"Failed to add column {col}: {alter_error}")
                        # Continue with other columns
                        
                if db_type == 'sqlite':
                    self.current_connection.commit()
            
            # Find columns in the table that are not in the new data
            extra_table_columns = existing_columns - new_columns
            
            # Add missing columns to the dataframe with NULL values
            for col in extra_table_columns:
                df[col] = None
            
            # Reorder dataframe columns to match table schema: existing columns first, new columns at the end
            all_columns = list(existing_columns) + sorted(list(missing_columns))
            df = df.reindex(columns=all_columns)
            
            # Convert all data to strings to prevent type conflicts
            for col in df.columns:
                df[col] = df[col].apply(lambda x: self.safe_string_convert(x) if x is not None else None)
            
            # Try to append the data
            try:
                df.to_sql(table_name, self.current_connection, if_exists='append', index=False)
                
                if db_type == 'sqlite':
                    self.current_connection.commit()
                    
            except Exception as append_error:
                print(f"Bulk append failed: {append_error}")
                
                # Try row-by-row insert as fallback
                try:
                    successful_rows = 0
                    placeholders = ", ".join(["?" for _ in df.columns])
                    insert_sql = f'INSERT INTO {table_name} VALUES ({placeholders})'
                    
                    for idx, row in df.iterrows():
                        try:
                            values = [self.safe_string_convert(val) for val in row.values]
                            self.current_connection.execute(insert_sql, values)
                            successful_rows += 1
                            
                            if successful_rows % 100 == 0 and db_type == 'sqlite':
                                self.current_connection.commit()
                                
                        except Exception as row_error:
                            print(f"Skipped problematic row {idx}: {row_error}")
                            continue
                    
                    if db_type == 'sqlite':
                        self.current_connection.commit()
                        
                    print(f"Flexible append completed: {successful_rows} rows inserted")
                    
                except Exception as row_error:
                    print(f"Row-by-row insert also failed: {row_error}")
                    raise
                
        except Exception as e:
            # Final fallback to safe import
            print(f"Flexible append completely failed: {e}")
            return self.safe_import_to_database(df, table_name, 'append')
    
    def clean_column_name(self, column_name):
        """Clean column name for database compatibility - capitalize and replace special chars with underscores"""
        # Convert to string and strip whitespace
        clean_name = str(column_name).strip()
        
        # Convert to uppercase
        clean_name = clean_name.upper()
        
        # Replace spaces and special characters with underscores
        clean_name = re.sub(r'[^A-Z0-9_]', '_', clean_name)
        
        # Remove consecutive underscores
        clean_name = re.sub(r'_+', '_', clean_name)
        
        # Remove leading/trailing underscores
        clean_name = clean_name.strip('_')
        
        # Ensure it starts with a letter or underscore
        if clean_name and not clean_name[0].isalpha() and clean_name[0] != '_':
            clean_name = f"COL_{clean_name}"
        
        # Handle empty names
        if not clean_name:
            clean_name = "UNNAMED_COLUMN"
        
        return clean_name
    
    def ensure_unique_table_name(self, base_name, mode='create'):
        """Ensure table name is unique and safe for database operations"""
        if not base_name:
            base_name = "imported_table"
        
        # Clean the base name
        clean_base = self.clean_table_name(base_name)
        
        # If we're creating a new table, check for conflicts
        if mode == 'create':
            return self.get_unique_table_name(clean_base)
        else:
            # For append/replace modes, just clean the name
            return clean_base
    
    def clean_table_name(self, table_name):
        """Clean table name to be database-safe"""
        if not table_name:
            return "imported_table"
        
        # Convert to string and strip whitespace
        clean_name = str(table_name).strip()
        
        # Replace problematic characters with underscores
        clean_name = re.sub(r'[^\w\s]', '_', clean_name)
        
        # Replace whitespace with underscores
        clean_name = re.sub(r'\s+', '_', clean_name)
        
        # Remove multiple consecutive underscores
        clean_name = re.sub(r'_+', '_', clean_name)
        
        # Remove leading/trailing underscores
        clean_name = clean_name.strip('_')
        
        # Ensure it doesn't start with a number
        if clean_name and clean_name[0].isdigit():
            clean_name = f"table_{clean_name}"
        
        # Ensure it's not empty and not too long
        if not clean_name:
            clean_name = "imported_table"
        elif len(clean_name) > 50:  # Reasonable limit for table names
            clean_name = clean_name[:50].rstrip('_')
        
        # Check against SQL reserved words
        sql_reserved = {
            'select', 'from', 'where', 'insert', 'update', 'delete', 'create', 'drop', 
            'alter', 'table', 'index', 'view', 'database', 'schema', 'primary', 'key',
            'foreign', 'references', 'constraint', 'unique', 'not', 'null', 'default',
            'check', 'order', 'by', 'group', 'having', 'union', 'join', 'inner', 'outer',
            'left', 'right', 'on', 'as', 'distinct', 'count', 'sum', 'avg', 'max', 'min'
        }
        
        if clean_name.lower() in sql_reserved:
            clean_name = f"tbl_{clean_name}"
        
        return clean_name.lower()
    
    def get_unique_table_name(self, base_name):
        """Get a unique table name by checking existing tables and adding suffix if needed"""
        if not self.current_connection or not self.current_connection_info:
            return base_name
        
        try:
            # Get existing table names
            existing_tables = set()
            
            if self.current_connection_info['type'].lower() == 'duckdb':
                tables_df = self.current_connection.execute("SHOW TABLES").fetchdf()
                existing_tables = set(tables_df['name'].str.lower()) if not tables_df.empty else set()
            else:  # SQLite
                cursor = self.current_connection.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                existing_tables = {row[0].lower() for row in cursor.fetchall()}
            
            # If base name is unique, use it
            if base_name.lower() not in existing_tables:
                return base_name
            
            # Find unique name with suffix
            counter = 1
            while True:
                candidate = f"{base_name}_{counter}"
                if candidate.lower() not in existing_tables:
                    return candidate
                counter += 1
                
                # Safety check to prevent infinite loop
                if counter > 1000:
                    import time
                    timestamp = int(time.time())
                    return f"{base_name}_{timestamp}"
                    
        except Exception as e:
            print(f"Error checking table uniqueness: {e}")
            # Fallback to timestamp-based naming
            import time
            timestamp = int(time.time())
            return f"{base_name}_{timestamp}"
    
    def validate_import_data(self, df, table_name, mode):
        """Validate data before import and provide helpful error messages"""
        errors = []
        warnings = []
        
        # Check if dataframe is empty
        if df is None or df.empty:
            errors.append("No data to import - the file appears to be empty")
            return errors, warnings
        
        # Check table name
        if not table_name or not table_name.strip():
            errors.append("Table name cannot be empty")
        elif len(table_name.strip()) > 50:
            warnings.append(f"Table name is very long ({len(table_name)} chars) - it will be truncated")
        
        # Check column names
        if df.columns.empty:
            errors.append("No columns found in the data")
        else:
            # Check for duplicate column names
            column_counts = df.columns.value_counts()
            duplicates = column_counts[column_counts > 1]
            if not duplicates.empty:
                warnings.append(f"Duplicate column names found: {list(duplicates.index)} - they will be renamed")
            
            # Check for problematic column names
            problematic_cols = []
            for col in df.columns:
                if not str(col).strip():
                    problematic_cols.append("(empty column name)")
                elif str(col).strip().isdigit():
                    problematic_cols.append(f"'{col}' (starts with number)")
            
            if problematic_cols:
                warnings.append(f"Problematic column names will be cleaned: {problematic_cols}")
        
        # Check data size
        row_count = len(df)
        col_count = len(df.columns)
        
        if row_count == 0:
            warnings.append("No data rows found (only headers)")
        elif row_count > 1000000:
            warnings.append(f"Large dataset ({row_count:,} rows) - import may take some time")
        
        if col_count > 100:
            warnings.append(f"Many columns ({col_count}) - consider if all are needed")
        
        # Check for mode-specific issues
        if mode in ['append', 'replace'] and self.current_connection:
            try:
                # Check if target table exists for append/replace
                table_exists = False
                if self.current_connection_info['type'].lower() == 'duckdb':
                    result = self.current_connection.execute(f"SELECT COUNT(*) FROM information_schema.tables WHERE table_name = '{table_name}'").fetchone()
                    table_exists = result[0] > 0 if result else False
                else:  # SQLite
                    cursor = self.current_connection.cursor()
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                    table_exists = cursor.fetchone() is not None
                
                if not table_exists:
                    if mode == 'append':
                        errors.append(f"Cannot append to table '{table_name}' - table does not exist")
                    elif mode == 'replace':
                        warnings.append(f"Table '{table_name}' does not exist - will create new table")
                        
            except Exception as e:
                warnings.append(f"Could not verify table existence: {str(e)}")
        
        return errors, warnings
    
    def handle_duplicate_columns(self, df):
        """Handle duplicate column names by renaming them"""
        if df is None or df.empty:
            return df
        
        # Get column names and find duplicates
        columns = list(df.columns)
        seen = {}
        new_columns = []
        
        for col in columns:
            col_str = str(col).strip()
            col_lower = col_str.lower()
            
            if col_lower in seen:
                # This is a duplicate, add a suffix
                seen[col_lower] += 1
                new_col = f"{col_str}_{seen[col_lower]}"
            else:
                # First occurrence
                seen[col_lower] = 0
                new_col = col_str
            
            new_columns.append(new_col)
        
        # Update dataframe columns
        df.columns = new_columns
        return df
    
    def execute_current_query(self):
        current_tab = self.tab_widget.currentWidget()
        if current_tab:
            current_tab.execute_query()
    
    def execute_selected_query(self):
        """Execute only the selected text as a query"""
        current_tab = self.tab_widget.currentWidget()
        if current_tab:
            cursor = current_tab.editor.textCursor()
            if cursor.hasSelection():
                current_tab.execute_query()  # Will automatically detect selection
            else:
                # If no selection, show a message
                self.statusBar().showMessage("No text selected. Select SQL code to execute or use F5 to run the entire query.", 3000)
    
    def export_current_results(self):
        """Export results from the current tab"""
        current_tab = self.tab_widget.currentWidget()
        if current_tab and hasattr(current_tab, 'show_export_menu'):
            current_tab.show_export_menu()
        else:
            QMessageBox.information(self, "No Results", "No query results to export. Please run a query first.")
    
    def save_query(self):
        current_tab = self.tab_widget.currentWidget()
        if not current_tab:
            return
            
        # Get the query text
        query_text = current_tab.editor.toPlainText()
        if not query_text.strip():
            return
            
        # Get file path
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Query", "", "SQL Files (*.sql);;All Files (*)"
        )
        
        if file_path:
            try:
                with open(file_path, 'w') as f:
                    f.write(query_text)
                self.statusBar().showMessage(f"Query saved to {file_path}", 3000)
                
                # Update tab title
                self.tab_widget.setTabText(
                    self.tab_widget.currentIndex(), os.path.basename(file_path)
                )
            except Exception as e:
                QMessageBox.critical(self, "Save Error", f"Failed to save query: {str(e)}")
    
    def save_query_as(self):
        self.save_query()
    
    def open_query(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Query", "", "SQL Files (*.sql);;All Files (*)"
        )
        
        if file_path:
            try:
                with open(file_path, 'r') as f:
                    query_text = f.read()
                
                # Create new tab or use current if empty
                current_tab = self.tab_widget.currentWidget()
                if current_tab and not current_tab.editor.toPlainText().strip():
                    tab = current_tab
                    tab_index = self.tab_widget.currentIndex()
                else:
                    tab = QueryTab(connection=self.current_connection, connection_info=self.current_connection_info)
                    tab_index = self.tab_widget.addTab(tab, os.path.basename(file_path))
                    self.tab_widget.setCurrentIndex(tab_index)
                
                # Set query text
                tab.editor.setPlainText(query_text)
                self.tab_widget.setTabText(tab_index, os.path.basename(file_path))
                self.statusBar().showMessage(f"Opened {file_path}", 3000)
            except Exception as e:
                QMessageBox.critical(self, "Open Error", f"Failed to open query: {str(e)}")

# Main entry point
def main():
    app = QApplication(sys.argv)
    window = SQLEditorApp()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
                