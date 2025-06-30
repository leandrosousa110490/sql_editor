"""
CSV Automation Module - Process Multiple CSV Sources with SQL
"""

import os
import glob
import pandas as pd
import duckdb
import time
import logging
import json
import gc
import psutil
import re
from typing import List, Dict, Optional
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import sqlite3

# Polars import for faster Excel processing
try:
    import polars as pl
    POLARS_AVAILABLE = True
except ImportError:
    POLARS_AVAILABLE = False
    print("Polars not installed. Using pandas fallback for Excel processing.")

from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QFileDialog,
    QProgressBar, QTextEdit, QGroupBox, QListWidget, QLineEdit, QTabWidget, 
    QWidget, QScrollArea, QFormLayout, QPlainTextEdit, QFrame, QMessageBox,
    QListWidgetItem, QInputDialog, QComboBox
)
from PyQt6.QtCore import QThread, pyqtSignal, Qt
from PyQt6.QtGui import QFont
import qtawesome as qta

from csv_merger import append_csv_files, get_csv_info

# Import SQL editor components from main app
try:
    from app import SQLTextEdit, SQLHighlighter, ColorScheme
    SQL_EDITOR_AVAILABLE = True
except ImportError:
    SQL_EDITOR_AVAILABLE = False

logger = logging.getLogger(__name__)


def clean_column_name(name: str) -> str:
    """Clean column name for SQL compatibility - capitalize and replace special chars with underscores"""
    name = str(name).strip()
    
    # Convert to uppercase
    name = name.upper()
    
    # Replace spaces and special characters with underscores
    name = re.sub(r'[^A-Z0-9_]', '_', name)  # Replace non-alphanumeric chars with underscore
    name = re.sub(r'_+', '_', name)          # Replace multiple underscores with single
    name = name.strip('_')                   # Remove leading/trailing underscores
    
    # Ensure it doesn't start with a number
    if name and name[0].isdigit():
        name = f"COL_{name}"
    
    return name or "UNNAMED_COLUMN"


def read_excel_optimized(file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """Read Excel file using Polars for maximum speed, fallback to pandas with robust error handling"""
    max_retries = 3
    retry_count = 0
    
    while retry_count < max_retries:
        try:
            if POLARS_AVAILABLE and retry_count == 0:
                # Try Polars first for maximum speed (only on first attempt)
                try:
                    if sheet_name:
                        df_pl = pl.read_excel(file_path, sheet_name=sheet_name)
                    else:
                        df_pl = pl.read_excel(file_path)
                    df = df_pl.to_pandas()
                    
                    # Clean column names
                    df.columns = [clean_column_name(col) for col in df.columns]
                    
                    # Validate result
                    if df.empty:
                        logger.warning(f"Polars returned empty dataframe for {file_path}")
                        raise ValueError("Empty dataframe returned")
                    
                    return df
                    
                except Exception as e:
                    logger.warning(f"Polars failed for {file_path} (sheet: {sheet_name}), trying pandas: {e}")
                    # Don't increment retry_count for Polars failure, just fall through to pandas
                    pass
            
            # Use pandas (either as fallback or if Polars not available)
            engines_to_try = []
            
            if file_path.lower().endswith('.xls'):
                engines_to_try = ['xlrd', 'openpyxl']
            else:
                engines_to_try = ['openpyxl', 'xlrd']
            
            last_error = None
            for engine in engines_to_try:
                try:
                    if sheet_name:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
                    else:
                        df = pd.read_excel(file_path, engine=engine)
                    
                    # Validate result
                    if df.empty:
                        logger.warning(f"Engine {engine} returned empty dataframe for {file_path}")
                        continue
                    
                    # Clean column names
                    df.columns = [clean_column_name(col) for col in df.columns]
                    return df
                    
                except Exception as e:
                    last_error = e
                    logger.warning(f"Engine {engine} failed for {file_path}: {e}")
                    continue
            
            # If we get here, all engines failed
            raise last_error or ValueError("All engines failed")
            
        except Exception as e:
            retry_count += 1
            if retry_count < max_retries:
                logger.warning(f"Retry {retry_count}/{max_retries} for {file_path}: {e}")
                time.sleep(0.5 * retry_count)  # Progressive backoff
            else:
                logger.error(f"Failed to read Excel file {file_path} after {max_retries} attempts: {e}")
                raise ValueError(f"Failed to read Excel file after {max_retries} attempts: {str(e)}")
    
    # Should never reach here
    raise ValueError(f"Unexpected error reading {file_path}")


def read_excel_all_sheets_optimized(file_path: str) -> Dict[str, pd.DataFrame]:
    """Read all sheets from Excel file using Polars optimization with robust error handling"""
    max_retries = 2
    retry_count = 0
    
    while retry_count < max_retries:
        try:
            if POLARS_AVAILABLE and retry_count == 0:
                try:
                    # Polars doesn't have direct all-sheets reading, so we'll get sheet names first
                    # Use pandas to get sheet names, then read each with Polars
                    excel_file = pd.ExcelFile(file_path)
                    sheet_names = excel_file.sheet_names
                    
                    if not sheet_names:
                        raise ValueError("No sheets found in Excel file")
                    
                    all_sheets = {}
                    failed_sheets = []
                    
                    for sheet_name in sheet_names:
                        try:
                            df_pl = pl.read_excel(file_path, sheet_name=sheet_name)
                            df = df_pl.to_pandas()
                            
                            # Validate sheet
                            if df.empty:
                                logger.warning(f"Sheet {sheet_name} is empty, skipping")
                                continue
                            
                            df.columns = [clean_column_name(col) for col in df.columns]
                            all_sheets[sheet_name] = df
                            
                        except Exception as e:
                            logger.warning(f"Polars failed for sheet {sheet_name}, will try pandas fallback: {e}")
                            failed_sheets.append(sheet_name)
                            continue
                    
                    # If we got some sheets with Polars, try pandas for failed sheets
                    if failed_sheets and all_sheets:
                        logger.info(f"Polars succeeded for {len(all_sheets)} sheets, trying pandas for {len(failed_sheets)} failed sheets")
                        for sheet_name in failed_sheets:
                            try:
                                engine = 'openpyxl' if not file_path.lower().endswith('.xls') else 'xlrd'
                                df = pd.read_excel(file_path, sheet_name=sheet_name, engine=engine)
                                if not df.empty:
                                    df.columns = [clean_column_name(col) for col in df.columns]
                                    all_sheets[sheet_name] = df
                            except Exception as e:
                                logger.warning(f"Pandas also failed for sheet {sheet_name}: {e}")
                                continue
                    
                    if all_sheets:
                        return all_sheets
                    else:
                        raise ValueError("No readable sheets found with Polars")
                        
                except Exception as e:
                    logger.warning(f"Polars batch reading failed for {file_path}, using pandas fallback: {e}")
                    # Fall through to pandas
                    pass
            
            # Pandas fallback with multiple engine attempts
            engines_to_try = ['openpyxl', 'xlrd'] if not file_path.lower().endswith('.xls') else ['xlrd', 'openpyxl']
            
            last_error = None
            for engine in engines_to_try:
                try:
                    all_sheets = pd.read_excel(file_path, sheet_name=None, engine=engine)
                    
                    if not all_sheets:
                        raise ValueError("No sheets returned")
                    
                    # Clean column names for all sheets and validate
                    valid_sheets = {}
                    for sheet_name, df in all_sheets.items():
                        try:
                            if not df.empty:
                                df.columns = [clean_column_name(col) for col in df.columns]
                                valid_sheets[sheet_name] = df
                        except Exception as e:
                            logger.warning(f"Failed to process sheet {sheet_name}: {e}")
                            continue
                    
                    if valid_sheets:
                        return valid_sheets
                    else:
                        raise ValueError("No valid sheets found")
                        
                except Exception as e:
                    last_error = e
                    logger.warning(f"Engine {engine} failed for all sheets in {file_path}: {e}")
                    continue
            
            # If we get here, all engines failed
            raise last_error or ValueError("All engines failed for all sheets")
            
        except Exception as e:
            retry_count += 1
            if retry_count < max_retries:
                logger.warning(f"Retry {retry_count}/{max_retries} for all sheets in {file_path}: {e}")
                time.sleep(1 * retry_count)  # Progressive backoff
            else:
                logger.error(f"Failed to read Excel file {file_path} after {max_retries} attempts: {e}")
                raise ValueError(f"Failed to read Excel file after {max_retries} attempts: {str(e)}")
    
    # Should never reach here
    raise ValueError(f"Unexpected error reading all sheets from {file_path}")


class CSVSourceWidget(QWidget):
    """Widget for configuring a single CSV source"""
    
    def __init__(self, parent=None, source_index=0, dialog=None):
        super().__init__(parent)
        self.source_index = source_index
        self.dialog = dialog  # Reference to the main dialog
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Header
        header_frame = QFrame()
        header_frame.setFrameStyle(QFrame.Shape.StyledPanel)
        header_layout = QHBoxLayout(header_frame)
        
        self.title_label = QLabel(f"CSV Source {self.source_index + 1}")
        self.title_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        header_layout.addWidget(self.title_label)
        
        header_layout.addStretch()
        
        # Mode toggle
        self.mode_combo = QComboBox()
        self.mode_combo.addItems([
            "CSV Folder (Multiple Files)", 
            "CSV Single File",
            "Excel Folder (Multiple Files)",
            "Excel Single File"
        ])
        self.mode_combo.currentTextChanged.connect(self.on_mode_changed)
        header_layout.addWidget(self.mode_combo)
        
        self.remove_btn = QPushButton("Remove")
        self.remove_btn.clicked.connect(self.request_remove)
        header_layout.addWidget(self.remove_btn)
        
        layout.addWidget(header_frame)
        
        # Configuration form
        self.form_layout = QFormLayout()
        
        # Path selection (folder or file depending on mode)
        path_layout = QHBoxLayout()
        self.path_line = QLineEdit()
        self.path_line.setPlaceholderText("Select CSV folder...")
        self.path_line.textChanged.connect(self.on_path_changed)
        
        self.browse_btn = QPushButton("Browse")
        self.browse_btn.clicked.connect(self.browse_path)
        
        path_layout.addWidget(self.path_line)
        path_layout.addWidget(self.browse_btn)
        
        # Store label references for dynamic updating
        self.path_label = QLabel("CSV Folder:")
        self.form_layout.addRow(self.path_label, path_layout)
        
        # Table name
        self.table_line = QLineEdit()
        self.table_line.setPlaceholderText("Enter table name")
        self.form_layout.addRow("Table Name:", self.table_line)
        
        # File pattern (only for folder mode)
        self.pattern_line = QLineEdit("*.csv")
        self.pattern_label = QLabel("File Pattern:")
        self.form_layout.addRow(self.pattern_label, self.pattern_line)
        
        # Excel sheet selection (only for Excel modes)
        self.sheet_combo = QComboBox()
        self.sheet_combo.addItems(["All sheets (combined)", "First sheet only", "Specific sheet..."])
        self.sheet_combo.currentTextChanged.connect(self.on_sheet_selection_changed)
        self.sheet_label = QLabel("Sheet Selection:")
        self.form_layout.addRow(self.sheet_label, self.sheet_combo)
        
        # Specific sheet name input
        self.sheet_name_line = QLineEdit()
        self.sheet_name_line.setPlaceholderText("Enter sheet name...")
        self.sheet_name_label = QLabel("Sheet Name:")
        self.form_layout.addRow(self.sheet_name_label, self.sheet_name_line)
        
        layout.addLayout(self.form_layout)
        
        # File preview
        self.file_list = QListWidget()
        self.file_list.setMaximumHeight(80)
        self.preview_title = QLabel("Files Preview:")
        layout.addWidget(self.preview_title)
        layout.addWidget(self.file_list)
        
        self.preview_label = QLabel("No folder selected")
        self.preview_label.setStyleSheet("color: gray;")
        layout.addWidget(self.preview_label)
        
        # Set initial mode
        self.current_mode = "csv_folder"
        self.file_type = "csv"
        self.update_ui_for_mode()
    
    def on_sheet_selection_changed(self):
        """Handle sheet selection changes"""
        selection = self.sheet_combo.currentText()
        is_specific = "Specific sheet" in selection
        self.sheet_name_line.setVisible(is_specific)
        self.sheet_name_label.setVisible(is_specific)
    
    def on_mode_changed(self):
        """Handle mode change between different file types and modes"""
        current_text = self.mode_combo.currentText()
        
        if "Excel Single File" in current_text:
            self.current_mode = "excel_file"
            self.file_type = "excel"
        elif "Excel Folder" in current_text:
            self.current_mode = "excel_folder" 
            self.file_type = "excel"
        elif "CSV Single File" in current_text:
            self.current_mode = "csv_file"
            self.file_type = "csv"
        else:  # CSV Folder (default)
            self.current_mode = "csv_folder"
            self.file_type = "csv"
            
        self.update_ui_for_mode()
        
    def update_ui_for_mode(self):
        """Update UI elements based on current mode"""
        if self.current_mode == "csv_file":
            # CSV Single file mode
            self.path_label.setText("CSV File:")
            self.path_line.setPlaceholderText("Select CSV file...")
            self.pattern_line.setVisible(False)
            self.pattern_label.setVisible(False)
            self.sheet_combo.setVisible(False)
            self.sheet_label.setVisible(False)
            self.sheet_name_line.setVisible(False)
            self.sheet_name_label.setVisible(False)
            self.preview_title.setText("File Preview:")
        elif self.current_mode == "excel_file":
            # Excel Single file mode
            self.path_label.setText("Excel File:")
            self.path_line.setPlaceholderText("Select Excel file...")
            self.pattern_line.setVisible(False)
            self.pattern_label.setVisible(False)
            self.sheet_combo.setVisible(True)
            self.sheet_label.setVisible(True)
            self.on_sheet_selection_changed()  # Update sheet name visibility
            self.preview_title.setText("File Preview:")
        elif self.current_mode == "excel_folder":
            # Excel Folder mode
            self.path_label.setText("Excel Folder:")
            self.path_line.setPlaceholderText("Select Excel folder...")
            self.pattern_line.setVisible(True)
            self.pattern_label.setVisible(True)
            self.pattern_line.setText("*.xlsx")
            self.pattern_label.setText("File Pattern:")
            self.sheet_combo.setVisible(True)
            self.sheet_label.setVisible(True)
            self.on_sheet_selection_changed()  # Update sheet name visibility
            self.preview_title.setText("Files Preview:")
        else:
            # CSV Folder mode (default)
            self.path_label.setText("CSV Folder:")
            self.path_line.setPlaceholderText("Select CSV folder...")
            self.pattern_line.setVisible(True)
            self.pattern_label.setVisible(True)
            self.pattern_line.setText("*.csv")
            self.pattern_label.setText("File Pattern:")
            self.sheet_combo.setVisible(False)
            self.sheet_label.setVisible(False)
            self.sheet_name_line.setVisible(False)
            self.sheet_name_label.setVisible(False)
            self.preview_title.setText("Files Preview:")
        
        # Update preview without clearing if we're loading configuration
        self.update_preview()
    
    def browse_path(self):
        """Browse for folder or file depending on mode"""
        if self.current_mode == "csv_file":
            file_path, _ = QFileDialog.getOpenFileName(
                self, 
                "Select CSV File", 
                "", 
                "CSV Files (*.csv);;All Files (*)"
            )
            if file_path:
                self.path_line.setText(file_path)
        elif self.current_mode == "excel_file":
            file_path, _ = QFileDialog.getOpenFileName(
                self, 
                "Select Excel File", 
                "", 
                "Excel Files (*.xlsx *.xls);;All Files (*)"
            )
            if file_path:
                self.path_line.setText(file_path)
        else:
            # Folder modes
            if self.file_type == "excel":
                folder = QFileDialog.getExistingDirectory(self, "Select Excel Folder")
            else:
                folder = QFileDialog.getExistingDirectory(self, "Select CSV Folder")
            if folder:
                self.path_line.setText(folder)
    
    def on_path_changed(self):
        """Handle path change for both folder and file modes"""
        path = self.path_line.text()
        if path and os.path.exists(path):
            self.update_preview()
            # Auto-suggest table name
            if not self.table_line.text():
                if self.current_mode in ["csv_file", "excel_file"]:
                    # Use filename without extension
                    suggested_name = os.path.splitext(os.path.basename(path))[0]
                else:
                    # Use folder name
                    suggested_name = os.path.basename(path.rstrip('/\\'))
                self.table_line.setText(self.clean_table_name(suggested_name))
    
    def update_preview(self):
        """Update preview for both folder and file modes"""
        self.file_list.clear()
        path = self.path_line.text()
        
        if not path or not os.path.exists(path):
            self.preview_label.setText("Invalid path")
            return
        
        if self.current_mode in ["csv_file", "excel_file"]:
            # Single file mode
            filename = os.path.basename(path)
            
            if self.current_mode == "csv_file" and path.lower().endswith('.csv'):
                self.file_list.addItem(filename)
                
                # Try to get CSV file info
                try:
                    file_size = os.path.getsize(path)
                    size_mb = file_size / (1024 * 1024)
                    
                    # Try to read first few rows to get column info
                    df_sample = pd.read_csv(path, nrows=0)  # Just headers
                    col_count = len(df_sample.columns)
                    
                    self.preview_label.setText(f"CSV: {filename} ({size_mb:.1f} MB, {col_count} columns)")
                except Exception as e:
                    self.preview_label.setText(f"CSV: {filename} (Unable to read: {str(e)})")
                    
            elif self.current_mode == "excel_file" and path.lower().endswith(('.xlsx', '.xls')):
                self.file_list.addItem(filename)
                
                # Try to get Excel file info
                try:
                    file_size = os.path.getsize(path)
                    size_mb = file_size / (1024 * 1024)
                    
                    # Get worksheet info
                    workbook = load_workbook(path, read_only=True)
                    sheet_names = workbook.sheetnames
                    workbook.close()
                    
                    # Get column info from first sheet
                    df_sample = pd.read_excel(path, nrows=0, sheet_name=sheet_names[0])
                    col_count = len(df_sample.columns)
                    
                    sheet_info = f"{len(sheet_names)} sheet(s)" if len(sheet_names) > 1 else f"Sheet: {sheet_names[0]}"
                    self.preview_label.setText(f"Excel: {filename} ({size_mb:.1f} MB, {col_count} columns, {sheet_info})")
                except Exception as e:
                    self.preview_label.setText(f"Excel: {filename} (Unable to read: {str(e)})")
            else:
                file_type = "Excel" if self.current_mode == "excel_file" else "CSV"
                self.preview_label.setText(f"Selected file is not a {file_type} file")
        else:
            # Folder mode
            if self.file_type == "excel":
                pattern = self.pattern_line.text() or "*.xlsx"
                default_pattern = "*.xlsx"
            else:
                pattern = self.pattern_line.text() or "*.csv"
                default_pattern = "*.csv"
                
            files = glob.glob(os.path.join(path, pattern))
            
            # Also include .xls files for Excel mode
            if self.file_type == "excel" and pattern == "*.xlsx":
                xls_files = glob.glob(os.path.join(path, "*.xls"))
                files.extend(xls_files)
            
            if files:
                for file_path in sorted(files)[:5]:  # Show max 5 files
                    filename = os.path.basename(file_path)
                    self.file_list.addItem(filename)
                
                if len(files) > 5:
                    self.file_list.addItem(f"... and {len(files) - 5} more files")
                
                file_type_name = "Excel" if self.file_type == "excel" else "CSV"
                self.preview_label.setText(f"Found {len(files)} {file_type_name} files")
            else:
                file_type_name = "Excel" if self.file_type == "excel" else "CSV"
                self.preview_label.setText(f"No {file_type_name} files found")
    
    def clean_table_name(self, name):
        import re
        name = str(name).lower().strip()
        name = re.sub(r'[^\w]', '_', name)
        name = re.sub(r'_+', '_', name)
        return name.strip('_') or f"table_{self.source_index + 1}"
    
    def request_remove(self):
        if self.dialog:
            self.dialog.remove_source(self.source_index)
    
    def get_config(self):
        config = {
            'table_name': self.table_line.text(),
            'mode': self.current_mode,
            'file_type': getattr(self, 'file_type', 'csv')
        }
        
        if self.current_mode in ["csv_file", "excel_file"]:
            config['file_path'] = self.path_line.text()
        else:
            config['folder_path'] = self.path_line.text()
            if self.file_type == "excel":
                config['file_pattern'] = self.pattern_line.text() or "*.xlsx"
            else:
                config['file_pattern'] = self.pattern_line.text() or "*.csv"
        
        # Add Excel sheet selection configuration
        if self.file_type == "excel":
            config['sheet_selection'] = self.sheet_combo.currentText()
            config['sheet_name'] = self.sheet_name_line.text()
        
        return config
    
    def set_config(self, config):
        """Load configuration into this source widget"""
        try:
            # Set table name
            self.table_line.setText(config.get('table_name', ''))
            
            # Handle mode-specific configuration
            mode = config.get('mode', 'csv_folder')
            file_type = config.get('file_type', 'csv')
            
            # Set up mode combo based on saved configuration
            if mode == 'csv_file':
                self.mode_combo.setCurrentText("CSV Single File")
            elif mode == 'excel_file':
                self.mode_combo.setCurrentText("Excel Single File")
            elif mode == 'excel_folder':
                self.mode_combo.setCurrentText("Excel Folder (Multiple Files)")
            else:  # csv_folder or legacy 'folder'
                self.mode_combo.setCurrentText("CSV Folder (Multiple Files)")
            
            # Update internal state
            self.current_mode = mode
            self.file_type = file_type
            self.update_ui_for_mode()
            
            # Set paths
            if mode in ['csv_file', 'excel_file']:
                self.path_line.setText(config.get('file_path', ''))
            else:
                self.path_line.setText(config.get('folder_path', ''))
                pattern = config.get('file_pattern', '*.xlsx' if file_type == 'excel' else '*.csv')
                self.pattern_line.setText(pattern)
            
            # Set Excel sheet selection if applicable
            if file_type == 'excel':
                sheet_selection = config.get('sheet_selection', 'All sheets (combined)')
                self.sheet_combo.setCurrentText(sheet_selection)
                
                sheet_name = config.get('sheet_name', '')
                self.sheet_name_line.setText(sheet_name)
                
                self.on_sheet_selection_changed()  # Update visibility
            
            # Trigger updates
            self.on_path_changed()
            
        except Exception as e:
            logger.error(f"Error setting source configuration: {e}")
    
    def is_valid(self):
        config = self.get_config()
        path_exists = False
        
        if self.current_mode in ["csv_file", "excel_file"]:
            path_exists = config.get('file_path') and os.path.exists(config['file_path'])
        else:
            path_exists = config.get('folder_path') and os.path.exists(config['folder_path'])
        
        return path_exists and config['table_name']


class CSVAutomationWorker(QThread):
    """Worker thread for CSV automation processing with enhanced crash prevention"""
    
    progress = pyqtSignal(int, str)
    finished = pyqtSignal(bool, str, dict)
    error = pyqtSignal(str)
    
    def __init__(self, connection, sources_config, output_config, sql_query=None, connection_info=None):
        super().__init__()
        
        # Store the connection info passed from main app
        self.connection_info = connection_info
        
        # Create a separate connection for automation using the same database file as main app
        try:
            if connection_info and connection_info.get('type') and connection_info.get('path'):
                db_type = connection_info['type'].lower()
                db_path = connection_info.get('file_path') or connection_info.get('path')
                
                if db_path and db_path != ':memory:':
                    if db_type == 'duckdb':
                        self.connection = duckdb.connect(db_path)
                        logger.info(f"Created separate DuckDB connection for automation: {db_path}")
                    elif db_type in ['sqlite', 'sqlite3']:
                        import sqlite3
                        self.connection = sqlite3.connect(db_path)
                        logger.info(f"Created separate SQLite connection for automation: {db_path}")
                    else:
                        # Fallback to using the provided connection
                        self.connection = connection
                        logger.info(f"Using provided connection for unsupported type: {db_type}")
                else:
                    # In-memory database, use provided connection
                    self.connection = connection
                    logger.info("Using provided connection for in-memory database")
            else:
                # No connection info provided, use provided connection
                self.connection = connection
                logger.info("Using provided connection (no connection info available)")
                
        except Exception as e:
            logger.warning(f"Failed to create separate connection, using provided: {e}")
            self.connection = connection
        
        self.sources_config = sources_config
        self.output_config = output_config
        self.sql_query = sql_query
        self.cancel_requested = False
        self.current_progress = 0
        
        # Validate connection on initialization
        if not self.validate_connection():
            logger.warning("Initial connection validation failed, will attempt reconnection when needed")
    
    def cancel(self):
        self.cancel_requested = True
    
    def validate_connection(self):
        """Validate and repair DuckDB connection if needed"""
        try:
            # Test the connection with a simple query
            result = self.connection.execute("SELECT 1").fetchone()
            return result is not None
        except Exception as e:
            logger.warning(f"Connection validation failed: {e}")
            return False
    
    def reconnect_if_needed(self):
        """Reconnect to database if connection is lost"""
        if not self.validate_connection():
            try:
                logger.info("Attempting to reconnect to database...")
                
                # Try to close the current connection
                if hasattr(self.connection, 'close'):
                    try:
                        self.connection.close()
                    except:
                        pass
                
                # Create new connection using the connection info from main app
                try:
                    if self.connection_info and self.connection_info.get('type'):
                        db_type = self.connection_info['type'].lower()
                        db_path = self.connection_info.get('file_path') or self.connection_info.get('path')
                        
                        if db_path and db_path != ':memory:':
                            if db_type == 'duckdb':
                                self.connection = duckdb.connect(db_path)
                                logger.info(f"Successfully reconnected to DuckDB: {db_path}")
                            elif db_type in ['sqlite', 'sqlite3']:
                                import sqlite3
                                self.connection = sqlite3.connect(db_path)
                                logger.info(f"Successfully reconnected to SQLite: {db_path}")
                            else:
                                logger.warning(f"Unsupported database type for reconnection: {db_type}")
                                return False
                        else:
                            logger.warning("No valid database path for reconnection")
                            return False
                    else:
                        # Fallback to main.duckdb (assuming this is the main app pattern)
                        self.connection = duckdb.connect("main.duckdb")
                        logger.info("Successfully reconnected to main.duckdb (fallback)")
                        
                except Exception as e:
                    logger.warning(f"Failed to reconnect to original database: {e}")
                    # Last resort: main.duckdb fallback
                    try:
                        self.connection = duckdb.connect("main.duckdb")
                        logger.info("Successfully reconnected using main.duckdb fallback")
                    except Exception as e2:
                        logger.error(f"All reconnection attempts failed: {e2}")
                        return False
                
                # Validate the new connection
                if self.validate_connection():
                    logger.info("Database reconnection successful")
                    return True
                else:
                    logger.error("Reconnected but validation failed")
                    return False
                    
            except Exception as e:
                logger.error(f"Failed to reconnect to database: {e}")
                return False
        return True
    
    def get_file_size_mb(self, file_path):
        """Get file size in MB"""
        try:
            return os.path.getsize(file_path) / (1024 * 1024)
        except:
            return 0
    
    def check_memory_usage(self):
        """Check current memory usage and trigger garbage collection if needed"""
        try:
            # Get process memory info
            process = psutil.Process()
            memory_info = process.memory_info()
            memory_usage_mb = memory_info.rss / (1024 * 1024)
            
            # Get system memory info
            virtual_memory = psutil.virtual_memory()
            available_memory_mb = virtual_memory.available / (1024 * 1024)
            
            # If memory usage is high or available memory is low, trigger garbage collection
            if memory_usage_mb > 2000 or available_memory_mb < 500:  # 2GB process or <500MB available
                gc.collect()
                
                # Re-check after garbage collection
                memory_info = process.memory_info()
                memory_usage_mb = memory_info.rss / (1024 * 1024)
                
                self.progress.emit(
                    self.current_progress,
                    f"Memory management: {memory_usage_mb:.0f}MB used, {available_memory_mb:.0f}MB available"
                )
            
            # Return True if we have enough memory to continue
            return available_memory_mb > 200  # At least 200MB available
            
        except Exception:
            # If we can't check memory, assume it's OK
            return True
    
    def force_cleanup(self):
        """Force cleanup of temporary variables and garbage collection"""
        gc.collect()
        import threading
        if hasattr(threading, 'active_count'):
            # Log thread count for debugging
            logger.debug(f"Active threads: {threading.active_count()}")
    
    def cleanup_temp_files(self):
        """Clean up all temporary files created during processing"""
        try:
            import glob
            # Clean up temp files with various patterns
            temp_patterns = [
                "temp_*.csv",
                "temp_chunk_*.csv"
            ]
            
            for pattern in temp_patterns:
                for temp_file in glob.glob(pattern):
                    try:
                        os.remove(temp_file)
                        logger.debug(f"Cleaned up temporary file: {temp_file}")
                    except Exception as e:
                        logger.warning(f"Could not remove temp file {temp_file}: {e}")
        except Exception as e:
            logger.warning(f"Error during temp file cleanup: {e}")
    
    def normalize_column_names(self, dataframes_list):
        """Normalize column names across multiple dataframes"""
        if not dataframes_list:
            return []
        
        # Collect all unique column names from all dataframes
        all_columns = set()
        for df in dataframes_list:
            all_columns.update(df.columns)
        
        # Sort columns for consistent ordering
        normalized_columns = sorted(all_columns)
        
        # Reindex all dataframes to have the same columns
        normalized_dfs = []
        for df in dataframes_list:
            df_normalized = df.reindex(columns=normalized_columns, fill_value=None)
            normalized_dfs.append(df_normalized)
        
        return normalized_dfs
    
    def read_excel_file(self, file_path, source_file_name, sheet_selection="All sheets (combined)", sheet_name=""):
        """Read an Excel file and return a DataFrame - Crash-resistant with Polars optimization"""
        max_attempts = 3
        attempt = 0
        
        while attempt < max_attempts:
            try:
                # Validate file exists and is readable
                if not os.path.exists(file_path):
                    raise FileNotFoundError(f"Excel file not found: {file_path}")
                
                if os.path.getsize(file_path) == 0:
                    raise ValueError(f"Excel file is empty: {file_path}")
                
                # Get sheet names with error handling
                try:
                    excel_file = pd.ExcelFile(file_path)
                    sheet_names = excel_file.sheet_names
                    excel_file.close()  # Explicitly close to free resources
                except Exception as e:
                    logger.warning(f"Could not read sheet names from {file_path}: {e}")
                    if attempt < max_attempts - 1:
                        attempt += 1
                        time.sleep(0.5)
                        continue
                    else:
                        raise
                
                if not sheet_names:
                    raise ValueError(f"No sheets found in Excel file: {file_path}")
                
                dataframes = []
                successful_sheets = 0
                failed_sheets = []
                
                if sheet_selection == "First sheet only":
                    # Read only the first sheet using crash-resistant method
                    if sheet_names:
                        try:
                            df = read_excel_optimized(file_path, sheet_names[0])
                            df['_source_file'] = source_file_name
                            df['_source_sheet'] = sheet_names[0]
                            dataframes.append(df)
                            successful_sheets += 1
                        except Exception as e:
                            logger.warning(f"Could not read first sheet '{sheet_names[0]}' from {file_path}: {e}")
                            failed_sheets.append(sheet_names[0])
                            
                elif "Specific sheet" in sheet_selection and sheet_name:
                    # Read specific sheet by name with fallback
                    if sheet_name in sheet_names:
                        try:
                            df = read_excel_optimized(file_path, sheet_name)
                            df['_source_file'] = source_file_name
                            df['_source_sheet'] = sheet_name
                            dataframes.append(df)
                            successful_sheets += 1
                        except Exception as e:
                            logger.warning(f"Could not read sheet '{sheet_name}' from {file_path}: {e}")
                            failed_sheets.append(sheet_name)
                    else:
                        logger.warning(f"Sheet '{sheet_name}' not found in {file_path}. Available sheets: {sheet_names}")
                        # Fall back to first sheet with error isolation
                        if sheet_names:
                            try:
                                df = read_excel_optimized(file_path, sheet_names[0])
                                df['_source_file'] = source_file_name
                                df['_source_sheet'] = sheet_names[0]
                                dataframes.append(df)
                                successful_sheets += 1
                                logger.info(f"Successfully used fallback sheet '{sheet_names[0]}' from {file_path}")
                            except Exception as e:
                                logger.warning(f"Could not read fallback sheet '{sheet_names[0]}' from {file_path}: {e}")
                                failed_sheets.append(sheet_names[0])
                else:
                    # Read all sheets with individual error isolation
                    try:
                        # Try batch reading first
                        all_sheets = read_excel_all_sheets_optimized(file_path)
                        
                        for sheet_name_iter, df in all_sheets.items():
                            try:
                                # Validate sheet data
                                if df.empty:
                                    logger.warning(f"Sheet '{sheet_name_iter}' is empty in {file_path}, skipping")
                                    continue
                                
                                # Add sheet info to track source
                                df['_source_file'] = source_file_name
                                df['_source_sheet'] = sheet_name_iter
                                dataframes.append(df)
                                successful_sheets += 1
                            except Exception as e:
                                logger.warning(f"Could not process sheet '{sheet_name_iter}' from {file_path}: {e}")
                                failed_sheets.append(sheet_name_iter)
                                continue
                                
                    except Exception as e:
                        # Fallback to individual sheet reading with error isolation
                        logger.warning(f"Batch reading failed, trying individual sheets: {e}")
                        for sheet_name_iter in sheet_names:
                            try:
                                df = read_excel_optimized(file_path, sheet_name_iter)
                                
                                # Validate sheet data
                                if df.empty:
                                    logger.warning(f"Sheet '{sheet_name_iter}' is empty in {file_path}, skipping")
                                    continue
                                
                                # Add sheet info to track source
                                df['_source_file'] = source_file_name
                                df['_source_sheet'] = sheet_name_iter
                                
                                dataframes.append(df)
                                successful_sheets += 1
                                
                            except Exception as e2:
                                logger.warning(f"Could not read sheet '{sheet_name_iter}' from {file_path}: {e2}")
                                failed_sheets.append(sheet_name_iter)
                                continue
                
                # Report results
                if failed_sheets:
                    logger.warning(f"Failed to read {len(failed_sheets)} sheet(s) from {file_path}: {failed_sheets}")
                
                if not dataframes:
                    if attempt < max_attempts - 1:
                        logger.warning(f"No readable sheets found in {file_path}, retrying... (attempt {attempt + 1}/{max_attempts})")
                        attempt += 1
                        time.sleep(1)
                        continue
                    else:
                        raise ValueError(f"No readable sheets found in {file_path} after {max_attempts} attempts. Failed sheets: {failed_sheets}")
                
                logger.info(f"Successfully read {successful_sheets} sheet(s) from {file_path}")
                
                # Handle single vs multiple dataframes with crash protection
                try:
                    if len(dataframes) == 1:
                        return dataframes[0]
                    
                    # Combine multiple sheets with error handling
                    if len(dataframes) > 1:
                        try:
                            # Check if column normalization is needed
                            first_columns = set(dataframes[0].columns)
                            needs_normalization = any(set(df.columns) != first_columns for df in dataframes[1:])
                            
                            if needs_normalization:
                                logger.info(f"Normalizing columns for {len(dataframes)} sheets from {file_path}")
                                normalized_dfs = self.normalize_column_names(dataframes)
                                combined_df = pd.concat(normalized_dfs, ignore_index=True, sort=False)
                            else:
                                # Faster concatenation when columns are already aligned
                                combined_df = pd.concat(dataframes, ignore_index=True)
                            
                            return combined_df
                            
                        except Exception as e:
                            logger.error(f"Failed to combine sheets from {file_path}: {e}")
                            # Return the first dataframe as fallback
                            logger.info(f"Falling back to first sheet only from {file_path}")
                            return dataframes[0]
                
                except Exception as e:
                    logger.error(f"Error processing dataframes from {file_path}: {e}")
                    if attempt < max_attempts - 1:
                        attempt += 1
                        time.sleep(1)
                        continue
                    else:
                        raise
                
                # Should not reach here
                return dataframes[0] if dataframes else None
                
            except Exception as e:
                attempt += 1
                if attempt < max_attempts:
                    logger.warning(f"Excel reading attempt {attempt}/{max_attempts} failed for {file_path}: {e}")
                    time.sleep(1 * attempt)  # Progressive backoff
                else:
                    logger.error(f"All {max_attempts} attempts failed for Excel file {file_path}: {e}")
                    raise ValueError(f"Failed to read Excel file after {max_attempts} attempts: {str(e)}")
        
        # Should never reach here
        raise ValueError(f"Unexpected error reading Excel file {file_path}")
    
    def process_large_excel_chunked(self, file_path, output_file, source_file_name, chunk_size=10000, sheet_selection="All sheets (combined)", sheet_name="", unified_columns=None):
        """Process large Excel files in chunks using Polars optimization"""
        try:
            total_rows = 0
            first_chunk = True
            
            # Get sheet names for processing (using pandas for discovery)
            excel_file = pd.ExcelFile(file_path)
            
            # Determine which sheets to process based on selection
            sheets_to_process = []
            if sheet_selection == "First sheet only":
                if excel_file.sheet_names:
                    sheets_to_process = [excel_file.sheet_names[0]]
            elif "Specific sheet" in sheet_selection and sheet_name:
                if sheet_name in excel_file.sheet_names:
                    sheets_to_process = [sheet_name]
                else:
                    logger.warning(f"Sheet '{sheet_name}' not found. Using first sheet as fallback.")
                    if excel_file.sheet_names:
                        sheets_to_process = [excel_file.sheet_names[0]]
            else:
                # All sheets (default)
                sheets_to_process = excel_file.sheet_names
            
            for sheet_name in sheets_to_process:
                if self.cancel_requested:
                    return None
                
                try:
                    # Read sheet using Polars optimization for maximum speed
                    df = read_excel_optimized(file_path, sheet_name)
                    df['_source_file'] = source_file_name
                    df['_source_sheet'] = sheet_name
                    
                    # Normalize to unified schema if provided
                    if unified_columns is not None:
                        df = self.normalize_dataframe_to_schema(df, unified_columns)
                    
                    # Process in chunks if the sheet is large
                    for i in range(0, len(df), chunk_size):
                        chunk = df.iloc[i:i+chunk_size].copy()
                        
                        # Write chunk
                        mode = 'w' if first_chunk else 'a'
                        header = first_chunk
                        chunk.to_csv(output_file, mode=mode, header=header, index=False, encoding='utf-8')
                        
                        total_rows += len(chunk)
                        first_chunk = False
                        
                        # Update progress
                        self.progress.emit(
                            self.current_progress + 15,
                            f"Processed {total_rows:,} rows from {source_file_name} (sheet: {sheet_name})..."
                        )
                        
                        if self.cancel_requested:
                            return None
                    
                except Exception as e:
                    logger.warning(f"Error processing sheet '{sheet_name}' in {file_path}: {e}")
                    continue
            
            return total_rows
            
        except Exception as e:
            logger.error(f"Error processing large Excel file {file_path}: {e}")
            raise ValueError(f"Failed to process large Excel file: {str(e)}")
    
    def process_large_csv_chunked(self, file_path, output_file, source_file_name, chunk_size=500000, unified_columns=None):
        """Process large CSV files in chunks to prevent memory issues"""
        try:
            total_rows = 0
            first_chunk = True
            
            # Read in chunks with flexible type handling
            try:
                # First try with normal pandas reading
                chunk_reader = pd.read_csv(file_path, encoding='utf-8', chunksize=chunk_size)
            except Exception as e:
                logger.warning(f"Standard CSV reading failed for {file_path}, trying with dtype=str: {e}")
                # Fallback to reading all columns as strings
                chunk_reader = pd.read_csv(file_path, encoding='utf-8', chunksize=chunk_size, dtype=str)
            
            for chunk_num, chunk in enumerate(chunk_reader):
                if self.cancel_requested:
                    return None
                
                # Add source file column
                chunk['_source_file'] = source_file_name
                
                # Normalize to unified schema if provided
                if unified_columns is not None:
                    chunk = self.normalize_dataframe_to_schema(chunk, unified_columns)
                
                # Write chunk (append after first chunk)
                mode = 'w' if first_chunk else 'a'
                header = first_chunk
                
                try:
                    chunk.to_csv(output_file, mode=mode, header=header, index=False, encoding='utf-8')
                except Exception as e:
                    logger.warning(f"Error writing chunk {chunk_num} from {file_path}: {e}")
                    # Try writing with string conversion
                    chunk_str = chunk.astype(str)
                    chunk_str.to_csv(output_file, mode=mode, header=header, index=False, encoding='utf-8')
                
                total_rows += len(chunk)
                first_chunk = False
                
                # Update progress
                self.progress.emit(
                    self.current_progress + 15,
                    f"Processed {total_rows:,} rows from {source_file_name}..."
                )
            
            return total_rows
            
        except Exception as e:
            logger.error(f"Error processing large file {file_path}: {str(e)}")
            raise ValueError(f"Failed to process large CSV file: {str(e)}")
    
    def process_csv_source_with_progress(self, source_config, output_file, source_name):
        """Process CSV source (folder or single file) with progress reporting and crash prevention"""
        
        # Memory management settings
        LARGE_FILE_THRESHOLD_MB = 100  # Files larger than 100MB are processed in chunks
        MAX_MEMORY_USAGE_MB = 1000     # Maximum memory usage before switching to chunked processing
        CHUNK_SIZE = 500000            # Rows per chunk for large files (increased for better performance)
        
        try:
            file_type = source_config.get('file_type', 'csv')
            mode = source_config.get('mode', 'csv_folder')
            
            if mode in ['csv_file', 'excel_file']:
                # Single file processing with memory management
                file_path = source_config.get('file_path')
                if not file_path or not os.path.exists(file_path):
                    raise FileNotFoundError(f"CSV file not found: '{file_path}'")
                
                file_size_mb = self.get_file_size_mb(file_path)
                file_name = os.path.basename(file_path)
                
                self.progress.emit(
                    self.current_progress,
                    f"Processing file: {file_name} ({file_size_mb:.1f} MB)"
                )
                
                # Use chunked processing for large files
                if file_size_mb > LARGE_FILE_THRESHOLD_MB:
                    self.progress.emit(
                        self.current_progress + 5,
                        f"Large file detected - using memory-safe processing..."
                    )
                    
                    if file_type == 'excel':
                        total_rows = self.process_large_excel_chunked(file_path, output_file, file_name, CHUNK_SIZE, unified_columns=None)
                    else:
                        total_rows = self.process_large_csv_chunked(file_path, output_file, file_name, CHUNK_SIZE, unified_columns=None)
                    
                    self.progress.emit(
                        self.current_progress + 25,
                        f"Completed {total_rows:,} rows from {source_name}"
                    )
                    
                    # Return a minimal DataFrame with row count info for results
                    return pd.DataFrame({'_row_count': [total_rows], '_source_file': [file_name]})
                
                else:
                    # Normal processing for smaller files
                    self.progress.emit(
                        self.current_progress + 10,
                        f"Reading {file_name}..."
                    )
                    
                    try:
                        # Handle different file types
                        if file_type == 'excel':
                            # Excel file processing
                            sheet_selection = source_config.get('sheet_selection', 'All sheets (combined)')
                            sheet_name = source_config.get('sheet_name', '')
                            df = self.read_excel_file(file_path, file_name, sheet_selection, sheet_name)
                        else:
                            # CSV file processing
                            df = pd.read_csv(file_path, encoding='utf-8')
                            df['_source_file'] = file_name
                        
                        self.progress.emit(
                            self.current_progress + 20,
                            f"Processing {len(df):,} rows from {source_name}..."
                        )
                        
                        # Save to output file
                        df.to_csv(output_file, index=False, encoding='utf-8')
                        
                        return df
                        
                    except (pd.errors.DtypeWarning, pd.errors.ParserError, ValueError) as e:
                        # Fallback for CSV files with type issues
                        if file_type == 'csv':
                            logger.warning(f"Type inference failed for {file_path}, reading as strings: {e}")
                            try:
                                df = pd.read_csv(file_path, encoding='utf-8', dtype=str)
                                df['_source_file'] = file_name
                                
                                self.progress.emit(
                                    self.current_progress + 20,
                                    f"Processing {len(df):,} rows from {source_name} (as text)..."
                                )
                                
                                df.to_csv(output_file, index=False, encoding='utf-8')
                                return df
                                
                            except Exception as e2:
                                logger.warning(f"String reading also failed for {file_path}, switching to chunked: {e2}")
                                total_rows = self.process_large_csv_chunked(file_path, output_file, file_name, CHUNK_SIZE, unified_columns=None)
                                return pd.DataFrame({'_row_count': [total_rows], '_source_file': [file_name]})
                        else:
                            # For Excel files, switch to chunked processing
                            logger.warning(f"Excel processing failed for {file_path}, switching to chunked: {e}")
                            sheet_selection = source_config.get('sheet_selection', 'All sheets (combined)')
                            sheet_name = source_config.get('sheet_name', '')
                            total_rows = self.process_large_excel_chunked(file_path, output_file, file_name, CHUNK_SIZE, sheet_selection, sheet_name, unified_columns=None)
                            return pd.DataFrame({'_row_count': [total_rows], '_source_file': [file_name]})
                        
                    except MemoryError:
                        # Fallback to chunked processing if we run out of memory
                        logger.warning(f"Memory error reading {file_path}, switching to chunked processing")
                        if file_type == 'excel':
                            sheet_selection = source_config.get('sheet_selection', 'All sheets (combined)')
                            sheet_name = source_config.get('sheet_name', '')
                            total_rows = self.process_large_excel_chunked(file_path, output_file, file_name, CHUNK_SIZE, sheet_selection, sheet_name, unified_columns=None)
                        else:
                            total_rows = self.process_large_csv_chunked(file_path, output_file, file_name, CHUNK_SIZE, unified_columns=None)
                        return pd.DataFrame({'_row_count': [total_rows], '_source_file': [file_name]})
            
            else:
                # Folder processing with memory management
                input_folder = source_config.get('folder_path')
                
                if file_type == 'excel':
                    file_pattern = source_config.get('file_pattern', '*.xlsx')
                    # Get both .xlsx and .xls files
                    files = glob.glob(os.path.join(input_folder, file_pattern))
                    if file_pattern == '*.xlsx':
                        files.extend(glob.glob(os.path.join(input_folder, '*.xls')))
                else:
                    file_pattern = source_config.get('file_pattern', '*.csv')
                    files = glob.glob(os.path.join(input_folder, file_pattern))
                
                if not files:
                    file_type_name = "Excel" if file_type == 'excel' else "CSV"
                    raise FileNotFoundError(f"No {file_type_name} files found in '{input_folder}' matching pattern '{file_pattern}'")
                
                # Calculate total size
                total_size_mb = sum(self.get_file_size_mb(f) for f in files)
                total_files = len(files)
                
                self.progress.emit(
                    self.current_progress,
                    f"Found {total_files} files ({total_size_mb:.1f} MB) in {source_name}"
                )
                
                # First pass: Scan all files to determine unified column schema
                self.progress.emit(
                    self.current_progress + 2,
                    f"Scanning {total_files} files to determine column schema..."
                )
                
                unified_columns = self.create_unified_column_schema(files, file_type, source_config)
                logger.info(f"Unified schema has {len(unified_columns)} columns: {unified_columns}")
                
                # Use chunked processing if total size is large
                use_chunked = total_size_mb > MAX_MEMORY_USAGE_MB
                
                if use_chunked:
                    self.progress.emit(
                        self.current_progress + 5,
                        f"Large dataset detected - using memory-safe processing with unified schema..."
                    )
                
                total_rows = 0
                first_file = True
                
                for i, file_path in enumerate(files):
                    if self.cancel_requested:
                        return None
                    
                    file_size_mb = self.get_file_size_mb(file_path)
                    file_name = os.path.basename(file_path)
                    
                    # Update progress for each file
                    file_progress = int((i / total_files) * 25)
                    self.progress.emit(
                        self.current_progress + file_progress,
                        f"Processing file {i+1}/{total_files}: {file_name} ({file_size_mb:.1f} MB)"
                    )
                    
                    try:
                        if use_chunked or file_size_mb > LARGE_FILE_THRESHOLD_MB:
                            # Process large files in chunks, append to output
                            import uuid
                            temp_output = f"temp_chunk_{i}_{uuid.uuid4().hex[:8]}.csv"
                            
                            if file_size_mb > LARGE_FILE_THRESHOLD_MB:
                                if file_type == 'excel':
                                    sheet_selection = source_config.get('sheet_selection', 'All sheets (combined)')
                                    sheet_name = source_config.get('sheet_name', '')
                                    rows_processed = self.process_large_excel_chunked(file_path, temp_output, file_name, CHUNK_SIZE, sheet_selection, sheet_name, unified_columns)
                                else:
                                    rows_processed = self.process_large_csv_chunked(file_path, temp_output, file_name, CHUNK_SIZE, unified_columns)
                            else:
                                # Normal read but write to temp file
                                if file_type == 'excel':
                                    sheet_selection = source_config.get('sheet_selection', 'All sheets (combined)')
                                    sheet_name = source_config.get('sheet_name', '')
                                    df = self.read_excel_file(file_path, file_name, sheet_selection, sheet_name)
                                else:
                                    df = pd.read_csv(file_path, encoding='utf-8')
                                    df['_source_file'] = file_name
                                
                                # Normalize to unified schema
                                df = self.normalize_dataframe_to_schema(df, unified_columns)
                                df.to_csv(temp_output, index=False, encoding='utf-8')
                                rows_processed = len(df)
                                del df  # Free memory immediately
                            
                            # Append temp file to main output
                            if first_file:
                                # First file - copy directly, remove destination if it exists
                                if os.path.exists(temp_output):
                                    try:
                                        # Remove destination file if it exists (Windows fix)
                                        if os.path.exists(output_file):
                                            os.remove(output_file)
                                        os.rename(temp_output, output_file)
                                    except OSError as e:
                                        # Fallback: copy content instead of rename
                                        logger.warning(f"Rename failed, copying file instead: {e}")
                                        with open(temp_output, 'r', encoding='utf-8') as src:
                                            with open(output_file, 'w', encoding='utf-8') as dst:
                                                dst.write(src.read())
                                        os.remove(temp_output)
                                first_file = False
                            else:
                                # Append to existing file
                                if os.path.exists(temp_output):
                                    try:
                                        with open(temp_output, 'r', encoding='utf-8') as temp_f:
                                            next(temp_f)  # Skip header
                                            with open(output_file, 'a', encoding='utf-8') as output_f:
                                                output_f.writelines(temp_f)
                                        os.remove(temp_output)
                                    except Exception as e:
                                        logger.error(f"Error appending temp file {temp_output}: {e}")
                                        # Clean up temp file even if append fails
                                        if os.path.exists(temp_output):
                                            os.remove(temp_output)
                            
                            total_rows += rows_processed
                            
                        else:
                            # Normal processing for small files
                            try:
                                if file_type == 'excel':
                                    sheet_selection = source_config.get('sheet_selection', 'All sheets (combined)')
                                    sheet_name = source_config.get('sheet_name', '')
                                    df = self.read_excel_file(file_path, file_name, sheet_selection, sheet_name)
                                else:
                                    df = pd.read_csv(file_path, encoding='utf-8')
                                    df['_source_file'] = file_name
                                
                                # Normalize to unified schema (works for both CSV and Excel)
                                df = self.normalize_dataframe_to_schema(df, unified_columns)
                                
                                # Write to output file
                                mode = 'w' if first_file else 'a'
                                header = first_file
                                df.to_csv(output_file, mode=mode, header=header, index=False, encoding='utf-8')
                                
                                total_rows += len(df)
                                first_file = False
                                del df  # Free memory immediately
                                
                            except (pd.errors.DtypeWarning, pd.errors.ParserError, ValueError) as e:
                                # Fallback to string reading for problematic CSV files
                                if file_type == 'csv':
                                    logger.warning(f"Type issues with {file_path}, reading as strings: {e}")
                                    df = pd.read_csv(file_path, encoding='utf-8', dtype=str)
                                    df['_source_file'] = file_name
                                    
                                    # Normalize to unified schema
                                    df = self.normalize_dataframe_to_schema(df, unified_columns)
                                    
                                    mode = 'w' if first_file else 'a'
                                    header = first_file
                                    df.to_csv(output_file, mode=mode, header=header, index=False, encoding='utf-8')
                                    
                                    total_rows += len(df)
                                    first_file = False
                                    del df  # Free memory immediately
                                else:
                                    # For Excel files, skip problematic files
                                    logger.warning(f"Skipping problematic Excel file {file_path}: {e}")
                                    continue
                        
                    except Exception as e:
                        logger.error(f"Error reading {file_path}: {str(e)}")
                        # Continue with other files instead of failing completely
                        continue
                
                if total_rows == 0:
                    raise ValueError("No CSV files could be successfully processed")
                
                self.progress.emit(
                    self.current_progress + 28,
                    f"Completed merging {total_files} files - {total_rows:,} total rows"
                )
                
                # Return minimal DataFrame with summary info  
                return pd.DataFrame({'_row_count': [total_rows], '_files_processed': [total_files]})
                
        except MemoryError as e:
            error_msg = f"Out of memory processing {source_name}. Try processing smaller batches or increase system memory."
            logger.error(error_msg)
            raise MemoryError(error_msg)
            
        except Exception as e:
            error_msg = f"Error processing {source_name}: {str(e)}"
            logger.error(error_msg)
            raise ValueError(error_msg)
    
    def run(self):
        try:
            results = {
                'sources_processed': 0,
                'total_rows': 0,
                'tables_created': [],
                'output_table': None,
                'execution_time': 0
            }
            
            start_time = time.time()
            total_sources = len(self.sources_config)
            
            # Calculate progress steps
            source_progress_step = 60 // total_sources if total_sources > 0 else 60  # 60% for all sources
            
            # Process each CSV source
            for i, source_config in enumerate(self.sources_config):
                if self.cancel_requested:
                    return
                
                # Check memory before processing each source
                if not self.check_memory_usage():
                    self.error.emit("Insufficient memory to continue processing. Please close other applications or restart the program.")
                    return
                
                self.current_progress = int(i * source_progress_step)
                
                self.progress.emit(
                    self.current_progress,
                    f"Processing source {i+1}/{total_sources}: {source_config['table_name']}"
                )
                
                # Merge CSV files from this source
                temp_output = f"temp_{source_config['table_name']}.csv"
                
                # Clean up any existing temp files from previous runs
                if os.path.exists(temp_output):
                    try:
                        os.remove(temp_output)
                    except Exception as e:
                        logger.warning(f"Could not remove existing temp file {temp_output}: {e}")
                
                # Clean up any chunk files that might be left over
                import glob
                chunk_pattern = f"temp_chunk_*_{source_config['table_name']}*.csv"
                for old_chunk in glob.glob(chunk_pattern):
                    try:
                        os.remove(old_chunk)
                    except:
                        pass
                
                try:
                    # Use our progress-enabled CSV processor (handles both folder and single file)
                    df = self.process_csv_source_with_progress(
                        source_config=source_config,
                        output_file=temp_output,
                        source_name=source_config['table_name']
                    )
                    
                    if df is None:  # Cancelled
                        return
                    
                    # Update progress for database loading
                    self.progress.emit(
                        self.current_progress + 30,
                        f"Loading {source_config['table_name']} into database..."
                    )
                    
                    # Load into DuckDB with connection validation and crash prevention
                    table_name = source_config['table_name']
                    
                    # Validate connection before database operations
                    if not self.reconnect_if_needed():
                        raise ConnectionError("Failed to establish database connection")
                    
                    # Check if table exists first for progress reporting with error handling
                    table_exists = False
                    try:
                        self.connection.execute("BEGIN TRANSACTION")  # Start transaction for safety
                        result = self.connection.execute(f"SELECT COUNT(*) FROM information_schema.tables WHERE table_name = '{table_name}'").fetchone()
                        table_exists = result[0] > 0 if result else False
                        self.connection.execute("COMMIT")  # Commit transaction
                    except Exception as e:
                        try:
                            self.connection.execute("ROLLBACK")  # Rollback on error
                        except:
                            pass
                        logger.warning(f"Could not check table existence for {table_name}: {e}")
                        table_exists = False
                    
                    if table_exists:
                        self.progress.emit(
                            self.current_progress + 25,
                            f"Replacing existing table '{table_name}'..."
                        )
                    
                    # Use absolute path for temp file to avoid path issues
                    temp_output_abs = os.path.abspath(temp_output)
                    
                    # Ensure temp file exists and is readable before trying to load it
                    if not os.path.exists(temp_output_abs):
                        raise FileNotFoundError(f"Temp file not found: {temp_output_abs}")
                    
                    if os.path.getsize(temp_output_abs) == 0:
                        raise ValueError(f"Temp file is empty: {temp_output_abs}")
                    
                    # Load with robust error handling and multiple fallback strategies
                    load_success = False
                    last_error = None
                    
                    # Strategy 1: DuckDB native CSV loading with transaction safety
                    try:
                        self.progress.emit(
                            self.current_progress + 30,
                            f"Loading {table_name} with DuckDB native reader..."
                        )
                        
                        # Start transaction for atomic operation
                        self.connection.execute("BEGIN TRANSACTION")
                        
                        # Drop table if exists (inside transaction)
                        self.connection.execute(f"DROP TABLE IF EXISTS {table_name}")
                        
                        # Load with proper path escaping and robust settings
                        escaped_path = temp_output_abs.replace('\\', '\\\\').replace("'", "''")
                        
                        self.connection.execute(f"""
                            CREATE TABLE {table_name} AS 
                            SELECT * FROM read_csv_auto('{escaped_path}', 
                                sample_size=10000,
                                ignore_errors=true,
                                null_padding=true,
                                header=true
                            )
                        """)
                        
                        # Validate table was created successfully
                        result = self.connection.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()
                        if result and result[0] >= 0:
                            self.connection.execute("COMMIT")  # Commit successful transaction
                            load_success = True
                            logger.info(f"Successfully loaded {result[0]} rows into {table_name}")
                        else:
                            raise ValueError("Table creation validation failed")
                            
                    except Exception as e1:
                        try:
                            self.connection.execute("ROLLBACK")  # Rollback failed transaction
                        except:
                            pass
                        last_error = e1
                        logger.warning(f"DuckDB native loading failed for {table_name}: {e1}")
                    
                    # Strategy 2: Force all VARCHAR columns if native loading failed
                    if not load_success:
                        try:
                            self.progress.emit(
                                self.current_progress + 32,
                                f"Loading {table_name} as text (fallback mode)..."
                            )
                            
                            if not self.reconnect_if_needed():
                                raise ConnectionError("Connection lost during fallback loading")
                            
                            self.connection.execute("BEGIN TRANSACTION")
                            self.connection.execute(f"DROP TABLE IF EXISTS {table_name}")
                            
                            self.connection.execute(f"""
                                CREATE TABLE {table_name} AS 
                                SELECT * FROM read_csv_auto('{escaped_path}', 
                                    all_varchar=true,
                                    ignore_errors=true,
                                    null_padding=true,
                                    header=true
                                )
                            """)
                            
                            # Validate table
                            result = self.connection.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()
                            if result and result[0] >= 0:
                                self.connection.execute("COMMIT")
                                load_success = True
                                logger.info(f"Successfully loaded {result[0]} rows into {table_name} (VARCHAR mode)")
                            else:
                                raise ValueError("Table creation validation failed")
                                
                        except Exception as e2:
                            try:
                                self.connection.execute("ROLLBACK")
                            except:
                                pass
                            last_error = e2
                            logger.warning(f"VARCHAR loading failed for {table_name}: {e2}")
                    
                    # Strategy 3: Basic CSV reader as last resort
                    if not load_success:
                        try:
                            self.progress.emit(
                                self.current_progress + 34,
                                f"Loading {table_name} with basic CSV reader (last resort)..."
                            )
                            
                            if not self.reconnect_if_needed():
                                raise ConnectionError("Connection lost during basic loading")
                            
                            self.connection.execute("BEGIN TRANSACTION")
                            self.connection.execute(f"DROP TABLE IF EXISTS {table_name}")
                            
                            self.connection.execute(f"""
                                CREATE TABLE {table_name} AS 
                                SELECT * FROM read_csv('{escaped_path}', 
                                    header=true,
                                    auto_detect=false,
                                    ignore_errors=true
                                )
                            """)
                            
                            # Validate table
                            result = self.connection.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()
                            if result and result[0] >= 0:
                                self.connection.execute("COMMIT")
                                load_success = True
                                logger.info(f"Successfully loaded {result[0]} rows into {table_name} (basic mode)")
                            else:
                                raise ValueError("Table creation validation failed")
                                
                        except Exception as e3:
                            try:
                                self.connection.execute("ROLLBACK")
                            except:
                                pass
                            last_error = e3
                            logger.error(f"All loading strategies failed for {table_name}: {e3}")
                    
                    if not load_success:
                        raise ValueError(f"Failed to load table {table_name} after trying all strategies. Last error: {last_error}")
                    
                    # Clean up temp file
                    if os.path.exists(temp_output_abs):
                        os.remove(temp_output_abs)
                    
                    results['sources_processed'] += 1
                    
                    # Handle both normal DataFrames and chunked processing results
                    if '_row_count' in df.columns:
                        # Chunked processing result
                        row_count = df['_row_count'].iloc[0] if len(df) > 0 else 0
                        results['total_rows'] += row_count
                    else:
                        # Normal DataFrame
                        results['total_rows'] += len(df)
                    
                    results['tables_created'].append(table_name)
                    
                    # Complete progress for this source
                    row_count_display = df['_row_count'].iloc[0] if '_row_count' in df.columns else len(df)
                    self.progress.emit(
                        int((i + 1) * source_progress_step),
                        f"Completed {source_config['table_name']} - {row_count_display:,} rows loaded"
                    )
                    
                    # Clean up memory after each source
                    del df
                    self.force_cleanup()
                    
                except MemoryError as e:
                    logger.error(f"Memory error processing source {source_config['table_name']}: {e}")
                    self.error.emit(f"Out of memory processing {source_config['table_name']}. Please try processing smaller files or free up system memory.")
                    return
                except FileNotFoundError as e:
                    logger.error(f"File not found for source {source_config['table_name']}: {e}")
                    self.error.emit(f"File not found for {source_config['table_name']}: {str(e)}")
                    return
                except PermissionError as e:
                    logger.error(f"Permission error processing source {source_config['table_name']}: {e}")
                    self.error.emit(f"Permission denied accessing files for {source_config['table_name']}. Please check file permissions.")
                    return
                except TimeoutError as e:
                    logger.error(f"Timeout error processing source {source_config['table_name']}: {e}")
                    self.error.emit(f"Operation timed out for {source_config['table_name']}. The dataset may be too large for available system resources.")
                    return
                except Exception as e:
                    logger.error(f"Error processing source {source_config['table_name']}: {e}")
                    self.error.emit(f"Error processing {source_config['table_name']}: {str(e)}")
                    return
            
            # Execute SQL query if provided
            if self.sql_query and self.output_config['table_name']:
                self.progress.emit(75, "Executing SQL transformation...")
                
                try:
                    output_table = self.output_config['table_name']
                    
                    # Check if output table exists first for progress reporting
                    output_table_exists = False
                    try:
                        result = self.connection.execute(f"SELECT 1 FROM {output_table} LIMIT 1").fetchone()
                        output_table_exists = result is not None
                    except:
                        output_table_exists = False
                    
                    if output_table_exists:
                        self.progress.emit(77, f"Replacing existing output table '{output_table}'...")
                    
                    # Always drop the output table if it exists to ensure clean replacement
                    self.connection.execute(f"DROP TABLE IF EXISTS {output_table}")
                    
                    # Create output table from SQL query
                    create_sql = f"CREATE TABLE {output_table} AS ({self.sql_query})"
                    self.connection.execute(create_sql)
                    
                    self.progress.emit(85, "Getting output table statistics...")
                    
                    # Get row count
                    result = self.connection.execute(f"SELECT COUNT(*) FROM {output_table}").fetchone()
                    output_rows = result[0] if result else 0
                    
                    results['output_table'] = output_table
                    results['output_rows'] = output_rows
                    
                    self.progress.emit(95, f"SQL transformation complete - {output_rows:,} rows in {output_table}")
                    
                except Exception as e:
                    logger.error(f"Error executing SQL query: {e}")
                    
                    # Enhanced error handling with more specific messages
                    error_msg = str(e)
                    
                    # Provide more helpful error messages for common issues
                    if "no such table" in error_msg.lower():
                        error_msg += "\n\nTip: Check if the table name matches your configured CSV source names."
                    elif "no such column" in error_msg.lower():
                        error_msg += "\n\nTip: Check if the column name exists in your CSV data. Remember column names are now capitalized with underscores."
                    elif "syntax error" in error_msg.lower():
                        error_msg += "\n\nTip: Check your SQL syntax. Common issues include missing quotes, incorrect keywords, or typos."
                    elif "alter table" in error_msg.lower():
                        error_msg += "\n\nTip: ALTER TABLE operations have limitations. Consider using CREATE TABLE AS SELECT instead."
                    elif "permission" in error_msg.lower() or "access" in error_msg.lower():
                        error_msg += "\n\nTip: Check if you have the necessary permissions to perform this operation."
                    
                    self.error.emit(f"Error executing SQL query: {error_msg}")
                    return
            
            results['execution_time'] = time.time() - start_time
            self.progress.emit(100, "Processing completed successfully!")
            
            # Clean up all temporary files
            self.cleanup_temp_files()
            
            self.finished.emit(True, "CSV automation completed successfully!", results)
            
        except Exception as e:
            logger.error(f"Unexpected error in CSV automation: {e}")
            # Clean up temporary files even on error
            self.cleanup_temp_files()
            self.error.emit(f"Unexpected error: {str(e)}")

    def create_unified_column_schema(self, files, file_type, source_config):
        """Scan all files to create a unified column schema with original order preserved"""
        # Keep track of column order - first file determines base order
        base_columns = []
        additional_columns = set()
        first_file_processed = False
        
        for i, file_path in enumerate(files):
            try:
                current_file_columns = []
                
                if file_type == 'excel':
                    # For Excel files, get columns from first sheet or specified sheet
                    sheet_selection = source_config.get('sheet_selection', 'All sheets (combined)')
                    sheet_name = source_config.get('sheet_name', '')
                    
                    if sheet_selection == "First sheet only":
                        # Read just the first few rows to get column names
                        df_sample = read_excel_optimized(file_path, sheet_name)
                        if df_sample is not None and not df_sample.empty:
                            current_file_columns = df_sample.columns.tolist()
                    else:
                        # For multiple sheets, we need to check all sheets
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(file_path, read_only=True)
                            sheet_columns = set()
                            for sheet_name_iter in wb.sheetnames:
                                try:
                                    df_sample = read_excel_optimized(file_path, sheet_name_iter)
                                    if df_sample is not None and not df_sample.empty:
                                        sheet_columns.update(df_sample.columns.tolist())
                                except Exception as e:
                                    logger.warning(f"Could not read sheet {sheet_name_iter} from {file_path}: {e}")
                                    continue
                            wb.close()
                            current_file_columns = list(sheet_columns)
                        except Exception as e:
                            logger.warning(f"Could not scan Excel file {file_path}: {e}")
                            # Fallback to reading first sheet
                            try:
                                df_sample = read_excel_optimized(file_path, None)
                                if df_sample is not None and not df_sample.empty:
                                    current_file_columns = df_sample.columns.tolist()
                            except:
                                continue
                else:
                    # For CSV files, read just the header
                    try:
                        df_sample = pd.read_csv(file_path, nrows=0, encoding='utf-8')
                        current_file_columns = df_sample.columns.tolist()
                    except Exception as e:
                        logger.warning(f"Could not read CSV header from {file_path}: {e}")
                        # Try with different encoding
                        try:
                            df_sample = pd.read_csv(file_path, nrows=0, encoding='latin-1')
                            current_file_columns = df_sample.columns.tolist()
                        except:
                            continue
                
                # Process the columns from this file
                if current_file_columns:
                    if not first_file_processed:
                        # First file - establish base column order
                        base_columns = current_file_columns.copy()
                        first_file_processed = True
                        logger.info(f"Base column order established from first file: {base_columns}")
                    else:
                        # Subsequent files - find new columns
                        for col in current_file_columns:
                            if col not in base_columns:
                                additional_columns.add(col)
                
            except Exception as e:
                logger.warning(f"Error scanning file {file_path}: {e}")
                continue
        
        # Build the unified schema: base columns + additional columns + source file column
        unified_columns = base_columns.copy()
        
        # Add new columns at the end, sorted for consistency
        if additional_columns:
            new_columns = sorted(list(additional_columns))
            unified_columns.extend(new_columns)
            logger.info(f"Added {len(new_columns)} new columns at the end: {new_columns}")
        
        # Add source file column at the very end if not already present
        if '_source_file' not in unified_columns:
            unified_columns.append('_source_file')
        
        logger.info(f"Final unified schema ({len(unified_columns)} columns): {unified_columns}")
        return unified_columns

    def normalize_dataframe_to_schema(self, df, unified_columns):
        """Normalize a dataframe to match the unified column schema"""
        if df is None or df.empty:
            return df
        
        # Create a new dataframe with all unified columns
        normalized_df = pd.DataFrame(index=df.index)
        
        # Copy existing columns
        for col in df.columns:
            if col in unified_columns:
                normalized_df[col] = df[col]
        
        # Add missing columns with None values
        for col in unified_columns:
            if col not in normalized_df.columns:
                normalized_df[col] = None
        
        # Reorder columns to match unified schema
        normalized_df = normalized_df[unified_columns]
        
        return normalized_df


class CSVAutomationDialog(QDialog):
    """Dialog for configuring and executing CSV automation"""
    
    def __init__(self, parent=None, connection=None, connection_info=None):
        super().__init__(parent)
        self.connection = connection
        self.connection_info = connection_info
        self.csv_sources = []
        self.worker = None
        
        # Store executed automation info for returning to main app
        self.executed_sql_query = None
        self.executed_output_table = None
        self.automation_results = None
        
        # Create automations directory if it doesn't exist
        self.automations_dir = "automations"
        os.makedirs(self.automations_dir, exist_ok=True)
        
        self.setWindowTitle("CSV Automation")
        self.setModal(True)
        self.resize(1200, 800)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Header
        header_label = QLabel("CSV Automation - Process Multiple CSV Sources")
        header_label.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        header_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(header_label)
        
        # Instructions
        instructions = QLabel(
            "1. Add CSV folder sources and specify table names\n"
            "2. Optional: Write SQL query to combine/transform data\n"
            "3. Specify output table name and execute automation"
        )
        instructions.setStyleSheet("color: gray; margin: 10px; padding: 10px; background-color: #f0f0f0;")
        layout.addWidget(instructions)
        
        # Main content with tabs
        self.tab_widget = QTabWidget()
        
        # Tab 1: Sources Configuration  
        sources_tab = QWidget()
        sources_main_layout = QHBoxLayout(sources_tab)
        
        # Left side - CSV Sources
        sources_left = QWidget()
        sources_layout = QVBoxLayout(sources_left)
        
        # Add source button
        self.add_source_btn = QPushButton("Add CSV Source")
        self.add_source_btn.clicked.connect(self.add_csv_source)
        sources_layout.addWidget(self.add_source_btn)
        
        # Sources scroll area
        self.sources_scroll = QScrollArea()
        self.sources_widget = QWidget()
        self.sources_layout = QVBoxLayout(self.sources_widget)
        self.sources_scroll.setWidget(self.sources_widget)
        self.sources_scroll.setWidgetResizable(True)
        sources_layout.addWidget(self.sources_scroll)
        
        sources_main_layout.addWidget(sources_left, 2)  # 2/3 of the space
        
        # Right side - Saved Automations
        automations_right = QWidget()
        automations_layout = QVBoxLayout(automations_right)
        
        automations_title = QLabel("Saved Automations")
        automations_title.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        automations_layout.addWidget(automations_title)
        
        # Automations list
        self.automations_list = QListWidget()
        self.automations_list.setMaximumWidth(300)
        self.automations_list.itemDoubleClicked.connect(self.load_selected_automation)
        self.automations_list.setToolTip("Double-click to load automation for editing\nUse buttons below to Load, Run, or Delete")
        automations_layout.addWidget(self.automations_list)
        
        # Automation management buttons
        automation_buttons = QHBoxLayout()
        
        self.refresh_automations_btn = QPushButton("Refresh")
        self.refresh_automations_btn.clicked.connect(self.refresh_automations_list)
        self.refresh_automations_btn.setToolTip("Refresh the list of saved automations")
        automation_buttons.addWidget(self.refresh_automations_btn)
        
        self.load_selected_btn = QPushButton("Load")
        self.load_selected_btn.clicked.connect(self.load_selected_automation)
        self.load_selected_btn.setToolTip("Load selected automation for editing and review")
        automation_buttons.addWidget(self.load_selected_btn)
        
        self.run_selected_btn = QPushButton("Run")
        self.run_selected_btn.clicked.connect(self.run_selected_automation)
        self.run_selected_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }")
        self.run_selected_btn.setToolTip("Load and immediately execute the selected automation")
        automation_buttons.addWidget(self.run_selected_btn)
        
        self.delete_automation_btn = QPushButton("Delete")
        self.delete_automation_btn.clicked.connect(self.delete_selected_automation)
        self.delete_automation_btn.setToolTip("Delete the selected automation permanently")
        automation_buttons.addWidget(self.delete_automation_btn)
        
        automations_layout.addLayout(automation_buttons)
        
        # Automation details
        self.automation_details = QLabel("Select an automation to see details")
        self.automation_details.setWordWrap(True)
        self.automation_details.setStyleSheet("color: gray; font-style: italic; padding: 10px;")
        automations_layout.addWidget(self.automation_details)
        
        automations_layout.addStretch()
        
        sources_main_layout.addWidget(automations_right, 1)  # 1/3 of the space
        
        self.tab_widget.addTab(sources_tab, "1. CSV Sources")
        
        # Tab 2: SQL Query
        sql_tab = QWidget()
        sql_layout = QVBoxLayout(sql_tab)
        
        # SQL header
        sql_title = QLabel("SQL Query (Optional)")
        sql_title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        sql_layout.addWidget(sql_title)
        
        sql_description = QLabel(
            "Write an SQL query to combine or transform your CSV data.\n"
            "Reference tables by the names specified in the Sources tab."
        )
        sql_description.setStyleSheet("color: gray;")
        sql_layout.addWidget(sql_description)
        
        # Use enhanced SQL editor if available
        if SQL_EDITOR_AVAILABLE:
            self.sql_editor = SQLTextEdit()
            self.sql_editor.setFont(QFont("Consolas", 10))
            
            # Apply syntax highlighting
            self.sql_highlighter = SQLHighlighter(self.sql_editor.document())
            
            # Set up SQL completions if we have connection info
            if self.connection_info:
                self.update_sql_completions()
        else:
            self.sql_editor = QPlainTextEdit()
            self.sql_editor.setFont(QFont("Consolas", 10))
        
        self.sql_editor.setPlaceholderText(
            "Example:\n"
            "SELECT table1.*, table2.additional_column\n"
            "FROM table1 \n"
            "LEFT JOIN table2 ON table1.id = table2.id\n"
            "WHERE table1.date >= '2024-01-01'"
        )
        sql_layout.addWidget(self.sql_editor)
        
        self.tab_widget.addTab(sql_tab, "2. SQL Query")
        
        # Tab 3: Output
        output_tab = QWidget()
        output_layout = QVBoxLayout(output_tab)
        
        output_title = QLabel("Output Configuration")
        output_title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        output_layout.addWidget(output_title)
        
        # Output table name
        output_form = QFormLayout()
        self.output_table_line = QLineEdit()
        self.output_table_line.setPlaceholderText("Enter output table name (only if using SQL query)")
        output_form.addRow("Output Table Name:", self.output_table_line)
        
        output_layout.addLayout(output_form)
        output_layout.addStretch()
        
        self.tab_widget.addTab(output_tab, "3. Output")
        
        layout.addWidget(self.tab_widget)
        
        # Progress section
        progress_group = QGroupBox("Progress")
        progress_layout = QVBoxLayout(progress_group)
        
        self.progress_bar = QProgressBar()
        self.progress_label = QLabel("Ready to start automation...")
        
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.progress_label)
        
        layout.addWidget(progress_group)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        # Save automation button
        self.save_automation_btn = QPushButton("Save Automation")
        self.save_automation_btn.clicked.connect(self.save_automation)
        button_layout.addWidget(self.save_automation_btn)
        
        button_layout.addStretch()  # Add some space between groups
        
        self.execute_btn = QPushButton("Execute Automation")
        self.execute_btn.clicked.connect(self.execute_automation)
        button_layout.addWidget(self.execute_btn)
        
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.cancel_automation)
        self.cancel_btn.setEnabled(False)
        button_layout.addWidget(self.cancel_btn)
        
        self.close_btn = QPushButton("Close")
        self.close_btn.clicked.connect(self.close)
        button_layout.addWidget(self.close_btn)
        
        layout.addLayout(button_layout)
        
        # Initialize automations directory
        self.automations_dir = "automations"
        os.makedirs(self.automations_dir, exist_ok=True)
        
        # Load saved automations list
        self.refresh_automations_list()
        
        # Connect list selection to show details
        self.automations_list.itemSelectionChanged.connect(self.show_automation_details)
        
        # Add initial CSV source
        self.add_csv_source()
    
    def add_csv_source(self):
        """Add a new CSV source widget"""
        source_widget = CSVSourceWidget(self.sources_widget, len(self.csv_sources), self)
        self.csv_sources.append(source_widget)
        self.sources_layout.addWidget(source_widget)
        
        # Connect table name changes to update SQL completions
        source_widget.table_line.textChanged.connect(self.update_sql_completions)
    
    def remove_source(self, index):
        """Remove a CSV source widget"""
        if len(self.csv_sources) <= 1:
            QMessageBox.warning(self, "Warning", "At least one CSV source is required.")
            return
        
        widget = self.csv_sources.pop(index)
        widget.setParent(None)
        widget.deleteLater()
        
        # Update indices
        for i, source in enumerate(self.csv_sources):
            source.source_index = i
            source.title_label.setText(f"CSV Source {i + 1}")
        
        # Update SQL completions after removing source
        self.update_sql_completions()
    
    def execute_automation(self):
        """Execute the CSV automation process"""
        # Ensure we have sources first
        if not self.csv_sources:
            QMessageBox.warning(self, "Warning", "No CSV sources configured. Please add at least one source.")
            return
            
        # Validate sources
        valid_sources = [s for s in self.csv_sources if s.is_valid()]
        
        if not valid_sources:
            QMessageBox.warning(self, "Warning", "At least one valid CSV source is required. Please check that all sources have valid folder paths and table names.")
            return
        
        # Check for duplicate table names
        table_names = [s.get_config()['table_name'] for s in valid_sources]
        if len(table_names) != len(set(table_names)):
            QMessageBox.warning(self, "Warning", "Table names must be unique.")
            return
        
        # Get configurations
        sources_config = [s.get_config() for s in valid_sources]
        
        sql_query = self.sql_editor.toPlainText().strip()
        output_table = self.output_table_line.text().strip()
        
        # Validate SQL query and output table
        if sql_query and not output_table:
            QMessageBox.warning(self, "Warning", "Please specify an output table name when using SQL query.")
            return
        
        output_config = {'table_name': output_table}
        
        # Start worker thread
        self.worker = CSVAutomationWorker(
            self.connection, sources_config, output_config, sql_query, self.connection_info
        )
        
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(self.automation_finished)
        self.worker.error.connect(self.automation_error)
        
        # Update UI
        self.execute_btn.setEnabled(False)
        self.cancel_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        
        self.worker.start()
    
    def cancel_automation(self):
        """Cancel the running automation"""
        if self.worker:
            self.worker.cancel()
            self.worker.wait()
            self.worker.deleteLater()
            self.worker = None
        
        self.execute_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        self.progress_label.setText("Automation cancelled")
    
    def update_progress(self, value, message):
        """Update progress bar and message"""
        self.progress_bar.setValue(value)
        self.progress_label.setText(message)
    
    def automation_finished(self, success, message, results):
        """Handle automation completion"""
        self.execute_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        
        # Properly clean up the worker thread
        if self.worker:
            self.worker.wait()  # Wait for thread to finish
            self.worker.deleteLater()  # Schedule for deletion
            self.worker = None
        
        if success:
            # Store automation results for main app
            self.automation_results = results
            self.executed_sql_query = self.sql_editor.toPlainText().strip()
            self.executed_output_table = results.get('output_table')
            
            result_text = f"Automation completed successfully!\n\n"
            result_text += f"Sources processed: {results['sources_processed']}\n"
            result_text += f"Total rows loaded: {results['total_rows']:,}\n"
            result_text += f"Tables created: {', '.join(results['tables_created'])}\n"
            
            if results.get('output_table'):
                result_text += f"Output table: {results['output_table']}\n"
                result_text += f"Output rows: {results.get('output_rows', 0):,}\n"
            
            result_text += f"Execution time: {results['execution_time']:.2f} seconds"
            
            # Add information about SQL query integration
            if self.executed_sql_query and self.executed_output_table:
                result_text += f"\n\n🚀 The executed SQL query will be displayed in the main query editor!"
            
            QMessageBox.information(self, "Success", result_text)
            
            # Refresh schema if possible
            if hasattr(self.parent(), 'refresh_schema_browser'):
                self.parent().refresh_schema_browser()
        else:
            QMessageBox.critical(self, "Error", message)
    
    def automation_error(self, error_message):
        """Handle automation error"""
        self.execute_btn.setEnabled(True)
        self.cancel_btn.setEnabled(False)
        
        # Properly clean up the worker thread
        if self.worker:
            self.worker.wait()  # Wait for thread to finish
            self.worker.deleteLater()  # Schedule for deletion
            self.worker = None
        
        QMessageBox.critical(self, "Error", f"Automation failed:\n{error_message}")
        self.progress_label.setText("Automation failed")
    
    def update_sql_completions(self):
        """Update SQL completions with table names from sources"""
        if not SQL_EDITOR_AVAILABLE or not hasattr(self.sql_editor, 'completer'):
            return
            
        # Get table names from current sources
        table_names = []
        for source in self.csv_sources:
            if source.is_valid():
                config = source.get_config()
                table_names.append(config['table_name'])
        
        # Update completions
        if hasattr(self.sql_editor, 'update_completions'):
            self.sql_editor.update_completions(table_names=table_names)
    
    def get_automation_config(self):
        """Get the current automation configuration as a dictionary"""
        config = {
            'version': '1.0',
            'created': datetime.now().isoformat(),
            'description': 'CSV Automation Configuration',
            'sources': [],
            'sql_query': '',
            'output_table': ''
        }
        
        # Get sources configuration
        for source in self.csv_sources:
            if source.is_valid():
                source_config = source.get_config()
                config['sources'].append(source_config)
        
        # Get SQL query
        if hasattr(self.sql_editor, 'toPlainText'):
            config['sql_query'] = self.sql_editor.toPlainText().strip()
        else:
            config['sql_query'] = self.sql_editor.toPlainText().strip()
        
        # Get output table
        config['output_table'] = self.output_table_line.text().strip()
        
        return config
    
    def set_automation_config(self, config):
        """Load an automation configuration"""
        try:
            # Clear existing sources (but ensure we always have at least one)
            while len(self.csv_sources) > 1:
                self.remove_source(len(self.csv_sources) - 1)
            
            # Clear the first source
            if self.csv_sources:
                first_source = self.csv_sources[0]
                first_source.path_line.clear()
                first_source.table_line.clear()
                first_source.pattern_line.setText("*.csv")
                # Reset to folder mode by default
                first_source.mode_combo.setCurrentText("Folder (Multiple Files)")
                first_source.current_mode = 'folder'
                first_source.update_ui_for_mode()
            
            # Load sources
            sources = config.get('sources', [])
            if not sources:
                # Just clear everything but don't return - user might want to add sources manually
                # Make sure we have at least one empty source to work with
                if not self.csv_sources:
                    self.add_csv_source()
                return  # No sources to load, but we have a clean slate
                
            for i, source_config in enumerate(sources):
                # Use existing source or add new one
                if i < len(self.csv_sources):
                    source_widget = self.csv_sources[i]
                else:
                    self.add_csv_source()
                    source_widget = self.csv_sources[-1]
                
                # Set source configuration using the new method
                source_widget.set_config(source_config)
            
            # Load SQL query
            sql_query = config.get('sql_query', '')
            if hasattr(self.sql_editor, 'setPlainText'):
                self.sql_editor.setPlainText(sql_query)
            else:
                self.sql_editor.setPlainText(sql_query)
            
            # Load output table
            self.output_table_line.setText(config.get('output_table', ''))
            
            # Update SQL completions
            self.update_sql_completions()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load automation configuration:\n{str(e)}")
    

    
    def load_automation(self):
        """Load an automation configuration from a JSON file"""
        try:
            # Open file dialog
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Load CSV Automation",
                "",
                "JSON Files (*.json);;All Files (*)"
            )
            
            if not file_path:
                return
            
            with open(file_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # Validate configuration
            if not isinstance(config, dict):
                raise ValueError("Invalid automation file format")
            
            if 'sources' not in config:
                raise ValueError("No sources found in automation file")
            
            # Load configuration
            self.set_automation_config(config)
            
            QMessageBox.information(
                self, 
                "Success", 
                f"Automation configuration loaded from:\n{file_path}\n\n"
                f"Sources loaded: {len(config.get('sources', []))}\n"
                f"SQL Query: {'Yes' if config.get('sql_query') else 'No'}\n"
                f"Output Table: {config.get('output_table', 'Not specified')}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load automation:\n{str(e)}")
    
    def refresh_automations_list(self):
        """Refresh the list of saved automations"""
        try:
            self.automations_list.clear()
            
            # Find all JSON files in automations directory
            json_files = glob.glob(os.path.join(self.automations_dir, "*.json"))
            
            if not json_files:
                self.automations_list.addItem("No saved automations")
                return
            
            # Sort by modification time (newest first)
            json_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            
            for json_file in json_files:
                try:
                    # Try to load and validate the file
                    with open(json_file, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    
                    # Create display name
                    filename = os.path.basename(json_file)
                    display_name = filename.replace('.json', '')
                    
                    # Add creation date if available
                    if 'created' in config:
                        try:
                            created_date = datetime.fromisoformat(config['created'].replace('Z', '+00:00'))
                            display_name += f" ({created_date.strftime('%Y-%m-%d %H:%M')})"
                        except:
                            pass
                    
                    # Add to list with full path as data
                    item = QListWidgetItem(display_name)
                    item.setData(Qt.ItemDataRole.UserRole, json_file)  # Store full path
                    self.automations_list.addItem(item)
                    
                except Exception as e:
                    # Skip invalid files
                    logger.error(f"Error reading automation file {json_file}: {e}")
                    continue
        
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to refresh automations list:\n{str(e)}")
    
    def show_automation_details(self):
        """Show details of the selected automation"""
        try:
            current_item = self.automations_list.currentItem()
            if not current_item or current_item.text() == "No saved automations":
                self.automation_details.setText("Select an automation to see details")
                return
            
            file_path = current_item.data(Qt.ItemDataRole.UserRole)
            if not file_path:
                return
            
            # Load automation config
            with open(file_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # Build details text
            details = []
            details.append(f"<b>File:</b> {os.path.basename(file_path)}")
            
            if 'created' in config:
                details.append(f"<b>Created:</b> {config['created']}")
            
            sources_count = len(config.get('sources', []))
            details.append(f"<b>Sources:</b> {sources_count}")
            
            for i, source in enumerate(config.get('sources', []), 1):
                table_name = source.get('table_name', 'Unknown')
                
                # Handle both file and folder modes
                if source.get('mode') == 'file':
                    file_path = source.get('file_path', 'Unknown')
                    if file_path != 'Unknown':
                        source_display = os.path.basename(file_path)
                    else:
                        source_display = 'Unknown file'
                else:
                    folder_path = source.get('folder_path', 'Unknown')
                    if folder_path != 'Unknown':
                        source_display = os.path.basename(folder_path)
                    else:
                        source_display = 'Unknown folder'
                
                details.append(f"  {i}. {table_name} ← {source_display}")
            
            if config.get('sql_query'):
                query_preview = config['sql_query'][:100]
                if len(config['sql_query']) > 100:
                    query_preview += "..."
                details.append(f"<b>SQL Query:</b> {query_preview}")
            else:
                details.append(f"<b>SQL Query:</b> None")
            
            if config.get('output_table'):
                details.append(f"<b>Output Table:</b> {config['output_table']}")
            
            self.automation_details.setText("<br/>".join(details))
            
        except Exception as e:
            self.automation_details.setText(f"Error reading automation: {str(e)}")
    
    def load_selected_automation(self):
        """Load the selected automation from the list"""
        try:
            current_item = self.automations_list.currentItem()
            if not current_item or current_item.text() == "No saved automations":
                QMessageBox.warning(self, "Warning", "Please select an automation to load.")
                return
            
            file_path = current_item.data(Qt.ItemDataRole.UserRole)
            if not file_path:
                return
            
            # Load and apply the configuration
            with open(file_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            self.set_automation_config(config)
            
            QMessageBox.information(
                self,
                "Success",
                f"Automation loaded successfully!\n\n"
                f"Sources: {len(config.get('sources', []))}\n"
                f"SQL Query: {'Yes' if config.get('sql_query') else 'No'}\n"
                f"Output Table: {config.get('output_table', 'Not specified')}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load selected automation:\n{str(e)}")
    
    def run_selected_automation(self):
        """Load and immediately run the selected automation"""
        try:
            current_item = self.automations_list.currentItem()
            if not current_item or current_item.text() == "No saved automations":
                QMessageBox.warning(self, "Warning", "Please select an automation to run.")
                return
            
            file_path = current_item.data(Qt.ItemDataRole.UserRole)
            if not file_path:
                return
            
            # Load configuration first
            with open(file_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # Show confirmation dialog with automation details
            sources_count = len(config.get('sources', []))
            has_sql = bool(config.get('sql_query', '').strip())
            output_table = config.get('output_table', '')
            
            confirmation_text = (
                f"Are you sure you want to run this automation?\n\n"
                f"File: {os.path.basename(file_path)}\n"
                f"Sources: {sources_count}\n"
                f"SQL Query: {'Yes' if has_sql else 'No'}\n"
                f"Output Table: {output_table if output_table else 'Not specified'}"
            )
            
            reply = QMessageBox.question(
                self,
                "Run Automation",
                confirmation_text,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # Load the configuration
                self.set_automation_config(config)
                
                # Switch to progress view
                self.tab_widget.setCurrentIndex(0)
                
                # Execute immediately
                self.execute_automation()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to run selected automation:\n{str(e)}")
    
    def delete_selected_automation(self):
        """Delete the selected automation file"""
        try:
            current_item = self.automations_list.currentItem()
            if not current_item or current_item.text() == "No saved automations":
                QMessageBox.warning(self, "Warning", "Please select an automation to delete.")
                return
            
            file_path = current_item.data(Qt.ItemDataRole.UserRole)
            if not file_path:
                return
            
            filename = os.path.basename(file_path)
            
            # Confirm deletion
            reply = QMessageBox.question(
                self,
                "Confirm Deletion",
                f"Are you sure you want to delete this automation?\n\n{filename}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                os.remove(file_path)
                self.refresh_automations_list()
                self.automation_details.setText("Automation deleted successfully")
                
                QMessageBox.information(self, "Success", f"Automation '{filename}' deleted successfully.")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to delete automation:\n{str(e)}")
    
    def save_automation(self):
        """Save the current automation configuration automatically to the automations directory"""
        try:
            config = self.get_automation_config()
            
            if not config['sources']:
                QMessageBox.warning(self, "Warning", "No valid sources configured to save.")
                return
            
            # Ask for automation name
            default_name = f"automation_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            automation_name, ok = QInputDialog.getText(
                self,
                "Save Automation",
                "Enter a name for this automation:",
                QLineEdit.EchoMode.Normal,
                default_name
            )
            
            if not ok or not automation_name.strip():
                return  # User cancelled or empty name
            
            # Clean the name for filename
            automation_name = automation_name.strip()
            # Remove invalid filename characters
            import re
            automation_name = re.sub(r'[<>:"/\\|?*]', '_', automation_name)
            
            # Ensure .json extension
            if not automation_name.endswith('.json'):
                automation_name += '.json'
            
            # Save to automations directory
            file_path = os.path.join(self.automations_dir, automation_name)
            
            # Check if file exists
            if os.path.exists(file_path):
                reply = QMessageBox.question(
                    self,
                    "File Exists",
                    f"An automation named '{automation_name}' already exists.\n\nDo you want to overwrite it?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                if reply == QMessageBox.StandardButton.No:
                    return
            
            # Save the automation
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            
            # Refresh the automations list
            self.refresh_automations_list()
            
            # Select the newly saved automation in the list
            self.select_automation_by_name(automation_name)
            
            QMessageBox.information(
                self, 
                "Success", 
                f"Automation '{automation_name}' saved successfully!\n\nIt's now available in the Saved Automations panel."
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save automation:\n{str(e)}")
    
    def select_automation_by_name(self, filename):
        """Select an automation in the list by filename"""
        try:
            for i in range(self.automations_list.count()):
                item = self.automations_list.item(i)
                if item and filename in item.text():
                    self.automations_list.setCurrentItem(item)
                    break
        except:
            pass  # If selection fails, it's not critical
    
    def get_executed_automation_info(self):
        """Get information about the executed automation for main app integration"""
        return {
            'sql_query': self.executed_sql_query,
            'output_table': self.executed_output_table,
            'results': self.automation_results,
            'has_sql_query': bool(self.executed_sql_query and self.executed_output_table)
        }


def show_csv_automation_dialog(parent=None, connection=None, connection_info=None):
    """Show the CSV automation dialog and return the dialog object for accessing results"""
    dialog = CSVAutomationDialog(parent, connection, connection_info)
    result = dialog.exec()
    # Return both the result and the dialog object so caller can access executed query info
    return result, dialog


if __name__ == "__main__":
    # Test the dialog
    from PyQt6.QtWidgets import QApplication
    import sys
    
    app = QApplication(sys.argv)
    
    # Create a test DuckDB connection
    conn = duckdb.connect(':memory:')
    conn_info = {'type': 'duckdb', 'database': ':memory:'}
    
    dialog = CSVAutomationDialog(None, conn, conn_info)
    dialog.show()
    
    sys.exit(app.exec())